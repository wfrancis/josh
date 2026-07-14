"""
FastAPI application for the Standard Interiors Bid Tool.
"""

import csv
import hashlib
import io
import json
import math
import os
import shutil
import sqlite3
import tempfile
from datetime import datetime, timezone
from functools import lru_cache
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Body
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from models import (
    init_db, save_job, load_job, list_jobs, delete_job,
    save_materials, save_sundries, save_labor, save_bundles,
    save_quotes, delete_quotes, update_quote, get_quote_job_id, search_all,
    get_settings, save_settings,
    save_labor_catalog_entries, get_labor_catalog_entries,
    update_labor_catalog_entry, delete_labor_catalog_entry,
    clear_labor_catalog, clear_price_list,
    save_price_list_entries, add_price_list_entry, update_price_list_entry,
    delete_price_list_entry, get_price_list_entries,
    get_company_rate, save_company_rate, get_all_company_rates,
    get_or_create_vendor, save_vendor_prices_from_quotes,
    list_vendors, get_vendor, update_vendor, search_vendor_prices,
    create_vendor, delete_vendor,
    get_price_history, import_vendor_prices_csv,
    create_notification, get_notifications, mark_notification_read,
    is_file_imported, record_imported_file, list_imported_files, record_job_artifact, list_job_artifacts,
    log_activity, get_activity, add_comment, get_comments,
    create_quote_request, list_quote_requests, update_quote_request, delete_quote_request,
    _normalize_product, _get_conn,
    import_price_book, search_price_book, match_price_book, get_price_book_summary,
    list_rules, get_rule, create_rule, update_rule, delete_rule,
    archive_rule, list_rule_versions,
    list_ruleset_versions, get_ruleset_version, rollback_ruleset_version,
    get_active_rules, seed_rules_registry_defaults,
    create_calculation_run, save_calculation_traces, complete_calculation_run,
    list_calculation_runs, get_latest_completed_calculation_run, get_calculation_traces,
    upsert_golden_job, get_golden_job_for_source, get_golden_replay,
    save_golden_replay, list_golden_replays_for_job, list_golden_replays_for_version,
    get_latest_golden_replay_for_version,
)
from rfms_parser import parse_rfms, ai_merge_materials
from quote_parser import MAX_QUOTE_FILE_BYTES, parse_quote_file, set_openai_config
from dropbox_scanner import match_folder
from sundry_calc import calculate_sundries_for_materials
from labor_calc import calculate_labor_for_materials, load_labor_catalog, load_labor_catalog_from_pdf, get_labor_catalog
from bid_assembler import assemble_bid
from pdf_generator import generate_bid_pdf, generate_proposal_pdf
from proposal_bundler import generate_proposal_data
from proposal_totals import effective_bundle_total, normalize_proposal_totals
from reproducibility import DEFAULT_TOLERANCE, apply_accepted_numeric_edits, make_golden_snapshot, replay_golden_job
from config import WASTE_FACTORS, SUNDRY_RULES, FREIGHT_RATES, LABOR_QTY_RULES, EXCLUSIONS_TEMPLATE, STAIR_SUNDRY_KITS
from email_agent import compose_quote_request, send_email, generate_quote_request_text
from ai_client import chat_complete, get_provider_info
from inbox_monitor import InboxMonitor
from audit_engine import AuditTraceBuilder
from build_info import build_manifest_for_snapshot, get_build_info
from readiness import evaluate_job_readiness, is_valid_material_classification, proposal_math_errors

# ── App Setup ─────────────────────────────────────────────────────────────────
app = FastAPI(title="SI Bid Tool", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def _match_price_book(material: dict) -> dict | None:
    """Match a material to a price_book_items entry (e.g. Schluter catalog).

    Looks for Schluter product references in the description like:
      "Schluter - Schiene #AE-100" -> product_line=SCHIENE, item_no=AE 100
      "Schluter - Reno-TK"         -> product_line=RENO-TK, default AE finish
      "Schluter - Jolly"           -> product_line=JOLLY, default AE finish
    Returns dict with unit_price (per LF), vendor, price_source.
    """
    desc = (material.get("description") or "")
    if "schluter" not in desc.lower():
        return None

    import re
    from models import _get_conn
    conn = _get_conn()
    try:
        # Extract product line and optional item number
        # "Schluter - Schiene #AE-100" -> line="Schiene", item="AE-100"
        # "Schluter - Reno-TK - Transition" -> line="Reno-TK", item=None
        m = re.search(r'schluter\s*[-–—]\s*([\w][\w-]*?)(?:\s*#([\w-]+)|\s)', desc, re.IGNORECASE)
        if not m:
            return None

        product_line = m.group(1).upper()  # SCHIENE, RENO-TK, JOLLY, FINEC, etc.
        raw_item = (m.group(2) or "").upper()  # AE-100, empty, etc.

        row = None
        if raw_item:
            # Normalize: "AE-100" -> "AE 100" (DB uses space separator)
            item_normalized = re.sub(r'[-]', ' ', raw_item).strip()
            row = conn.execute(
                "SELECT net_price, length FROM price_book_items WHERE UPPER(product_line)=? AND UPPER(REPLACE(TRIM(item_no), '-', ' '))=?",
                (product_line, item_normalized)
            ).fetchone()

        if not row:
            # Fallback: match product_line with anodized aluminum finish (default for SI)
            row = conn.execute(
                "SELECT net_price, length FROM price_book_items WHERE UPPER(product_line)=? AND material_finish LIKE '%anodized%' ORDER BY net_price ASC LIMIT 1",
                (product_line,)
            ).fetchone()

        if not row:
            return None

        net_price = row[0]  # price per stick
        length_str = row[1] or ""
        # Parse stick length: "2.5 m - 8' 2-1/2" length" -> ~8.2 LF
        stick_lf = 8.208  # default Schluter stick length (2.5m)
        lf_match = re.search(r"(\d+)'\s*(\d+)?", length_str)
        if lf_match:
            feet = int(lf_match.group(1))
            inches = int(lf_match.group(2)) if lf_match.group(2) else 0
            stick_lf = feet + inches / 12.0

        price_per_lf = round(net_price / stick_lf, 4)
        return {
            "unit_price": price_per_lf,
            "stick_price": net_price,
            "stick_lf": stick_lf,
            "vendor": "Schluter",
            "price_source": "price_book",
        }
    finally:
        conn.close()


def _match_price_list(material: dict, price_list: list[dict]) -> dict | None:
    """Match a material to a price list entry by item_code, description, or material_type."""
    item_code = (material.get("item_code") or "").strip().lower()
    description = (material.get("description") or "").strip().lower()
    material_type = (material.get("material_type") or "").strip().lower()

    # Pass 1: exact item_code match against product_name
    if item_code and len(item_code) >= 3:
        for entry in price_list:
            entry_name = (entry.get("product_name") or "").strip().lower()
            if item_code == entry_name or item_code in entry_name or entry_name in item_code:
                return entry

    # Pass 2: description substring match
    if description and len(description) >= 5:
        for entry in price_list:
            entry_name = (entry.get("product_name") or "").strip().lower()
            if entry_name and len(entry_name) >= 5:
                if entry_name in description or description in entry_name:
                    return entry

    # Pass 3: material_type match (weakest, gives generic pricing)
    if material_type:
        for entry in price_list:
            if (entry.get("material_type") or "").strip().lower() == material_type:
                return entry

    return None


_artifact_default = "/data/artifacts" if os.path.isdir("/data") else os.path.join(os.path.dirname(__file__), "artifacts")
ARTIFACT_ROOT = os.environ.get("ARTIFACT_ROOT", _artifact_default)
MAX_RFMS_FILE_BYTES = 50 * 1024 * 1024
UPLOAD_DIR = os.path.join(ARTIFACT_ROOT, "shared", "uploads")
PDF_DIR = os.path.join(ARTIFACT_ROOT, "shared", "pdfs")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)


def _safe_artifact_name(name: str) -> str:
    """Keep uploaded artifact names inside the intended job directory."""
    import re as _re
    clean = os.path.basename(name or "artifact")
    return _re.sub(r"[^A-Za-z0-9._-]+", "_", clean) or "artifact"


def _job_artifact_dir(job_id: int, kind: str) -> str:
    path = os.path.join(ARTIFACT_ROOT, str(job_id), kind)
    os.makedirs(path, exist_ok=True)
    return path


def _job_upload_path(job_id: int, filename: str, prefix: str) -> str:
    return os.path.join(_job_artifact_dir(job_id, "uploads"), f"{prefix}_{_safe_artifact_name(filename)}")


def _job_pdf_path(job_id: int, kind: str) -> str:
    return os.path.join(_job_artifact_dir(job_id, "pdfs"), f"{kind}_{job_id}.pdf")


@lru_cache(maxsize=512)
def _file_hash_for_stat(
    path: str,
    size: int,
    modified_ns: int,
    changed_ns: int,
) -> str | None:
    try:
        digest = hashlib.sha256()
        with open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError:
        return None


def _file_hash(path: str) -> str | None:
    try:
        stat = os.stat(path)
    except OSError:
        return None
    return _file_hash_for_stat(
        os.path.realpath(path),
        stat.st_size,
        stat.st_mtime_ns,
        stat.st_ctime_ns,
    )


def _record_artifact(job_id: int, path: str, artifact_kind: str) -> None:
    """Persist a receipt for a durable file after it has been written."""
    digest = _file_hash(path)
    if not digest:
        return
    try:
        size = os.path.getsize(path)
    except OSError:
        size = 0
    record_job_artifact(
        job_id,
        artifact_kind,
        os.path.relpath(path, ARTIFACT_ROOT),
        digest,
        size,
    )


def _require_artifact_receipt(job_id: int, path: str, artifact_kind: str) -> None:
    """Reject a download when its durable receipt is missing or no longer matches."""
    expected_path = os.path.relpath(path, ARTIFACT_ROOT)
    receipt = next(
        (
            item for item in list_job_artifacts(job_id)
            if item.get("artifact_kind") == artifact_kind
            and item.get("artifact_path") == expected_path
        ),
        None,
    )
    if not receipt:
        raise HTTPException(
            status_code=409,
            detail="This PDF has no durable storage receipt. Generate it again before downloading.",
        )
    if _file_hash(path) != receipt.get("file_hash"):
        raise HTTPException(
            status_code=409,
            detail="This PDF no longer matches its recorded hash. Generate it again before downloading.",
        )


def _persist_automated_quote_evidence(
    job_id: int,
    temp_files: list[str],
    filenames: list[str],
    parsed_indexes: set[int],
    source: str,
) -> None:
    """Copy temporary inbox artifacts to durable job storage before cleanup."""
    for index, temp_path in enumerate(temp_files):
        digest = _file_hash(temp_path)
        if not digest:
            continue
        original_name = filenames[index] if index < len(filenames) else os.path.basename(temp_path)
        durable_name = f"{digest[:12]}_{original_name or 'vendor_quote'}"
        durable_path = _job_upload_path(job_id, durable_name, "quote")
        if not os.path.exists(durable_path):
            shutil.copy2(temp_path, durable_path)
        _record_artifact(job_id, durable_path, "vendor_quote")
        if index in parsed_indexes and not is_file_imported(job_id, digest):
            record_imported_file(
                job_id,
                original_name or "vendor_quote",
                digest,
                os.path.getsize(durable_path),
                source=source,
                artifact_path=os.path.relpath(durable_path, ARTIFACT_ROOT),
                artifact_kind="vendor_quote",
            )


def _job_artifact_manifest(job_id: int) -> list[dict]:
    root = os.path.join(ARTIFACT_ROOT, str(job_id))
    manifest = []
    if not os.path.isdir(root):
        return manifest
    for current_root, _, names in os.walk(root):
        for name in sorted(names):
            path = os.path.join(current_root, name)
            try:
                size = os.path.getsize(path)
            except OSError:
                continue
            manifest.append({
                "path": os.path.relpath(path, root),
                "size": size,
                "sha256": _file_hash(path),
            })
    known = {item.get("path") for item in manifest}
    for receipt in list_job_artifacts(job_id):
        rel_from_artifact_root = receipt.get("artifact_path") or ""
        path = os.path.join(ARTIFACT_ROOT, rel_from_artifact_root)
        rel_from_job = os.path.relpath(path, root) if path.startswith(root + os.sep) else None
        if rel_from_job and rel_from_job not in known and os.path.isfile(path):
            manifest.append({
                "path": rel_from_job,
                "size": receipt.get("file_size") or 0,
                "sha256": receipt.get("file_hash"),
            })
    return manifest


def _artifact_readiness(
    job_id: int,
    proposal_pdf_path: str,
    materials: list[dict] | None = None,
) -> tuple[str, str, list[str]]:
    """Verify durable source and PDF receipts without changing the job."""
    failures = []
    warnings = []
    artifact_root = os.path.realpath(ARTIFACT_ROOT)

    def checked_path(relative_path: str) -> str | None:
        if not relative_path:
            return None
        candidate = os.path.realpath(os.path.join(artifact_root, relative_path))
        try:
            if os.path.commonpath([artifact_root, candidate]) != artifact_root:
                return None
        except ValueError:
            return None
        return candidate

    def add_failure(message: str) -> None:
        if message not in failures:
            failures.append(message)

    imported_files = list_imported_files(job_id)
    for imported in imported_files:
        label = imported.get("file_name") or "source file"
        relative_path = imported.get("artifact_path") or ""
        if not relative_path:
            warnings.append(f"{label} is a legacy import without a durable path")
            continue
        path = checked_path(relative_path)
        if not path or not os.path.isfile(path):
            add_failure(f"{label} is missing from durable storage")
            continue
        expected_hash = imported.get("file_hash")
        if expected_hash and _file_hash(path) != expected_hash:
            add_failure(f"{label} no longer matches its recorded hash")

    artifact_receipts = list_job_artifacts(job_id)
    for receipt in artifact_receipts:
        relative_path = receipt.get("artifact_path") or ""
        label = receipt.get("artifact_kind") or relative_path or "artifact"
        path = checked_path(relative_path)
        if not path or not os.path.isfile(path):
            add_failure(f"Recorded {label} artifact is missing from durable storage")
            continue
        expected_hash = receipt.get("file_hash")
        if expected_hash and _file_hash(path) != expected_hash:
            add_failure(f"Recorded {label} artifact no longer matches its hash")

    has_rfms_evidence = any(
        item.get("artifact_kind") == "rfms"
        for item in (*imported_files, *artifact_receipts)
    )
    if materials and not has_rfms_evidence:
        warnings.append("No durable RFMS source workbook is recorded for this structured job")

    vendor_priced = [
        material.get("item_code") or material.get("description") or "material"
        for material in (materials or [])
        if isinstance(material, dict)
        and str(material.get("price_source") or "").strip().lower() == "vendor_quote"
    ]
    has_vendor_evidence = any(
        item.get("artifact_kind") == "vendor_quote"
        for item in (*imported_files, *artifact_receipts)
    )
    if vendor_priced and not has_vendor_evidence:
        add_failure(
            f"{len(vendor_priced)} material(s) use vendor-quote pricing without a durable quote file"
        )

    expected_pdf_rel = os.path.relpath(proposal_pdf_path, ARTIFACT_ROOT)
    pdf_receipt = next(
        (
            item for item in artifact_receipts
            if item.get("artifact_kind") == "proposal_pdf"
            and item.get("artifact_path") == expected_pdf_rel
        ),
        None,
    )
    if not os.path.isfile(proposal_pdf_path):
        add_failure("The proposal PDF is missing from durable storage")
    elif not pdf_receipt:
        add_failure("The proposal PDF does not have a durable artifact receipt")
    elif _file_hash(proposal_pdf_path) != pdf_receipt.get("file_hash"):
        add_failure("The proposal PDF no longer matches its recorded hash")

    if failures:
        return "fail", "Durable artifact verification failed.", failures
    if warnings:
        return "warn", "Current artifacts are intact, but some historical imports predate durable storage.", warnings
    return "pass", "Recorded source files, vendor evidence, and the proposal PDF pass their hash checks.", []


def _find_incoming_quote_job(job_reference, subject: str) -> int | None:
    reference = str(job_reference or "").strip()
    if reference:
        direct = load_job(reference)
        if direct:
            return direct["id"]
    search_term = reference or str(subject or "")[:80]
    if not search_term:
        return None
    results = search_all(search_term)
    jobs = results.get("jobs") or []
    normalized = "".join(character.lower() for character in search_term if character.isalnum())
    exact = [
        item for item in jobs
        if normalized and normalized in {
            "".join(character.lower() for character in str(item.get("project_name") or "") if character.isalnum()),
            "".join(character.lower() for character in str(item.get("slug") or "") if character.isalnum()),
        }
    ]
    if len(exact) == 1:
        return exact[0]["id"]
    if len(jobs) == 1 and len(normalized) >= 6:
        return jobs[0]["id"]
    if jobs:
        print(f"[InboxMonitor] Ambiguous job match for '{search_term}'; leaving the email unread")
    return None


def _ingest_automated_quote(
    *,
    job_reference=None,
    temp_files=None,
    vendor_email=None,
    filenames=None,
    subject="",
    source: str,
    simulated: bool = False,
) -> bool:
    """Persist, parse, and import one automated vendor response idempotently."""
    temp_files = temp_files or []
    filenames = filenames or []
    job_id = _find_incoming_quote_job(job_reference, subject)
    log_prefix = "SimWatcher" if simulated else "InboxMonitor"
    if not job_id:
        print(f"[{log_prefix}] No matching job for: {job_reference or subject}")
        return False

    # Preserve the original evidence before any AI call. A failed parse stays
    # unread/retryable, while the estimator still has the source artifact.
    _persist_automated_quote_evidence(
        job_id,
        temp_files,
        filenames,
        set(),
        source=source,
    )

    all_products = []
    parsed_indexes = set()
    already_imported = set()
    for index, file_path in enumerate(temp_files):
        digest = _file_hash(file_path)
        if digest and is_file_imported(job_id, digest):
            already_imported.add(index)
            continue
        try:
            products = parse_quote_file(file_path)
            if products:
                filename = filenames[index] if index < len(filenames) else os.path.basename(file_path)
                for product in products:
                    product["file_name"] = filename
                    product["_source_hash"] = digest
                all_products.extend(products)
                parsed_indexes.add(index)
        except Exception as exc:
            print(f"[{log_prefix}] Error parsing {file_path}: {exc}")

    if not all_products:
        if temp_files and len(already_imported) == len(temp_files):
            return True
        print(f"[{log_prefix}] No products extracted from: {str(subject)[:60]}")
        return False

    save_quotes(job_id, all_products)
    matched = _auto_match_quotes(job_id, all_products)
    save_vendor_prices_from_quotes(job_id, all_products)
    _link_upload_to_requests(job_id, all_products)
    _persist_automated_quote_evidence(
        job_id,
        temp_files,
        filenames,
        parsed_indexes,
        source=source,
    )

    vendor_name = all_products[0].get("vendor", vendor_email or "Unknown")
    file_names_str = ", ".join(filenames[:3]) if filenames else "email"
    sim_prefix = "[SIM] " if simulated else ""
    create_notification(
        job_id,
        "quote_received",
        f"{sim_prefix}Quote received from {vendor_name} - {len(all_products)} products parsed, "
        f"{matched} auto-matched ({file_names_str})",
    )
    log_activity(
        job_id,
        "agent_quote_imported",
        f"{sim_prefix}Auto-imported quote from {vendor_name} via {'test mode' if simulated else 'email monitor'}",
        {"vendor": vendor_name, "products": len(all_products), "matched": matched, "sim": simulated},
    )
    print(f"[{log_prefix}] Imported {len(all_products)} products for job #{job_id} from {vendor_name}")
    return True


_inbox_monitor: InboxMonitor | None = None
_sim_watcher = None  # SimFolderWatcher instance (lazy import to avoid circular)


def _start_sim_watcher():
    """Start the SimFolderWatcher if vendor quote test mode is enabled."""
    global _sim_watcher
    settings = get_settings()
    test_mode = str(settings.get("vendor_quote_test_mode", "false")).lower() == "true"

    # Stop existing watcher if running
    if _sim_watcher and _sim_watcher.is_running:
        _sim_watcher.stop()
        _sim_watcher = None

    if not test_mode:
        return

    from sim_email import SimFolderWatcher

    def on_sim_quote_received(*, job_reference=None, temp_files=None, vendor_email=None,
                              filenames=None, subject=""):
        """Callback for sim watcher — same pipeline as InboxMonitor."""
        return _ingest_automated_quote(
            job_reference=job_reference,
            temp_files=temp_files,
            vendor_email=vendor_email,
            filenames=filenames,
            subject=subject,
            source="simulator",
            simulated=True,
        )

    _sim_watcher = SimFolderWatcher(on_quote_received=on_sim_quote_received)
    _sim_watcher.start()


def _start_inbox_monitor():
    """Start the inbox monitor if email automation is enabled."""
    global _inbox_monitor
    import json as _json
    settings = get_settings()
    # Don't start real inbox monitor when in test mode
    if str(settings.get("vendor_quote_test_mode", "false")).lower() == "true":
        if _inbox_monitor and _inbox_monitor.is_running:
            _inbox_monitor.stop()
        _inbox_monitor = None
        return
    if settings.get("email_automation_enabled") != "true":
        if _inbox_monitor and _inbox_monitor.is_running:
            _inbox_monitor.stop()
        _inbox_monitor = None
        return
    try:
        config = _json.loads(settings.get("email_config", "{}"))
    except Exception:
        if _inbox_monitor and _inbox_monitor.is_running:
            _inbox_monitor.stop()
        _inbox_monitor = None
        return
    imap_host = config.get("imap_host")
    imap_port = int(config.get("imap_port", 993))
    email_addr = config.get("email_address")
    email_pass = config.get("email_password")
    if not all([imap_host, email_addr, email_pass]):
        if _inbox_monitor and _inbox_monitor.is_running:
            _inbox_monitor.stop()
        _inbox_monitor = None
        return

    def on_quote_received(*, job_reference=None, temp_files=None, vendor_email=None,
                          filenames=None, subject=""):
        """Callback when inbox monitor detects a vendor response.
        Signature matches InboxMonitor._process_email kwargs."""
        return _ingest_automated_quote(
            job_reference=job_reference,
            temp_files=temp_files,
            vendor_email=vendor_email,
            filenames=filenames,
            subject=subject,
            source="email_monitor",
        )

    if _inbox_monitor and _inbox_monitor.is_running:
        _inbox_monitor.stop()

    _inbox_monitor = InboxMonitor(
        imap_config={
            "host": imap_host,
            "port": imap_port,
            "username": email_addr,
            "password": email_pass,
            "use_ssl": imap_port == 993,
        },
        on_quote_received=on_quote_received,
        poll_interval=300,
    )
    _inbox_monitor.start()
    print(f"[InboxMonitor] Started monitoring {email_addr}")


@app.on_event("startup")
def startup():
    init_db()
    _apply_openai_config()
    _seed_company_rates()
    _seed_rules_registry()
    _auto_import_price_books()
    _start_inbox_monitor()
    _start_sim_watcher()


def _auto_import_price_books():
    """Auto-import price books from JSON files on startup if DB is empty."""
    summary = get_price_book_summary()
    if any(s["vendor"] == "Schluter" for s in summary):
        return  # Already imported
    json_path = os.path.join(os.path.dirname(__file__), "schluter_prices.json")
    if os.path.exists(json_path):
        import json as _json
        with open(json_path) as f:
            items = _json.load(f)
        count = import_price_book("Schluter", items, discount_pct=0.55, category="transitions")
        print(f"Auto-imported Schluter price book: {count} items (45% of list)")


def _seed_rules_registry():
    """Seed hard estimating rules if they are not already in the registry."""
    result = seed_rules_registry_defaults()
    if result.get("inserted"):
        print(f"[seed] Seeded {result['inserted']} estimating rules")


# ── Pydantic Models ──────────────────────────────────────────────────────────

class JobCreate(BaseModel):
    project_name: str
    gc_name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip: Optional[str] = None
    tax_rate: float = 0.0
    gpm_pct: float = 0.0
    unit_count: int = 0
    tub_shower_count: int = 0
    salesperson: Optional[str] = None
    notes: Optional[str] = None
    architect: Optional[str] = None
    designer: Optional[str] = None
    textura_fee: int = 0


class MaterialUpdate(BaseModel):
    materials: list[dict]
    base_source_fingerprint: Optional[str] = None
    deletion_reasons: dict[str, str] = Field(default_factory=dict)


class NotesUpdate(BaseModel):
    notes: str = ""


class BulkDeleteRequest(BaseModel):
    job_ids: list[int]


class SettingsUpdate(BaseModel):
    openai_api_key: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    openai_model: Optional[str] = None
    multi_pass_count: Optional[int] = None
    email_automation_enabled: Optional[str] = None
    email_config: Optional[str] = None
    bid_folder_path: Optional[str] = None
    vendor_quote_test_mode: Optional[str] = None


class RuleCreate(BaseModel):
    rule_id: str
    name: str
    category: str = ""
    stage: str = ""
    status: str = "draft"
    priority: int = 0
    condition_json: Optional[dict] = None
    action_json: Optional[dict] = None
    source: str = ""
    description: str = ""
    effective_from: Optional[str] = None
    effective_to: Optional[str] = None
    version: int = 1
    implementation_ref: str = ""
    test_ref: str = ""
    notes: str = ""
    changed_by: Optional[str] = None
    change_note: Optional[str] = None


class RuleUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    stage: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[int] = None
    condition_json: Optional[dict] = None
    action_json: Optional[dict] = None
    source: Optional[str] = None
    description: Optional[str] = None
    effective_from: Optional[str] = None
    effective_to: Optional[str] = None
    version: Optional[int] = None
    implementation_ref: Optional[str] = None
    test_ref: Optional[str] = None
    notes: Optional[str] = None
    changed_by: Optional[str] = None
    change_note: Optional[str] = None


class RuleChangeMeta(BaseModel):
    changed_by: Optional[str] = None
    change_note: Optional[str] = None


class RuleDraftRequest(BaseModel):
    lesson_text: str
    changed_by: Optional[str] = "Josh"


class RulesetRollbackRequest(BaseModel):
    changed_by: Optional[str] = "Josh"
    change_note: Optional[str] = None


class GoldenBaselineRequest(BaseModel):
    jr_quote_id: Optional[str] = ""
    target_totals: Optional[dict] = None
    tolerance: Optional[dict] = None
    notes: Optional[str] = ""
    reviewer_name: Optional[str] = ""


class GoldenReplayRequest(BaseModel):
    mode: str = "baseline"


# ── Routes ────────────────────────────────────────────────────────────────────

def _rule_with_engine_contract(rule: dict, build: dict | None = None) -> dict:
    build = build or get_build_info()
    item = dict(rule or {})
    implementation_ref = str(item.get("implementation_ref") or item.get("source") or "").strip()
    implementation_path = implementation_ref.split("#", 1)[0].split(":", 1)[0]
    implementation_file = os.path.basename(implementation_path)
    implementation_hash = (build.get("engine_files") or {}).get(implementation_file)
    contribution_payload = {
        "rule_id": item.get("rule_id"),
        "rule_version": item.get("version"),
        "implementation_ref": implementation_ref,
        "test_ref": item.get("test_ref"),
        "implementation_hash": implementation_hash,
        "config_fingerprint": build.get("config_fingerprint"),
    }
    item["rule_version"] = item.get("version")
    item["engine_config_fingerprint_contribution"] = hashlib.sha256(
        json.dumps(contribution_payload, sort_keys=True, default=str, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    runtime_reference_verified = bool(implementation_ref and implementation_hash)
    item["calculation_contract"] = "implemented" if runtime_reference_verified else "metadata_only"
    item["runtime_implementation_referenced"] = runtime_reference_verified
    item["registry_effect"] = "metadata_only"
    return item


def _ruleset_with_engine_contract(ruleset: dict | None) -> dict | None:
    if not ruleset:
        return ruleset
    result = dict(ruleset)
    snapshot = dict(result.get("snapshot") or {})
    build = get_build_info()
    snapshot["rules"] = [_rule_with_engine_contract(rule, build) for rule in (snapshot.get("rules") or [])]
    snapshot["engine_fingerprint"] = build.get("engine_fingerprint")
    snapshot["config_fingerprint"] = build.get("config_fingerprint")
    result["snapshot"] = snapshot
    result["engine_fingerprint"] = build.get("engine_fingerprint")
    result["config_fingerprint"] = build.get("config_fingerprint")
    return result

@app.get("/api/system/build")
def api_system_build():
    """Return the deployed build and engine identity used for trust checks."""
    return get_build_info()


def _vendor_import_idempotency_status() -> dict:
    required_columns = {
        "job_quotes": "source_hash",
        "vendor_prices": "source_hash",
    }
    required_indexes = {
        "idx_job_quotes_source_product",
        "idx_vendor_prices_source_product",
    }
    missing = []
    conn = _get_conn()
    try:
        for table, column in required_columns.items():
            columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
            if column not in columns:
                missing.append(f"{table}.{column}")
        indexes = set()
        for table in required_columns:
            for row in conn.execute(f"PRAGMA index_list({table})").fetchall():
                if int(row["unique"] or 0):
                    indexes.add(row["name"])
        missing.extend(sorted(required_indexes - indexes))
    finally:
        conn.close()
    return {
        "verified": not missing,
        "mechanism": "source_hash_database_uniqueness",
        "missing": missing,
    }


@app.get("/api/system/vendor-ingestion")
def api_system_vendor_ingestion():
    """Return non-secret health evidence for vendor-pricing ingestion."""
    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    provider = get_provider_info(api_key)
    test_mode = str(settings.get("vendor_quote_test_mode", "false")).lower() == "true"
    automation_enabled = str(settings.get("email_automation_enabled", "false")).lower() == "true"
    monitor_running = bool(_inbox_monitor and _inbox_monitor.is_running)
    simulator_running = bool(_sim_watcher and _sim_watcher.is_running)
    idempotency = _vendor_import_idempotency_status()
    if not idempotency["verified"] or not provider.get("available"):
        status = "blocked"
    elif test_mode and not simulator_running:
        status = "warning"
    elif automation_enabled and not test_mode and not monitor_running:
        status = "warning"
    else:
        status = "ready"
    try:
        multi_pass_count = max(1, min(5, int(settings.get("multi_pass_count", "2"))))
    except (TypeError, ValueError):
        multi_pass_count = 2
    return {
        "status": status,
        "ai_parser": {
            "available": bool(provider.get("available")),
            "provider": provider.get("provider"),
            "model": settings.get("openai_model", "gpt-5-mini"),
            "multi_pass_count": multi_pass_count,
        },
        "email_monitor": {
            "enabled": automation_enabled,
            "running": monitor_running,
            "simulator_running": simulator_running,
            "test_mode": test_mode,
            "retry_failed_unread": True,
            "idempotency": idempotency["mechanism"],
            "idempotency_verified": idempotency["verified"],
            "idempotency_missing": idempotency["missing"],
            "job_match": "stable_subject_tag_then_unambiguous_project_match",
        },
        "dropbox_import": {
            "mode": "browser_folder_picker",
            "requires_user_action": True,
            "supported_extensions": [".eml", ".msg", ".pdf", ".txt", ".csv", ".xlsx"],
            "max_file_mb": MAX_QUOTE_FILE_BYTES // (1024 * 1024),
        },
        "durable_artifact_root": ARTIFACT_ROOT,
    }

@app.get("/api/rules")
def api_list_rules(category: str = None, stage: str = None, status: str = None):
    """List hard estimating rules with optional category/stage/status filters."""
    rules = list_rules(category=category, stage=stage, status=status)
    build = get_build_info()
    return {"rules": [_rule_with_engine_contract(rule, build) for rule in rules], "count": len(rules)}


@app.get("/api/rules/active")
def api_get_active_rules(stage: str = None, category: str = None, as_of: str = None):
    """List currently active estimating rules for audit/calculation consumers."""
    rules = get_active_rules(stage=stage, category=category, as_of=as_of)
    build = get_build_info()
    return {"rules": [_rule_with_engine_contract(rule, build) for rule in rules], "count": len(rules)}


def _rules_registry_contract() -> dict:
    active = get_active_rules()
    gaps = [
        rule.get("rule_id")
        for rule in active
        if not str(rule.get("implementation_ref") or "").strip()
        or not str(rule.get("test_ref") or "").strip()
    ]
    return {
        "mode": "metadata_with_implementation_refs",
        "status": "warning" if gaps else "pass",
        "implementation_gaps": gaps,
        "active_rule_count": len(active),
        "calculation_behavior": "engine_code_and_config",
        "registry_changes_are_metadata_only": True,
        "behavior_change_signal": "engine_and_config_fingerprints",
    }


@app.post("/api/rules/seed")
def api_seed_rules(overwrite: bool = False):
    """Seed built-in hard estimating rules."""
    return seed_rules_registry_defaults(overwrite=overwrite)


@app.get("/api/rulesets")
def api_list_rulesets(limit: int = 25):
    """List whole-registry ruleset versions."""
    versions = list_ruleset_versions(limit=limit)
    current = versions[0] if versions else None
    return {
        "versions": versions,
        "current": current,
        "current_version": current["version"] if current else None,
        "count": len(versions),
        "contract": _rules_registry_contract(),
    }


@app.get("/api/rulesets/{version}")
def api_get_ruleset(version: int):
    """Fetch a whole-registry ruleset snapshot."""
    ruleset = get_ruleset_version(version)
    if not ruleset:
        raise HTTPException(status_code=404, detail="Ruleset version not found")
    return _ruleset_with_engine_contract(ruleset)


@app.post("/api/rulesets/{version}/rollback")
def api_rollback_ruleset(version: int, body: Optional[RulesetRollbackRequest] = None):
    """Restore rules to a previous whole-registry snapshot as a new ruleset version."""
    body = body or RulesetRollbackRequest()
    try:
        new_version = rollback_ruleset_version(
            version,
            changed_by=body.changed_by or "Josh",
            change_note=body.change_note or f"Rolled registry back to ruleset v{version}.",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {
        "status": "ok",
        "rolled_back_to": version,
        "new_version": new_version,
        "current": get_ruleset_version(new_version),
    }


@app.post("/api/rules/draft-from-lesson")
def api_draft_rule_from_lesson(body: RuleDraftRequest):
    """Use AI to turn a spoken/plain-English lesson into a rule draft."""
    lesson = (body.lesson_text or "").strip()
    if len(lesson) < 8:
        raise HTTPException(status_code=400, detail="Tell me a little more about the rule.")

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")
    provider = get_provider_info(api_key)
    if not provider["available"]:
        raise HTTPException(status_code=400, detail="AI is not configured. Add an API key in Settings first.")

    import json as _json
    import re as _re

    system_msg = """You turn spoken estimating lessons into draft rules for a commercial flooring bid platform.
Return ONLY valid JSON. Do not include markdown.

Output shape:
{
  "rule_id": "custom.short.stable.id",
  "name": "Short human name",
  "category": "material|pricing|labor|sundry|freight|tax|proposal|classification|audit",
  "stage": "classification|pricing|sundry|labor|proposal|audit|rfms_parse|quote_parse|sundry_calc|labor_calc|proposal_generate",
  "status": "draft",
  "priority": 10,
  "description": "What Josh wants this to do.",
  "condition_json": {},
  "action_json": {},
  "source": "Josh spoken lesson",
  "notes": "Important assumptions or examples.",
  "change_note": "Initial spoken lesson from Josh.",
  "assumptions": [],
  "needs_review": true
}

Rules:
- Use status "draft" unless the lesson is extremely precise.
- Never claim the app already enforces the rule.
- Use condition_json for WHEN the rule applies.
- Use action_json for WHAT should happen.
- Keep JSON simple and readable for a human reviewer.
- If the spoken lesson is ambiguous, preserve the ambiguity in notes/assumptions instead of inventing specifics."""

    user_msg = f"""Josh said this rule out loud:

{lesson}

Draft the rule fields for the registry. Use a stable rule_id starting with custom."""

    try:
        raw = chat_complete(
            system=system_msg,
            user=user_msg,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        parsed = _json.loads(raw)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not draft rule: {e}")

    if not isinstance(parsed, dict):
        raise HTTPException(status_code=500, detail="AI returned an invalid rule draft.")

    def _slug(value: str) -> str:
        value = _re.sub(r"[^a-z0-9]+", ".", (value or "").lower()).strip(".")
        return value[:72] or "spoken.lesson"

    name = str(parsed.get("name") or lesson[:80]).strip()
    rule_id = str(parsed.get("rule_id") or "").strip()
    if not rule_id.startswith("custom."):
        rule_id = f"custom.{_slug(rule_id or name)}"

    condition_json = parsed.get("condition_json")
    action_json = parsed.get("action_json")
    if not isinstance(condition_json, dict):
        condition_json = {"spoken_condition": str(condition_json or lesson)}
    if not isinstance(action_json, dict):
        action_json = {"spoken_action": str(action_json or "Needs review")}

    try:
        priority = int(parsed.get("priority") or 10)
    except (TypeError, ValueError):
        priority = 10

    return {
        "draft": {
            "rule_id": rule_id,
            "name": name,
            "category": str(parsed.get("category") or "material").strip() or "material",
            "stage": str(parsed.get("stage") or "classification").strip() or "classification",
            "status": "draft",
            "priority": priority,
            "description": str(parsed.get("description") or lesson).strip(),
            "condition_json": condition_json,
            "action_json": action_json,
            "source": str(parsed.get("source") or "Josh spoken lesson").strip(),
            "implementation_ref": "",
            "test_ref": "",
            "notes": str(parsed.get("notes") or "").strip(),
            "changed_by": body.changed_by or "Josh",
            "change_note": str(parsed.get("change_note") or "Initial spoken lesson from Josh.").strip(),
        },
        "assumptions": parsed.get("assumptions") if isinstance(parsed.get("assumptions"), list) else [],
        "needs_review": bool(parsed.get("needs_review", True)),
        "transcript": lesson,
    }


@app.get("/api/rules/{rule_id}/versions")
def api_get_rule_versions(rule_id: str):
    """List all saved versions for a rule."""
    rule = get_rule(rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    versions = list_rule_versions(rule_id)
    build = get_build_info()
    contracted_versions = []
    for version in versions:
        item = dict(version)
        item["snapshot"] = _rule_with_engine_contract(item.get("snapshot") or {}, build)
        contracted_versions.append(item)
    return {"rule": _rule_with_engine_contract(rule, build), "versions": contracted_versions, "count": len(versions)}


@app.get("/api/rules/{rule_id}")
def api_get_rule(rule_id: str):
    """Get a single estimating rule."""
    rule = get_rule(rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    return _rule_with_engine_contract(rule)


@app.post("/api/rules")
def api_create_rule(body: RuleCreate):
    """Create a hard estimating rule."""
    data = body.model_dump()
    data["condition_json"] = data.get("condition_json") or {}
    data["action_json"] = data.get("action_json") or {}
    try:
        rule_id = create_rule(data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Rule already exists")
    return _rule_with_engine_contract(get_rule(rule_id))


@app.put("/api/rules/{rule_id}")
def api_update_rule(rule_id: str, body: RuleUpdate):
    """Update a hard estimating rule. Each edit creates a new version."""
    updates = body.model_dump(exclude_unset=True)
    changed_by = updates.pop("changed_by", None) or "Rules Registry"
    change_note = updates.pop("change_note", None) or "Rule updated from registry."
    if updates.get("name") is None and "name" in updates:
        raise HTTPException(status_code=400, detail="name cannot be null")
    for key in ("category", "stage", "status", "source", "description", "implementation_ref", "test_ref", "notes"):
        if updates.get(key) is None and key in updates:
            updates[key] = ""
    for key in ("priority", "version"):
        if updates.get(key) is None and key in updates:
            raise HTTPException(status_code=400, detail=f"{key} cannot be null")
    if "condition_json" in updates and updates["condition_json"] is None:
        updates["condition_json"] = {}
    if "action_json" in updates and updates["action_json"] is None:
        updates["action_json"] = {}
    if not updates:
        raise HTTPException(status_code=400, detail="No rule fields supplied")
    try:
        updated = update_rule(rule_id, updates, changed_by=changed_by, change_note=change_note)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except sqlite3.IntegrityError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not updated:
        raise HTTPException(status_code=404, detail="Rule not found")
    return _rule_with_engine_contract(get_rule(rule_id))


@app.post("/api/rules/{rule_id}/archive")
def api_archive_rule(rule_id: str, body: Optional[RuleChangeMeta] = None):
    """Archive a rule without erasing its history."""
    body = body or RuleChangeMeta()
    archived = archive_rule(
        rule_id,
        changed_by=(body.changed_by or "Rules Registry"),
        change_note=(body.change_note or "Rule archived from registry."),
    )
    if not archived:
        raise HTTPException(status_code=404, detail="Rule not found")
    return _rule_with_engine_contract(get_rule(rule_id))


@app.delete("/api/rules/{rule_id}")
def api_delete_rule(rule_id: str):
    """Archive a hard estimating rule. History is preserved for old bids."""
    if not delete_rule(rule_id):
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"message": "Rule archived", "rule": _rule_with_engine_contract(get_rule(rule_id))}


@app.get("/api/jobs")
def api_list_jobs():
    """List all jobs."""
    return list_jobs()


@app.get("/api/jobs/match")
def api_match_job(q: str = ""):
    """Fuzzy-match an email subject or project reference to a job.
    Used by the local quote agent to find which job a vendor email belongs to."""
    if not q or len(q) < 3:
        return {"job_id": None, "project_name": None, "gc_name": None, "score": 0}

    # Clean the query — strip RE:, FW:, [EXT], etc.
    import re as _re
    cleaned = _re.sub(r"^(re|fw|fwd|ext|\[ext\])[\s:]+", "", q, flags=_re.IGNORECASE).strip()
    cleaned = _re.sub(r"^(quote\s*request|pricing|price\s*list|proposal)\s*[-:–—]\s*", "", cleaned, flags=_re.IGNORECASE).strip()

    # Search using existing search_all
    results = search_all(cleaned[:80])
    jobs_found = results.get("jobs", [])
    if jobs_found:
        best = jobs_found[0]
        return {
            "job_id": best["id"],
            "project_name": best.get("project_name", ""),
            "gc_name": best.get("gc_name", ""),
            "score": 0.8,
        }

    # Try word-level matching against all jobs
    all_jobs = list_jobs()
    cleaned_lower = cleaned.lower()
    cleaned_words = set(_re.sub(r"[^a-z0-9\s]", "", cleaned_lower).split())

    best_job = None
    best_score = 0
    for job in all_jobs:
        name = (job.get("project_name") or "").lower()
        gc = (job.get("gc_name") or "").lower()
        combined_words = set(_re.sub(r"[^a-z0-9\s]", "", f"{gc} {name}").split())
        if not combined_words or not cleaned_words:
            continue
        overlap = cleaned_words & combined_words
        score = len(overlap) / max(len(cleaned_words), 1)
        if name in cleaned_lower or cleaned_lower in name:
            score += 0.3
        if score > best_score:
            best_score = score
            best_job = job

    if best_job and best_score >= 0.3:
        return {
            "job_id": best_job["id"],
            "project_name": best_job.get("project_name", ""),
            "gc_name": best_job.get("gc_name", ""),
            "score": round(best_score, 2),
        }

    return {"job_id": None, "project_name": None, "gc_name": None, "score": 0}


@app.post("/api/jobs")
def api_create_job(job: JobCreate):
    """Create a new job."""
    job_id = save_job(job.model_dump())
    created = load_job(job_id)
    log_activity(job_id, "job_created", f"Job '{job.project_name}' created")
    return {"id": job_id, "slug": created.get("slug", ""), "message": "Job created"}


def _resolve_job_id(job_id: str) -> int:
    """Resolve a job_id string (could be slug or numeric ID) to a numeric DB id."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job["id"]


@app.get("/api/jobs/{job_id}")
def api_get_job(job_id: str):
    """Get job details by ID or slug."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job["materials_source_fingerprint"] = _materials_source_fingerprint(job.get("materials") or [])
    # Known prices are read-only suggestions. Opening a job must never change
    # accepted source data or stale an existing proposal audit.
    _enrich_known_prices(job)
    return job


def _enrich_known_prices(job: dict):
    """Add transient price suggestions without mutating persisted job inputs."""
    materials = job.get("materials", [])
    applied_count = 0
    if not materials:
        return 0
    price_list = get_price_list_entries()

    # Build a map of normalized vendor prices (latest price per product)
    conn = _get_conn()
    try:
        rows = conn.execute("""
            SELECT vp.product_normalized, vp.unit_price, vp.vendor_name, vp.unit,
                   vp.created_at AS latest_date
            FROM vendor_prices vp
            WHERE vp.unit_price > 0
              AND vp.id = (
                  SELECT newest.id
                  FROM vendor_prices newest
                  WHERE newest.product_normalized = vp.product_normalized
                    AND newest.unit_price > 0
                  ORDER BY newest.created_at DESC, newest.id DESC
                  LIMIT 1
              )
            ORDER BY vp.created_at DESC
        """).fetchall()
        vendor_map = {r["product_normalized"]: dict(r) for r in rows}
    finally:
        conn.close()

    for mat in materials:
        if mat.get("unit_price") and mat["unit_price"] > 0:
            continue  # already has a price, skip

        item_code = (mat.get("item_code") or "").strip().lower()
        description = (mat.get("description") or "").strip().lower()

        # Check price list first
        pl_match = _match_price_list(mat, price_list)
        if pl_match and pl_match.get("unit_price"):
            order_qty = mat.get("order_qty") or mat.get("installed_qty") or 1
            mat["known_price"] = round(pl_match["unit_price"] * order_qty, 2)
            mat["known_unit_price"] = pl_match["unit_price"]
            mat["known_price_source"] = "price_list"
            mat["known_price_vendor"] = pl_match.get("vendor", "")
            applied_count += 1
            continue

        # Check vendor price history
        normalized = _normalize_product(item_code or description)
        if normalized and len(normalized) >= 3:
            for key, vp in vendor_map.items():
                if normalized in key or key in normalized:
                    order_qty = mat.get("order_qty") or mat.get("installed_qty") or 1
                    mat["known_price"] = round(vp["unit_price"] * order_qty, 2)
                    mat["known_unit_price"] = vp["unit_price"]
                    mat["known_price_source"] = "vendor_history"
                    mat["known_price_vendor"] = vp.get("vendor_name", "")
                    applied_count += 1
                    break

    return applied_count


@app.post("/api/jobs/bulk-delete")
def api_bulk_delete(body: BulkDeleteRequest):
    """Delete multiple jobs."""
    deleted = 0
    for jid in body.job_ids:
        if delete_job(jid):
            deleted += 1
    return {"deleted": deleted}


@app.delete("/api/jobs/{job_id}")
def api_delete_job(job_id: str):
    """Delete a job by ID or slug and all related data."""
    db_id = _resolve_job_id(job_id)
    if not delete_job(db_id):
        raise HTTPException(status_code=404, detail="Job not found")
    return {"message": "Job deleted"}


@app.post("/api/jobs/{job_id}/duplicate")
def api_duplicate_job(job_id: str):
    """Duplicate a job and all its materials."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Create new job with same fields
    new_job = {
        "project_name": job["project_name"] + " (Copy)",
        "gc_name": job.get("gc_name"),
        "address": job.get("address"),
        "city": job.get("city"),
        "state": job.get("state"),
        "zip": job.get("zip"),
        "tax_rate": job.get("tax_rate", 0),
        "gpm_pct": job.get("gpm_pct", 0),
        "unit_count": job.get("unit_count", 0),
        "tub_shower_count": job.get("tub_shower_count", 0),
        "salesperson": job.get("salesperson"),
        "notes": job.get("notes"),
        "exclusions": job.get("exclusions"),
        "markup_pct": job.get("markup_pct", 0),
        "architect": job.get("architect"),
        "designer": job.get("designer"),
        "textura_fee": job.get("textura_fee", 0),
    }
    new_id = save_job(new_job)

    # Copy materials (strip id and job_id)
    materials = job.get("materials", [])
    copied = []
    for m in materials:
        mat = {k: v for k, v in m.items() if k not in ("id", "job_id")}
        copied.append(mat)
    if copied:
        save_materials(new_id, copied)

    created = load_job(new_id)
    log_activity(new_id, "job_created", f"Duplicated from '{job['project_name']}'", {"source_job_id": job["id"]})
    return {"id": new_id, "slug": created.get("slug", "")}


@app.put("/api/jobs/{job_id}/notes")
def api_update_notes(job_id: str, body: NotesUpdate):
    """Update job notes."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job["notes"] = body.notes
    save_job(job)
    log_activity(job["id"], "notes_updated", "Notes updated")
    return {"message": "Notes saved"}


class JobUpdate(BaseModel):
    markup_pct: Optional[float] = None
    gpm_pct: Optional[float] = None
    project_name: Optional[str] = None
    gc_name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip: Optional[str] = None
    tax_rate: Optional[float] = None
    unit_count: Optional[int] = None
    tub_shower_count: Optional[int] = None
    salesperson: Optional[str] = None
    notes: Optional[str] = None
    architect: Optional[str] = None
    designer: Optional[str] = None
    textura_fee: Optional[int] = None

@app.put("/api/jobs/{job_id}")
def api_update_job(job_id: str, body: JobUpdate):
    """Update job fields."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    updates = body.model_dump(exclude_none=True)
    changes = {}
    for key, val in updates.items():
        old_val = job.get(key)
        if old_val != val:
            changes[key] = {"old": old_val, "new": val}
    for key, val in updates.items():
        job[key] = val
    save_job(job)
    if changes:
        changed_keys = ", ".join(changes.keys())
        log_activity(job["id"], "job_updated", f"Updated {changed_keys}", {"changes": changes})
    return {"message": "Job updated"}


@app.post("/api/jobs/{job_id}/upload-rfms")
async def api_upload_rfms(job_id: str, request: Request, files: list[UploadFile] = File(default=None)):
    """Upload one or more RFMS pivot tables, parse them, return merged materials."""
    # Debug: log what we received
    ct = request.headers.get("content-type", "")
    print(f"[rfms_upload] Content-Type: {ct}")
    print(f"[rfms_upload] files param: {files}, type: {type(files)}")

    # If 'files' field is missing, try reading from the raw form
    if not files:
        form = await request.form()
        print(f"[rfms_upload] Raw form keys: {list(form.keys())}")
        files = form.getlist("files") or form.getlist("file")
        print(f"[rfms_upload] Extracted files: {files}")
        if not files:
            raise HTTPException(status_code=422, detail=f"No files received. Form keys: {list(form.keys())}")

    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)

    all_materials_raw = []
    rfms_job_info = {}
    imported_uploads = []

    for file in files:
        if os.path.splitext(file.filename or "")[1].lower() != ".xlsx":
            raise HTTPException(status_code=400, detail=f"RFMS file '{file.filename}' must be an .xlsx workbook.")
        content = await file.read(MAX_RFMS_FILE_BYTES + 1)
        if len(content) > MAX_RFMS_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"RFMS file '{file.filename}' exceeds the {MAX_RFMS_FILE_BYTES // (1024 * 1024)} MB limit.",
            )
        file_hash = hashlib.sha256(content).hexdigest()
        file_path = _job_upload_path(db_id, f"{file_hash[:12]}_{file.filename}", "rfms")
        with open(file_path, "wb") as f:
            f.write(content)
        _record_artifact(db_id, file_path, "rfms")

        try:
            result = parse_rfms(file_path)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse RFMS file '{file.filename}': {e}")
        imported_uploads.append((file.filename, file_hash, len(content), file_path))

        # Use job info from first file that has it
        file_job_info = result.get("job_info", {})
        if not rfms_job_info and any(file_job_info.values()):
            rfms_job_info = file_job_info

        # Detect area_type from filename: "common area(s)" or "amenity" → common, else unit
        fname_lower = (file.filename or "").lower()
        area_type = "common" if ("common area" in fname_lower or "amenity" in fname_lower or "common_area" in fname_lower) else "unit"
        for mat in result.get("materials", []):
            mat["area_type"] = area_type
        all_materials_raw.extend(result.get("materials", []))

    for filename, file_hash, file_size, file_path in imported_uploads:
        record_imported_file(
            db_id,
            filename,
            file_hash,
            file_size,
            source="rfms",
            artifact_path=os.path.relpath(file_path, ARTIFACT_ROOT),
            artifact_kind="rfms",
        )

    # Update job info from RFMS if available
    job_update = {
        "id": db_id,
        "project_name": rfms_job_info.get("project_name") or job["project_name"],
        "gc_name": rfms_job_info.get("gc_name") or job.get("gc_name"),
        "address": rfms_job_info.get("address") or job.get("address"),
        "city": rfms_job_info.get("city") or job.get("city"),
        "state": rfms_job_info.get("state") or job.get("state"),
        "zip": rfms_job_info.get("zip") or job.get("zip"),
        "tax_rate": job.get("tax_rate", 0),
        "gpm_pct": job.get("gpm_pct", 0),
        "unit_count": job.get("unit_count", 0),
        "tub_shower_count": job.get("tub_shower_count", 0),
        "salesperson": job.get("salesperson"),
        "notes": job.get("notes"),
        "exclusions": job.get("exclusions"),
        "markup_pct": job.get("markup_pct", 0),
        "bid_data": job.get("bid_data"),
        "proposal_data": job.get("proposal_data"),
        "architect": job.get("architect"),
        "designer": job.get("designer"),
        "textura_fee": job.get("textura_fee", 0),
    }
    save_job(job_update)

    # AI merge if job already has materials, otherwise use raw parsed
    existing_materials = job.get("materials", [])
    if existing_materials:
        print(f"[rfms_upload] Job has {len(existing_materials)} existing materials, running AI merge")
        merged_raw = ai_merge_materials(existing_materials, all_materials_raw)
    else:
        # First upload — just use the parsed materials directly
        merged_raw = []
        for m in all_materials_raw:
            merged_raw.append({
                "item_code": m.get("item_code"),
                "description": m.get("description"),
                "material_type": m.get("material_type", "unknown"),
                "installed_qty": m.get("qty", 0),
                "unit": m.get("unit"),
                "area_type": m.get("area_type", "unit"),
                "tack_strip_lf": m.get("tack_strip_lf", 0),
                "seam_tape_lf": m.get("seam_tape_lf", 0),
                "pad_sy": m.get("pad_sy", 0),
                "is_mosaic": m.get("is_mosaic", False),
                "is_penny_hex": m.get("is_penny_hex", False),
                "crack_isolation_sf": m.get("crack_isolation_sf", 0),
            })

    # Deduplicate by item_code — if the same item_code appears with different
    # area_types (e.g. from uploading unit + common area files), keep only the first
    seen_codes = {}
    deduped = []
    for m in merged_raw:
        code = m.get("item_code")
        if code and code in seen_codes:
            # Skip duplicate — same item_code already present
            continue
        if code:
            seen_codes[code] = True
        deduped.append(m)
    merged_raw = deduped

    # Load waste factors from DB (falls back to config defaults)
    import json as _json
    _waste_data = get_company_rate("waste_factors")
    _waste_factors = _json.loads(_waste_data) if _waste_data else WASTE_FACTORS

    # Load price list for auto-pricing
    _price_list = get_price_list_entries()

    # Apply waste factors to the final merged list
    materials = []
    for m in merged_raw:
        material_type = m.get("material_type", "unknown")
        waste_pct = _waste_factors.get(material_type, 0)
        installed_qty = m.get("installed_qty", m.get("qty", 0))
        order_qty = installed_qty * (1 + waste_pct)

        # Auto-price from internal price list and price books
        unit_price = m.get("unit_price", 0)
        vendor = m.get("vendor", "")
        price_source = m.get("price_source")
        if not unit_price and _price_list:
            matched = _match_price_list(m, _price_list)
            if matched:
                unit_price = matched["unit_price"]
                vendor = matched.get("vendor", "")
                price_source = "price_list"
        # Check price_book_items (e.g. Schluter catalog)
        # Schluter products come in 8' sticks — round up to full sticks
        unit_override = None
        if not unit_price:
            pb_match = _match_price_book(m)
            if pb_match:
                import math
                stick_price = pb_match.get("stick_price", 0)
                stick_lf = pb_match.get("stick_lf", 8.208)
                sticks_needed = math.ceil(order_qty / stick_lf) if order_qty > 0 else 0
                vendor = pb_match.get("vendor", "")
                price_source = pb_match.get("price_source", "price_book")
                # Price by full sticks rounded up
                order_qty = sticks_needed
                unit_price = stick_price
                unit_override = "EA"

        # Set quote_status for unpriced materials
        quote_status = m.get("quote_status")
        if not unit_price and not quote_status:
            quote_status = "needs_quote"

        materials.append({
            "item_code": m.get("item_code"),
            "description": m.get("description"),
            "material_type": material_type,
            "installed_qty": round(installed_qty, 2),
            "unit": unit_override or m.get("unit"),
            "waste_pct": waste_pct,
            "order_qty": round(order_qty, 2),
            "vendor": vendor,
            "unit_price": unit_price,
            "extended_cost": round(unit_price * round(order_qty, 2), 2),
            "ai_confidence": m.get("ai_confidence"),
            "quote_status": quote_status,
            "price_source": price_source,
            "area_type": m.get("area_type", "unit"),
            "tack_strip_lf": m.get("tack_strip_lf", 0),
            "seam_tape_lf": m.get("seam_tape_lf", 0),
            "pad_sy": m.get("pad_sy", 0),
            "is_mosaic": m.get("is_mosaic", False),
            "is_penny_hex": m.get("is_penny_hex", False),
            "crack_isolation_sf": m.get("crack_isolation_sf", 0),
        })

    material_ids = save_materials(db_id, materials)

    # Attach IDs to returned materials
    for mat, mid in zip(materials, material_ids):
        mat["id"] = mid

    file_names = [f.filename for f in files if hasattr(f, 'filename')]
    log_activity(db_id, "rfms_uploaded", f"Uploaded {len(file_names)} RFMS file(s), {len(materials)} materials parsed", {"files": file_names, "material_count": len(materials)})

    return {"job_info": rfms_job_info, "materials": materials}


def _apply_fob_freight(mat: dict, prod: dict, freight_rates: dict | None = None):
    """When vendor freight is FOB (we pay shipping), apply internal freight rates
    based on material type. CPT/carpet tile uses cpt_tile rate, LVT uses lvt rate."""
    freight_val = prod.get("freight") or ""
    if not isinstance(freight_val, str) or "fob" not in freight_val.lower():
        return  # Not FOB — freight is either included or a dollar amount

    mat_type = (mat.get("material_type") or "").lower()
    unit = (mat.get("unit") or "").upper()
    description = (mat.get("description") or "").lower()

    # Determine freight rate from the editable company table, with config as the
    # startup fallback. This same table is snapshotted for golden replay.
    freight_rates = freight_rates if isinstance(freight_rates, dict) else FREIGHT_RATES
    rate = None
    if mat_type in ("carpet_tile", "cpt", "cpt_tile") or "carpet tile" in description or "cpt" in (mat.get("item_code") or "").lower():
        rate = freight_rates.get("cpt_tile", 1.25)  # per SY
    elif mat_type in ("lvt", "unit_lvt") or "lvt" in description or "lvt" in (mat.get("item_code") or "").lower() or "vinyl plank" in description:
        # Determine LVT thickness from description
        if "5mm" in description or "4.5mm" in description or "5.0mm" in description:
            rate = freight_rates.get("lvt_5mm", 0.25)  # per SF
        else:
            rate = freight_rates.get("lvt_2mm", 0.11)  # per SF
    elif mat_type in ("broadloom",) or "broadloom" in description:
        rate = freight_rates.get("broadloom", 0.65)  # per SY

    if rate is not None:
        mat["freight_per_unit"] = rate
        mat["freight_source"] = "internal_rate"
        print(f"[freight] FOB detected for {mat.get('item_code', '?')} — applied internal rate ${rate}/{unit}")


def _auto_match_quotes(job_id: int, products: list[dict]) -> int:
    """Try to match parsed quote products to existing materials.
    Phase 1: exact item_code matching (fast, no AI).
    Phase 2: AI fuzzy matching for remaining unmatched items."""
    job = load_job(job_id)
    if not job:
        return 0
    materials = job.get("materials", [])
    if not materials:
        return 0

    matched = 0
    updated = False
    matched_mat_indices = set()
    matched_prod_indices = set()
    freight_rates = (get_all_company_rates().get("freight_rates") or FREIGHT_RATES)

    # Phase 1: Fast matching — item_code AND description-based product identifiers
    # Extract searchable identifiers from material descriptions.
    # e.g. "Interface - Woven Gradience - WG100 - 108051 Onyx" → ["wg100", "108051", "woven gradience"]
    import re as _re

    def _extract_vendor_from_desc(description: str) -> str:
        """Extract vendor name from RFMS description.
        RFMS format: '(Standard) - T-200 - Arizona Tile - Flash - Ivory...'
        The vendor name is typically the part after the item code, before the product line.
        Only returns a vendor if the description has the standard ' - ' delimited format
        with at least 3 parts (code - vendor - product)."""
        parts = [p.strip() for p in description.split(" - ") if p.strip()]
        if len(parts) < 3:
            return ""  # Not enough parts for code - vendor - product format
        # Skip option prefixes and item codes to find the vendor name
        for part in parts:
            clean = part.strip()
            # Skip option prefixes like (Standard), (Premium), (Alternate)
            if _re.match(r'^\(', clean):
                continue
            # Skip item codes like CPT-200, T-202, F103, LVT-200, B-101, TS-100
            if _re.match(r'^[A-Z]{1,4}-?\d{1,4}', clean):
                continue
            # Skip numeric-only parts
            if _re.match(r'^\d+', clean):
                continue
            # This should be the vendor name
            return clean.lower()
        return ""

    def _extract_identifiers(description: str) -> list[str]:
        """Extract product line identifiers from a material description.
        Splits on ' - ' delimiters and returns meaningful tokens."""
        parts = [p.strip().lower() for p in description.split(" - ") if p.strip()]
        identifiers = []
        for part in parts:
            # Skip the vendor name (first part) and very short/generic tokens
            if len(part) < 2:
                continue
            identifiers.append(part)
            # Also extract individual alphanumeric codes (WG100, 108051, etc.)
            codes = _re.findall(r'[a-z]*\d+[a-z]*\d*', part)
            identifiers.extend(codes)
        return identifiers

    # Generic words that should NOT count as meaningful matches on their own
    GENERIC_WORDS = {
        "tile", "carpet", "floor", "flooring", "interface", "mohawk", "shaw",
        "matte", "glossy", "polished", "honed", "satin", "brushed",  # finishes
        "black", "white", "grey", "gray", "brown", "beige", "cream", "ivory",  # colors
        "wall", "base", "trim", "edge", "cove", "corner",  # generic parts
        "custom", "standard", "premium", "commercial", "residential",
        "rubber", "vinyl", "porcelain", "ceramic", "glass", "stone", "marble",
        "daltile", "johnsonite", "schluter", "mannington",  # vendor names
        "rectangular", "square", "round", "linear", "straight",
        "size", "type", "style", "color", "finish", "series",
    }

    # Also collect ALL quote products across the DB for this job (not just current upload)
    # so we can score against the full universe of quotes
    conn = _get_conn()
    try:
        all_quotes = conn.execute(
            "SELECT id, product_name, vendor, unit_price, unit, file_name, freight, lead_time, notes FROM job_quotes WHERE job_id=?",
            (job_id,),
        ).fetchall()
    finally:
        conn.close()
    all_quote_products = [dict(q) for q in all_quotes] if all_quotes else products

    for mat_idx, mat in enumerate(materials):
        if mat.get("unit_price") and mat["unit_price"] > 0:
            continue  # already priced
        item_code = (mat.get("item_code") or "").strip().lower()
        description = (mat.get("description") or "").strip().lower()
        if not item_code and not description:
            continue

        # Build list of identifiers to match against
        mat_identifiers = _extract_identifiers(description)

        # Extract vendor name from RFMS description for hard filtering
        # e.g. "(Standard) - T-200 - Arizona Tile - Flash" → "arizona tile"
        rfms_vendor = _extract_vendor_from_desc(mat.get("description") or "")

        # Score ALL products and pick the best match instead of first-match-wins
        best_score = 0
        best_prod = None
        best_prod_idx = None

        # Use all_quote_products for scoring (includes previous uploads)
        scoring_products = all_quote_products if all_quote_products else products

        for prod_idx, prod in enumerate(scoring_products):
            if prod.get("error"):
                continue
            unit_price = prod.get("unit_price", 0)
            if not unit_price:
                continue
            prod_name = (prod.get("product_name") or "").strip().lower()
            prod_desc = (prod.get("description") or "").strip().lower()
            prod_vendor = (prod.get("vendor") or "").strip().lower()
            prod_text = f"{prod_name} {prod_desc}"

            # HARD FILTER: If RFMS description names a vendor, only match quotes
            # from that vendor. "T-200 - Arizona Tile - Flash" must match Arizona Tile
            # quotes, never Metropolitan Floors or anyone else.
            # Vendor aliases: parent companies own subsidiaries (Daltile=Marazzi, etc.)
            _VENDOR_ALIASES = {
                "marazzi": ["daltile"],
                "flor": ["interface"], "interface": ["flor"],
                "mohawk": ["daltile", "marazzi"], "daltile": ["marazzi", "mohawk"],
            }
            if rfms_vendor and len(rfms_vendor) >= 3:
                vendor_match = False
                # Direct match
                if rfms_vendor in prod_vendor or prod_vendor in rfms_vendor:
                    vendor_match = True
                # Check aliases (e.g. Marazzi material can match Daltile quote)
                if not vendor_match:
                    aliases = _VENDOR_ALIASES.get(rfms_vendor, [])
                    for alias in aliases:
                        if alias in prod_vendor or prod_vendor in alias:
                            vendor_match = True
                            break
                # Also check product name/file for vendor name
                prod_file = (prod.get("file_name") or "").lower()
                if rfms_vendor in prod_file:
                    vendor_match = True
                if not vendor_match:
                    continue  # SKIP — wrong vendor, don't even score

            score = 0

            # Check 1: item_code in product name/desc (strong signal: +10)
            if item_code and len(item_code) >= 3:
                if item_code in prod_name or item_code in prod_desc:
                    score += 10

            # Check 2: product identifiers from description match quote product
            if mat_identifiers:
                for ident in mat_identifiers:
                    if len(ident) >= 3 and ident in prod_text:
                        # Weight by specificity: codes with digits worth more
                        if _re.search(r'\d', ident):
                            score += 5  # alphanumeric codes like "d617", "wg100"
                        elif ident not in GENERIC_WORDS and len(ident) >= 4:
                            score += 2  # meaningful product names
                        elif ident not in GENERIC_WORDS:
                            score += 1

            # Check 3: quote product name found in material description
            if prod_name and len(prod_name) >= 4:
                prod_parts = [p.strip() for p in prod_name.split(" - ") if len(p.strip()) >= 3]
                for pp in prod_parts:
                    if pp in description:
                        score += 3

                # Individual significant words from product name
                prod_words = [w for w in _re.findall(r'[a-z]+\d*\S*', prod_name) if len(w) >= 4]
                desc_words = set(_re.findall(r'[a-z]+\d*\S*', description))
                for pw in prod_words:
                    if pw in desc_words and pw not in GENERIC_WORDS:
                        score += 2

            # Check 4: Word-level overlap scoring
            desc_tokens = set(_re.findall(r'[a-z]+', description))
            prod_tokens = set(_re.findall(r'[a-z]+', prod_text))
            meaningful_overlap = (desc_tokens & prod_tokens) - GENERIC_WORDS - {"and", "the", "for", "with"}
            meaningful_overlap = {w for w in meaningful_overlap if len(w) >= 4}
            score += len(meaningful_overlap)

            # Check 5: Dimension/size match (e.g. "1x1", "12x24", "4x12")
            desc_dims = set(_re.findall(r'(\d+)\s*["\']?\s*x\s*["\']?\s*(\d+)', description))
            prod_dims = set(_re.findall(r'(\d+)\s*["\']?\s*x\s*["\']?\s*(\d+)', prod_text))
            if desc_dims and prod_dims and desc_dims & prod_dims:
                score += 3

            if score > best_score:
                best_score = score
                best_prod = prod
                best_prod_idx = prod_idx

        # Require a minimum score of 3 to accept a match (prevents single generic word matches)
        if best_score >= 3 and best_prod:
            mat["unit_price"] = best_prod["unit_price"]
            mat["vendor"] = best_prod.get("vendor", "")
            mat["quote_status"] = "quoted"
            mat["price_source"] = "vendor_quote"
            # If freight is FOB, apply internal freight rates by material type
            _apply_fob_freight(mat, best_prod, freight_rates)
            order_qty = mat.get("order_qty", 0)
            mat["extended_cost"] = round(order_qty * mat["unit_price"], 2)
            matched += 1
            updated = True
            matched_mat_indices.add(mat_idx)

    # Phase 2: AI fuzzy matching for remaining unmatched
    unmatched_mats = [(i, m) for i, m in enumerate(materials) if i not in matched_mat_indices and (not m.get("unit_price") or m["unit_price"] == 0)]
    unmatched_prods = [(i, p) for i, p in enumerate(products) if i not in matched_prod_indices and not p.get("error") and p.get("unit_price")]

    if unmatched_mats and unmatched_prods:
        ai_matched = _ai_match_quotes(unmatched_mats, unmatched_prods)
        for mat_idx, prod_idx in ai_matched:
            mat = materials[mat_idx]
            prod = products[prod_idx]
            mat["unit_price"] = prod["unit_price"]
            mat["vendor"] = prod.get("vendor", "")
            mat["quote_status"] = "quoted"
            mat["price_source"] = "vendor_quote"
            _apply_fob_freight(mat, prod, freight_rates)
            order_qty = mat.get("order_qty", 0)
            mat["extended_cost"] = round(order_qty * mat["unit_price"], 2)
            matched += 1
            updated = True

    # Phase 3: Apply transition default rules (Carpet→LVT = Silver Pin, etc.)
    still_unpriced = [i for i, m in enumerate(materials) if i not in matched_mat_indices and (not m.get("unit_price") or m["unit_price"] == 0)]
    if still_unpriced:
        td_matched = _apply_transition_defaults(materials, still_unpriced)
        if td_matched > 0:
            matched += td_matched
            updated = True

    # Phase 4: Price book matching for remaining unpriced materials
    # Check if any unpriced materials match a vendor price book (e.g. Schluter transitions)
    still_unpriced2 = [i for i, m in enumerate(materials) if i not in matched_mat_indices and (not m.get("unit_price") or m["unit_price"] == 0)]
    if still_unpriced2:
        pb_matched = _price_book_match(materials, still_unpriced2)
        if pb_matched > 0:
            matched += pb_matched
            updated = True

    # Phase 5: Apply labor rates to all Schluter transitions (even those matched by vendor quotes)
    for mat in materials:
        if (mat.get("vendor") or "").lower() == "schluter" and not mat.get("labor_rate_lf"):
            desc = (mat.get("description") or "").lower()
            is_premium = any(line in desc for line in SCHLUTER_PREMIUM_LABOR_LINES)
            mat["labor_rate_lf"] = SCHLUTER_LABOR_RATE_PREMIUM if is_premium else SCHLUTER_LABOR_RATE_DEFAULT
            mat["labor_catalog"] = "Schluter Schiene"
            updated = True

    if updated:
        save_materials(job_id, materials)

    # Also try to link to open quote requests
    _link_upload_to_requests(job_id, products)

    return matched


def _ai_match_quotes(unmatched_mats: list, unmatched_prods: list) -> list:
    """Use AI to fuzzy-match vendor products to job materials."""
    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if not provider["available"]:
        return []

    try:
        import json

        mat_lines = []
        for i, (idx, m) in enumerate(unmatched_mats):
            mat_lines.append(f"M{i}: [{m.get('item_code', '')}] {m.get('description', '')} (type: {m.get('material_type', '')})")

        prod_lines = []
        for i, (idx, p) in enumerate(unmatched_prods):
            prod_lines.append(f"P{i}: {p.get('product_name', '')} — {p.get('description', '')} (vendor: {p.get('vendor', '')}, ${p.get('unit_price', 0)}/{p.get('unit', 'unit')})")

        prompt = f"""Match vendor-quoted products to our job materials. These are commercial flooring products.

Our unmatched materials:
{chr(10).join(mat_lines)}

Vendor quoted products:
{chr(10).join(prod_lines)}

Match products to materials based on: brand name, product line, color, style number, dimensions.
Only match if you are 80%+ confident they are the same product.

Return ONLY a JSON object with a "matches" key: {{"matches": [{{"material": "M0", "product": "P0", "confidence": 0.95}}]}}
Return {{"matches": []}} if no confident matches."""

        raw = chat_complete(
            system="You are a commercial flooring product matching assistant. Return JSON only.",
            user=prompt,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        result = json.loads(raw)

        # Parse result
        matches_raw = result if isinstance(result, list) else result.get("matches", result.get("results", result.get("data", [])))
        if not isinstance(matches_raw, list):
            matches_raw = []

        pairs = []
        for m in matches_raw:
            conf = m.get("confidence", 0)
            if conf < 0.8:
                continue
            mat_ref = m.get("material", "")
            prod_ref = m.get("product", "")
            try:
                mat_local_idx = int(mat_ref.replace("M", ""))
                prod_local_idx = int(prod_ref.replace("P", ""))
                if 0 <= mat_local_idx < len(unmatched_mats) and 0 <= prod_local_idx < len(unmatched_prods):
                    pairs.append((unmatched_mats[mat_local_idx][0], unmatched_prods[prod_local_idx][0]))
            except (ValueError, IndexError):
                continue

        return pairs
    except Exception as e:
        print(f"AI quote matching failed (non-fatal): {e}")
        return []


# ─── Transition Default Rules ───────────────────────────────────────────────
# These rules define the default product, price, and labor rate for each
# transition type. Applied BEFORE AI/price-book matching so they take priority.
TRANSITION_DEFAULTS = [
    {
        "match": "carpet to lvt",  # substring match on description
        "product": "Silver Pin Metal",
        "vendor": "Silver Pin Metal",
        "price_per_piece": 7.94,
        "stick_length_lf": 12.0,
        "labor_rate_lf": 0,  # no install labor for pin metal
        "labor_catalog": "",
        "price_source": "default_rule",
    },
    {
        "match": "vertical exposed edge",
        "product": "Schluter Jolly J 100 AE",
        "vendor": "Schluter",
        "price_per_piece": 9.78,  # Jolly J 100 AE net
        "stick_length_lf": 8.0 + 2.0 / 12.0,  # 8'-2"
        "labor_rate_lf": 0.50,
        "labor_catalog": "Schluter Schiene",
        "price_source": "price_book",
    },
    {
        "match": "tile to lvt",
        "product": "Schluter Reno-U AEU 100",
        "vendor": "Schluter",
        "price_per_piece": 12.15,  # Reno-U AEU 100 net
        "stick_length_lf": 8.0 + 2.0 / 12.0,
        "labor_rate_lf": 0.50,
        "labor_catalog": "Schluter Schiene",
        "price_source": "price_book",
    },
    {
        "match": "tile to carpet",  # tile to CPT
        "match_alt": "tile to cpt",
        "product": "Schluter Reno-TK AETK 100",
        "vendor": "Schluter",
        "price_per_piece": 12.80,  # Reno-TK AETK 100 net
        "stick_length_lf": 8.0 + 2.0 / 12.0,
        "labor_rate_lf": 0.50,
        "labor_catalog": "Schluter Schiene",
        "price_source": "price_book",
    },
]

# Catch-all default for any transition not matched by specific rules above.
# Generic transitions (Carpet to Tile, Carpet to VCT, @Courtyard, etc.)
# default to Schluter Reno-TK AE 100.
TRANSITION_CATCHALL = {
    "product": "Schluter Reno-TK AETK 100",
    "vendor": "Schluter",
    "price_per_piece": 12.80,  # Reno-TK AETK 100 net
    "stick_length_lf": 8.0 + 2.0 / 12.0,
    "labor_rate_lf": 0.50,
    "labor_catalog": "Schluter Schiene",
    "price_source": "price_book",
}

# Schluter labor rate exceptions: Dilex, Rondec, Quadec = $1.07/LF (all others = $0.50/LF)
SCHLUTER_LABOR_RATE_DEFAULT = 0.50
SCHLUTER_LABOR_RATE_PREMIUM = 1.07
SCHLUTER_PREMIUM_LABOR_LINES = {"dilex", "rondec", "quadec"}


def _apply_transition_defaults(materials: list[dict], unpriced_indices: list[int]) -> int:
    """Apply default transition product/pricing rules BEFORE AI matching.
    Returns number of materials matched."""
    import math
    matched = 0

    for mat_idx in unpriced_indices:
        mat = materials[mat_idx]
        mat_type = (mat.get("material_type") or "").lower()
        if mat_type != "transitions":
            continue

        description = (mat.get("description") or "").lower()
        order_qty_lf = mat.get("order_qty", 0)
        fixture_count = mat.get("fixture_count", 0) or 0

        # Check each rule
        for rule in TRANSITION_DEFAULTS:
            match_str = rule["match"]
            match_alt = rule.get("match_alt", "")
            if match_str in description or (match_alt and match_alt in description):
                price_per_piece = rule["price_per_piece"]
                stick_lf = rule["stick_length_lf"]

                # Calculate pieces
                if rule["vendor"] == "Schluter":
                    pieces = _calc_schluter_pieces(order_qty_lf, fixture_count, stick_lf)
                else:
                    pieces = math.ceil(order_qty_lf / stick_lf) if order_qty_lf > 0 else 0

                mat["unit_price"] = price_per_piece
                mat["vendor"] = rule["vendor"]
                mat["price_source"] = rule["price_source"]
                mat["quote_status"] = "price_book"
                mat["extended_cost"] = round(pieces * price_per_piece, 2)

                # Set labor rate
                labor_rate = rule["labor_rate_lf"]
                mat["labor_rate_lf"] = labor_rate
                mat["labor_catalog"] = rule.get("labor_catalog", "")

                matched += 1
                break  # first matching rule wins
        else:
            # No specific rule matched — skip named Schluter products (they'll match via price book)
            # For generic transitions, apply catch-all default (Reno-TK AE 100)
            if "schluter" not in description:
                rule = TRANSITION_CATCHALL
                stick_lf = rule["stick_length_lf"]
                pieces = _calc_schluter_pieces(order_qty_lf, fixture_count, stick_lf)
                mat["unit_price"] = rule["price_per_piece"]
                mat["vendor"] = rule["vendor"]
                mat["price_source"] = rule["price_source"]
                mat["quote_status"] = "price_book"
                mat["extended_cost"] = round(pieces * rule["price_per_piece"], 2)
                mat["labor_rate_lf"] = rule["labor_rate_lf"]
                mat["labor_catalog"] = rule["labor_catalog"]
                matched += 1

        # If no rule matched but it's a named Schluter product, apply labor rates
        if (mat.get("vendor") or "").lower() == "schluter" and not mat.get("labor_rate_lf"):
            is_premium = any(line in description for line in SCHLUTER_PREMIUM_LABOR_LINES)
            mat["labor_rate_lf"] = SCHLUTER_LABOR_RATE_PREMIUM if is_premium else SCHLUTER_LABOR_RATE_DEFAULT
            mat["labor_catalog"] = "Schluter Schiene"

    return matched


def _calc_schluter_pieces(order_qty_lf: float, fixture_count: int, piece_lf: float) -> int:
    """Calculate number of Schluter pieces needed.
    If fixture_count is set, each fixture needs full pieces per side (no splicing across fixtures).
    E.g. 200 showers at 7'-6" each side: each side = 1 piece, 2 sides = 2 pieces/fixture = 400 total.
    Without fixture_count, just divides total LF by piece length."""
    import math
    if order_qty_lf <= 0:
        return 0
    if fixture_count and fixture_count > 0:
        # Fixture-based: each fixture needs full pieces, can't reuse leftover across fixtures
        # Assume 2 sides per fixture (tub/shower has left + right)
        sides = 2
        lf_per_side = order_qty_lf / (fixture_count * sides)
        pieces_per_side = math.ceil(lf_per_side / piece_lf)
        return fixture_count * sides * pieces_per_side
    else:
        return math.ceil(order_qty_lf / piece_lf)


def _price_book_match(materials: list[dict], unpriced_indices: list[int]) -> int:
    """Match unpriced materials against vendor price books (e.g. Schluter).
    Only matches when description explicitly contains a Schluter product line name
    as a whole word AND has additional identifying info (item number, size, etc.)."""
    import re as _re

    # Known Schluter product lines — only match as whole words to avoid false positives
    # e.g. "deco" must not match "decorative", "trep" must not match "trepidation"
    SCHLUTER_LINES = [
        "schiene", "reno-t", "reno-tk", "reno-u", "reno-v", "reno-ramp",
        "jolly", "ditra", "kerdi", "kerdi-band", "kerdi-board",
        "dilex-ahka", "dilex-ahk", "dilex", "quadec", "rondec", "trep-e", "trep-b", "trep-s",
        "trep-fl", "trep-ek", "trep-se", "trep-tap",
    ]
    # Compile whole-word patterns (avoid substring matches like "deco" in "decorative")
    SCHLUTER_PATTERNS = {
        line: _re.compile(r'\b' + _re.escape(line) + r'\b', _re.IGNORECASE)
        for line in SCHLUTER_LINES
    }

    matched = 0
    ai_candidates = []
    for mat_idx in unpriced_indices:
        mat = materials[mat_idx]
        description = (mat.get("description") or "").lower()
        item_code = (mat.get("item_code") or "").lower()
        mat_type = (mat.get("material_type") or "").lower()

        # STRICT GATE: Only match if material is a known Schluter-applicable type
        # OR the description explicitly says "schluter"
        is_schluter_type = mat_type in ("transitions", "waterproofing", "tread_riser")
        has_schluter_name = "schluter" in description

        if not is_schluter_type and not has_schluter_name:
            # For unknown types (no AI classification), require "schluter" in description
            # Do NOT fall through to product line name matching — too many false positives
            continue

        # Try to find the product line and item number in the description
        # Descriptions look like: "Schluter SCHIENE A 100 AE" or "RENO-TK ETK 80"
        best_match = None
        best_score = 0

        for line, pattern in SCHLUTER_PATTERNS.items():
            if pattern.search(description):
                # Found a product line as a whole word — search the price book
                results = match_price_book(line.upper())
                if not results:
                    results = match_price_book(line.upper().replace("-", ""))
                if not results:
                    continue

                # Try to narrow down by item number or size from description
                for pb_item in results:
                    score = 1  # base score for product line match
                    pb_item_lower = pb_item["item_no"].lower()

                    # Check if the item number appears in the description
                    if pb_item_lower and pb_item_lower in description:
                        score += 5  # strong match

                    # Check size match
                    if pb_item["size_mm"] and pb_item["size_mm"] in description:
                        score += 2
                    if pb_item["size_inches"] and pb_item["size_inches"] in description:
                        score += 2

                    # Check material/finish match
                    finish_lower = pb_item["material_finish"].lower()
                    if finish_lower and any(w in description for w in finish_lower.split() if len(w) > 3):
                        score += 1

                    if score > best_score:
                        best_score = score
                        best_match = pb_item

        if best_match and best_score >= 3:
            # Apply price book net price (already discounted)
            # Schluter transitions are sold per PIECE (each piece = 8'-2" = 8.1667 LF)
            import math
            SCHLUTER_PIECE_LF = 8.0 + 2.0 / 12.0  # 8'-2" = 8.1667 LF
            price_per_piece = best_match["net_price"]
            order_qty_lf = mat.get("order_qty", 0)
            fixture_count = mat.get("fixture_count", 0) or 0
            pieces_needed = _calc_schluter_pieces(order_qty_lf, fixture_count, SCHLUTER_PIECE_LF)
            mat["unit_price"] = price_per_piece
            mat["vendor"] = "Schluter"
            mat["quote_status"] = "price_book"
            mat["price_source"] = "price_book"
            mat["extended_cost"] = round(pieces_needed * price_per_piece, 2)
            # Apply labor rate
            is_premium = any(line in description for line in SCHLUTER_PREMIUM_LABOR_LINES)
            mat["labor_rate_lf"] = SCHLUTER_LABOR_RATE_PREMIUM if is_premium else SCHLUTER_LABOR_RATE_DEFAULT
            mat["labor_catalog"] = "Schluter Schiene"
            matched += 1
        elif (is_schluter_type or has_schluter_name) and (not best_match or best_score < 3):
            # No rule-based match or low-confidence match — queue for AI matching
            ai_candidates.append(mat_idx)

    # Phase 2: AI matching for remaining Schluter materials
    if ai_candidates:
        ai_matched = _ai_price_book_match(materials, ai_candidates)
        matched += ai_matched

    return matched


def _ai_price_book_match(materials: list[dict], candidate_indices: list[int]) -> int:
    """Use AI to match Schluter materials to the price book when rule-based matching fails."""
    import json as _json

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")
    provider = get_provider_info(api_key)
    if not provider["available"]:
        return 0

    # Group candidates by detected product line to minimize AI calls
    # Gather all price book entries for relevant product lines
    all_pb_items = search_price_book("", vendor="Schluter")
    if not all_pb_items:
        return 0

    # Build material list for AI
    mat_lines = []
    for i, idx in enumerate(candidate_indices):
        mat = materials[idx]
        desc = mat.get("description", "")
        unit = mat.get("unit", "")
        mat_lines.append(f"M{i}: {desc} (unit: {unit})")

    # Build price book summary grouped by product line
    pb_by_line = {}
    for pb in all_pb_items:
        line = pb["product_line"]
        if line not in pb_by_line:
            pb_by_line[line] = []
        pb_by_line[line].append(pb)

    pb_lines = []
    for line, items in sorted(pb_by_line.items()):
        # Show a sample of items per line to keep prompt reasonable
        samples = items[:8]
        for s in samples:
            pb_lines.append(
                f"  {line} | {s['item_no']} | {s['material_finish']} | "
                f"size: {s.get('size_inches', '')} | ${s['net_price']}/{s.get('unit', 'length')}"
            )
        if len(items) > 8:
            pb_lines.append(f"  ... and {len(items) - 8} more {line} items")

    prompt = f"""Match these Schluter transition materials to the correct product from our Schluter price book.

Materials to match:
{chr(10).join(mat_lines)}

Schluter Price Book:
{chr(10).join(pb_lines)}

Rules:
- Match based on product line (Reno-TK, Reno-Ramp, Dilex, Schiene, etc.), material/finish, and size
- DEFAULT: Unless the description explicitly specifies a different finish or size, always default to AE (satin anodized aluminum) finish and 100 (10mm / 3/8") size. This is standard estimating practice.
- Only use a different finish (ATGB, ATG, AK, etc.) if the description explicitly calls it out in the finish schedule
- If size is not specified, use 100 (10mm)
- If the exact product line is not in the price book, check if a similar or parent product line exists (e.g. DILEX-AHKA may relate to DECO or JOLLY). If nothing similar exists, return no match for that material
- Only match if you are confident the product line and material type are correct

Return JSON: {{"matches": [{{"material": "M0", "item_no": "AETK 80", "product_line": "RENO-TK", "net_price": 9.80, "confidence": 0.9, "reason": "Reno-TK anodized aluminum 5/16 inch"}}]}}
Return {{"matches": []}} for any materials you cannot confidently match."""

    try:
        # Sanitize prompt to avoid encoding issues on Windows
        prompt = prompt.encode("ascii", errors="replace").decode("ascii")

        raw = chat_complete(
            system="You are a Schluter product matching expert for commercial flooring. Return JSON only.",
            user=prompt,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        result = _json.loads(raw)
        matches_raw = result if isinstance(result, list) else result.get("matches", [])
        if not isinstance(matches_raw, list):
            matches_raw = []

        matched = 0
        for m in matches_raw:
            conf = m.get("confidence", 0)
            if conf < 0.7:
                continue
            mat_ref = m.get("material", "")
            try:
                local_idx = int(mat_ref.replace("M", ""))
                mat_idx = candidate_indices[local_idx]
            except (ValueError, IndexError):
                continue

            net_price = m.get("net_price", 0)
            if not net_price or net_price <= 0:
                continue

            mat = materials[mat_idx]
            # Schluter transitions are sold per PIECE (each piece = 8'-2" = 8.1667 LF)
            import math
            SCHLUTER_PIECE_LF = 8.0 + 2.0 / 12.0  # 8'-2" = 8.1667 LF
            price_per_piece = net_price
            order_qty_lf = mat.get("order_qty", 0)
            fixture_count = mat.get("fixture_count", 0) or 0
            pieces_needed = _calc_schluter_pieces(order_qty_lf, fixture_count, SCHLUTER_PIECE_LF)
            mat["unit_price"] = price_per_piece
            mat["vendor"] = "Schluter"
            mat["quote_status"] = "price_book"
            mat["price_source"] = "price_book"
            mat["extended_cost"] = round(pieces_needed * price_per_piece, 2)
            # Apply labor rate
            desc_lower = (mat.get("description") or "").lower()
            is_premium = any(line in desc_lower for line in SCHLUTER_PREMIUM_LABOR_LINES)
            mat["labor_rate_lf"] = SCHLUTER_LABOR_RATE_PREMIUM if is_premium else SCHLUTER_LABOR_RATE_DEFAULT
            mat["labor_catalog"] = "Schluter Schiene"
            matched += 1
            print(f"[price_book] AI matched: {mat.get('description', '')[:60]} -> {m.get('product_line')} {m.get('item_no')} ${net_price}/pc x {pieces_needed}pc")

        return matched
    except Exception as e:
        print(f"[price_book] AI matching error: {e}")
        return 0


def _link_upload_to_requests(job_id: int, products: list[dict]):
    """Detect which open quote requests match uploaded vendor quotes.
    Returns list of matched requests for frontend confirmation."""
    requests = list_quote_requests(job_id)
    if not requests:
        return []

    # Detect vendors and file names from uploaded products
    upload_vendors = {}  # vendor_lower -> {vendor, file_name}
    for p in products:
        v = (p.get("vendor") or "").strip()
        if v:
            upload_vendors[v.lower()] = {
                "vendor": v,
                "file_name": p.get("file_name", ""),
            }

    if not upload_vendors:
        return []

    import re
    def _normalize_vendor(name):
        """Normalize vendor name for fuzzy matching: lowercase, strip punctuation, collapse spaces."""
        return re.sub(r'[^a-z0-9 ]', '', name.lower()).strip()

    matched = []
    for req in requests:
        if req.get("received_at"):
            continue  # already marked received
        req_vendor = _normalize_vendor(req.get("vendor_name") or "")
        for uv_lower, uv_info in upload_vendors.items():
            uv_norm = _normalize_vendor(uv_lower)
            if req_vendor in uv_norm or uv_norm in req_vendor or req_vendor == uv_norm:
                matched.append({
                    "request_id": req["id"],
                    "vendor_name": req.get("vendor_name", ""),
                    "sent_at": req.get("sent_at", ""),
                    "response_file": uv_info.get("file_name", ""),
                })
                break

    return matched


@app.post("/api/jobs/{job_id}/upload-quotes")
async def api_upload_quotes(job_id: str, files: list[UploadFile] = File(...)):
    """Upload vendor quote files, parse them, return pricing."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    import hashlib as _hashlib

    # Ensure AI config is loaded (key may have been added since server start)
    _apply_openai_config()

    all_products = []
    skipped_files = []
    file_errors = []
    for upload in files:
        content = await upload.read(MAX_QUOTE_FILE_BYTES + 1)
        if len(content) > MAX_QUOTE_FILE_BYTES:
            file_errors.append({
                "file": upload.filename,
                "error": f"File exceeds the {MAX_QUOTE_FILE_BYTES // (1024 * 1024)} MB limit.",
            })
            continue

        # Dedup: check file hash before parsing
        file_hash = _hashlib.sha256(content).hexdigest()
        if is_file_imported(db_id, file_hash):
            skipped_files.append(upload.filename)
            continue

        file_path = _job_upload_path(db_id, f"{file_hash[:12]}_{upload.filename}", "quote")
        with open(file_path, "wb") as f:
            f.write(content)
        _record_artifact(db_id, file_path, "vendor_quote")

        try:
            products = parse_quote_file(file_path)
            for p in products:
                p["file_name"] = upload.filename
                p["_source_hash"] = file_hash
            all_products.extend(products)
            if not products:
                file_errors.append({"file": upload.filename, "error": "No pricing products were extracted."})
            # Only record as imported if we actually extracted products
            if products:
                record_imported_file(
                    db_id,
                    upload.filename,
                    file_hash,
                    len(content),
                    source="manual",
                    artifact_path=os.path.relpath(file_path, ARTIFACT_ROOT),
                    artifact_kind="vendor_quote",
                )
        except Exception as e:
            file_errors.append({"error": str(e), "file": upload.filename})

    if not all_products and file_errors:
        detail = "; ".join(f"{item['file']}: {item['error']}" for item in file_errors[:5])
        raise HTTPException(status_code=422, detail=f"No vendor pricing was imported. {detail}")

    # Persist quotes to DB
    save_quotes(db_id, all_products)

    # Auto-match prices to materials
    auto_matched = _auto_match_quotes(db_id, all_products)

    # Save to vendor pricing database
    save_vendor_prices_from_quotes(db_id, all_products)

    # Detect matching quote requests (don't auto-link — frontend will confirm)
    linked_requests = _link_upload_to_requests(db_id, all_products)

    file_names = [u.filename for u in files if hasattr(u, 'filename')]
    vendors_found = list(set(p.get("vendor", "Unknown") for p in all_products if p.get("vendor")))
    log_activity(db_id, "quotes_uploaded", f"Uploaded {len(file_names)} quote file(s), {len(all_products)} products, {auto_matched} auto-matched", {"files": file_names, "vendors": vendors_found, "product_count": len(all_products), "auto_matched": auto_matched})

    refreshed = load_job(db_id) or {}
    return {
        "products": refreshed.get("quotes") or [],
        "parsed_products": all_products,
        "auto_matched": auto_matched,
        "linked_requests": linked_requests,
        "skipped_files": skipped_files,
        "file_errors": file_errors,
    }


@app.get("/api/jobs/{job_id}/imported-files")
def api_imported_files(job_id: str):
    """List all files that have been imported for this job (for dedup)."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return list_imported_files(job["id"])


@app.delete("/api/jobs/{job_id}/quotes")
def api_clear_quotes(job_id: str):
    """Clear all parsed quotes for a job."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    delete_quotes(job["id"])
    log_activity(job["id"], "quotes_cleared", "All vendor quotes cleared")
    return {"message": "Quotes cleared"}


@app.put("/api/quotes/{quote_id}")
def api_update_quote(quote_id: int, body: dict = Body(...)):
    """Update a single quote entry and re-match against materials."""
    job_id = get_quote_job_id(quote_id)
    if not job_id:
        raise HTTPException(status_code=404, detail="Quote not found")
    update_quote(quote_id, body)
    # Re-run auto-match so the updated price flows to materials
    job = load_job(job_id)
    if job:
        quotes = job.get("quotes", [])
        _auto_match_quotes(job_id, quotes)
    log_activity(job_id, "quote_updated", f"Quote #{quote_id} updated")
    return {"ok": True}


# ── Dropbox Scanner Endpoints ────────────────────────────────────────────────

@app.post("/api/jobs/{job_id}/match-dropbox-folder")
def api_match_dropbox_folder(job_id: str, body: dict = Body(...)):
    """Fuzzy-match job project name against a list of folder names from the browser.
    The browser reads the local Dropbox folder via File System Access API and sends folder names here.
    """
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    folder_names = body.get("folder_names", [])
    if not folder_names:
        return {"folder_found": False, "folder_name": None, "score": 0}

    result = match_folder(
        job.get("project_name", ""),
        job.get("gc_name", ""),
        folder_names,
    )

    if not result:
        return {"folder_found": False, "folder_name": None, "score": 0}

    return {"folder_found": True, "folder_name": result["folder_name"], "score": result["score"]}


@app.post("/api/jobs/{job_id}/calculate")
def api_calculate(job_id: str):
    """Run sundry + labor calculators, return results."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    materials = job.get("materials", [])

    # Stamp job-level counts for sundry calculations and Schluter fixture counts
    unit_count = job.get("unit_count", 0) or 0
    tub_shower_count = job.get("tub_shower_count", 0) or 0
    for mat in materials:
        mtype = mat.get("material_type", "")
        desc = (mat.get("description") or "").lower()
        if mtype == "backsplash":
            mat["unit_count"] = unit_count
        if mtype == "tub_shower_surround":
            mat["tub_shower_total"] = tub_shower_count
        # Schluter transitions at tub/shower surrounds get fixture_count from tub_shower_count
        # This enables _calc_schluter_pieces to compute pieces per fixture (2 sides each)
        if "schluter" in desc and ("tub" in desc or "shower" in desc or "surround" in desc or "rr" in desc or "wash" in desc):
            mat["fixture_count"] = tub_shower_count

    # Calculate sundries
    trace = AuditTraceBuilder(job["id"])

    sundries = calculate_sundries_for_materials(materials, trace=trace)
    save_sundries(job["id"], sundries)

    # Calculate labor
    labor_items = calculate_labor_for_materials(materials, trace=trace)
    save_labor(job["id"], labor_items)

    run_id = create_calculation_run(
        job["id"],
        "bid_calculation",
        metadata=_audit_metadata({"endpoint": "calculate"}),
    )
    trace_count = save_calculation_traces(job["id"], run_id, trace.records)
    complete_calculation_run(run_id, summary=trace.summary())

    log_activity(job["id"], "bid_calculated", f"Calculated {len(sundries)} sundries and {len(labor_items)} labor items")

    return {
        "sundries": sundries,
        "labor": labor_items,
        "audit": {"run_id": run_id, "trace_count": trace_count, "summary": trace.summary()},
    }


@app.put("/api/jobs/{job_id}/materials")
def api_update_materials(job_id: str, body: MaterialUpdate):
    """Update materials (edited pricing, waste, etc.)."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    deletion_reasons = {
        str(key): str(value or "").strip()
        for key, value in (body.deletion_reasons or {}).items()
    }

    def deletion_key(material: dict) -> str:
        return str(material.get("item_code") or material.get("id") or "")

    conn = _get_conn()
    deleted_materials = []
    try:
        conn.execute("BEGIN IMMEDIATE")
        current_materials = [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM job_materials WHERE job_id=? ORDER BY id",
                (job["id"],),
            ).fetchall()
        ]
        current_fingerprint = _materials_source_fingerprint(current_materials)
        if (
            body.base_source_fingerprint
            and body.base_source_fingerprint != current_fingerprint
        ):
            raise HTTPException(
                status_code=409,
                detail=(
                    "These materials changed in another tab or workflow. Your stale pricing copy "
                    "was not saved. Reload the job and review the newer values before editing again."
                ),
            )

        existing = {material["id"]: material for material in current_materials}
        incoming_ids = {
            str(material.get("id"))
            for material in body.materials
            if material.get("id") is not None
        }
        deleted_materials = [
            material for material in existing.values()
            if str(material.get("id")) not in incoming_ids
        ]
        missing_deletion_reasons = [
            deletion_key(material)
            for material in deleted_materials
            if not deletion_reasons.get(deletion_key(material))
        ]
        if missing_deletion_reasons:
            raise HTTPException(
                status_code=400,
                detail=(
                    "A reason is required before removing takeoff materials: "
                    f"{', '.join(missing_deletion_reasons[:5])}."
                ),
            )

        updated = []
        for material in body.materials:
            base = existing.get(material.get("id"), {})
            merged = {**base, **{key: value for key, value in material.items() if value is not None}}
            if material.get("material_type") and base.get("material_type") != material.get("material_type"):
                merged["ai_confidence"] = 1.0

            waste_pct = merged.get("waste_pct", 0)
            installed_qty = merged.get("installed_qty", 0)
            unit_price = merged.get("unit_price", 0)
            order_qty = (
                material["order_qty"]
                if "order_qty" in material and material["order_qty"] is not None
                else installed_qty * (1 + waste_pct)
            )

            if material.get("price_source") == "manual" and material.get("extended_cost") is not None:
                extended_cost = material["extended_cost"]
            elif (
                merged.get("price_source") in ("price_book", "default_rule")
                and (merged.get("material_type") or "").lower() == "transitions"
            ):
                import math
                vendor = (merged.get("vendor") or "").lower()
                stick_lf = 12.0 if "silver pin" in vendor else 8.0 + 2.0 / 12.0
                fixture_count = merged.get("fixture_count", 0) or 0
                pieces = (
                    _calc_schluter_pieces(order_qty, fixture_count, stick_lf)
                    if vendor == "schluter"
                    else math.ceil(order_qty / stick_lf) if order_qty > 0 else 0
                )
                extended_cost = pieces * unit_price
            else:
                extended_cost = order_qty * unit_price

            merged["order_qty"] = round(order_qty, 2)
            merged["extended_cost"] = round(extended_cost, 2)
            updated.append(merged)

        material_ids = save_materials(job["id"], updated, conn=conn)
        for material, material_id in zip(updated, material_ids):
            material["id"] = material_id
        persisted_materials = [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM job_materials WHERE job_id=? ORDER BY id",
                (job["id"],),
            ).fetchall()
        ]
        source_fingerprint = _materials_source_fingerprint(persisted_materials)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # Keep read-only price suggestions visible after an autosave without
    # including them in the durable source fingerprint.
    response_job = {"materials": persisted_materials}
    _enrich_known_prices(response_job)
    response_materials = response_job["materials"]

    if deleted_materials:
        log_activity(
            job["id"],
            "materials_deleted",
            f"Removed {len(deleted_materials)} takeoff material(s) with estimator reasons",
            {
                "deletions": [
                    {
                        "material_key": deletion_key(material),
                        "description": material.get("description"),
                        "reason": deletion_reasons.get(deletion_key(material)),
                    }
                    for material in deleted_materials
                ],
            },
        )
    log_activity(job["id"], "materials_updated", f"Updated pricing for {len(body.materials)} materials")

    return {
        "materials": response_materials,
        "materials_source_fingerprint": source_fingerprint,
    }


@app.post("/api/jobs/{job_id}/materials/{material_idx}/estimate-price")
def api_estimate_price(job_id: str, material_idx: int):
    """Use AI to estimate a material's unit price based on its description."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    materials = job.get("materials", [])
    if material_idx < 0 or material_idx >= len(materials):
        raise HTTPException(status_code=404, detail="Material not found")

    m = materials[material_idx]
    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if not provider["available"]:
        raise HTTPException(status_code=400, detail="No AI API key configured (set OpenAI or ANTHROPIC_API_KEY)")

    import json as _json

    # Check price history and price list for existing data to inform the estimate
    history = get_price_history(item_code=m.get("item_code"), product=m.get("description"))
    price_list = get_price_list_entries()
    price_list_match = None
    item_code_lower = (m.get("item_code") or "").lower()
    desc_lower = (m.get("description") or "").lower()
    for entry in price_list:
        ename = (entry.get("product_name") or "").lower()
        if item_code_lower and item_code_lower in ename:
            price_list_match = entry
            break
        if desc_lower and ename and ename in desc_lower:
            price_list_match = entry
            break

    # Build context for AI
    context_lines = []
    if history.get("records"):
        context_lines.append(f"Price history: min=${history['min']}, max=${history['max']}, avg=${history['avg']}")
        latest = history["latest"]
        if latest:
            context_lines.append(f"Latest quote: ${latest.get('unit_price')} from {latest.get('vendor_name', 'unknown')} ({latest.get('created_at', '')})")
    if price_list_match:
        context_lines.append(f"Internal price list: ${price_list_match.get('unit_price')} per {price_list_match.get('unit', 'unit')}")

    history_context = "\n".join(context_lines) if context_lines else "No historical pricing data available."

    prompt = f"""Estimate the unit price for this flooring/interior material.
Return JSON: {{"estimated_price": <number>, "confidence": <0-1>, "reasoning": "<brief>"}}

Material: {m.get('description', '')}
Item Code: {m.get('item_code', '')}
Type: {m.get('material_type', '')}
Unit: {m.get('unit', '')}
Order Qty: {m.get('order_qty', 0)}

Historical Data:
{history_context}

If historical data is available, weight it heavily in your estimate. Otherwise, base your estimate on typical commercial flooring/interior material pricing.
The price should be per {m.get('unit', 'unit')}. Be conservative — estimate on the higher side."""

    try:
        raw = chat_complete(
            system="You are a commercial flooring estimator. Return only valid JSON.",
            user=prompt,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        result = _json.loads(raw)
        estimated_price = float(result.get("estimated_price", 0))
        confidence = float(result.get("confidence", 0.5))
        reasoning = result.get("reasoning", "")

        # Update the material with the AI estimate
        m["unit_price"] = round(estimated_price, 2)
        m["price_source"] = "ai_estimate"
        m["order_qty"] = round(m.get("installed_qty", 0) * (1 + m.get("waste_pct", 0)), 2)
        m["extended_cost"] = round(m["order_qty"] * m["unit_price"], 2)
        materials[material_idx] = m
        save_materials(job["id"], materials)

        log_activity(job["id"], "ai_estimate",
                     f"AI estimated price for {m.get('item_code', m.get('description', 'material'))}: ${estimated_price:.2f}/{m.get('unit', 'unit')}")

        return {
            "estimated_price": estimated_price,
            "confidence": confidence,
            "reasoning": reasoning,
            "material": m,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI estimation failed: {e}")


@app.post("/api/jobs/{job_id}/generate-bid")
def api_generate_bid(job_id: str):
    """Assemble bid + generate PDF, return bid data."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _validate_bid_job_ready(job)

    materials = job.get("materials", [])
    sundries = job.get("sundries", [])
    labor_items = job.get("labor", [])

    # Build job_info dict for assembler
    job_info = {
        "id": job["id"],
        "project_name": job["project_name"],
        "gc_name": job.get("gc_name"),
        "address": job.get("address"),
        "city": job.get("city"),
        "state": job.get("state"),
        "zip": job.get("zip"),
        "tax_rate": job.get("tax_rate", 0),
        "unit_count": job.get("unit_count", 0),
        "salesperson": job.get("salesperson"),
        "markup_pct": job.get("markup_pct", 0),
        "gpm_pct": job.get("gpm_pct", 0),
    }

    # Parse job-specific exclusions (stored as JSON string)
    import json as _json
    custom_exclusions = None
    raw_exclusions = job.get("exclusions")
    if raw_exclusions:
        try:
            custom_exclusions = _json.loads(raw_exclusions)
        except (ValueError, TypeError):
            pass

    bid_data = assemble_bid(job_info, materials, sundries, labor_items, exclusions=custom_exclusions)
    bid_audit = _record_bid_audit(job["id"], bid_data)

    # Save bundles
    save_bundles(job["id"], bid_data["bundles"])

    # Persist full bid data to job record (bundles + totals)
    bid_persist = {
        "bundles": bid_data["bundles"],
        "subtotal": bid_data["subtotal"],
        "markup_pct": bid_data["markup_pct"],
        "markup_amount": bid_data["markup_amount"],
        "gpm_pct": bid_data.get("gpm_pct", 0),
        "gpm_profit": bid_data.get("gpm_profit", 0),
        "total_cost": bid_data.get("total_cost", 0),
        "tax_rate": bid_data["tax_rate"],
        "tax_amount": bid_data["tax_amount"],
        "grand_total": bid_data["grand_total"],
        "exclusions": bid_data.get("exclusions", []),
        "audit": {
            "run_id": bid_audit["run"]["id"],
            "trace_count": bid_audit["trace_count"],
            "summary": bid_audit["run"].get("summary", {}),
            "ruleset_version": bid_audit["run"].get("metadata", {}).get("ruleset_version"),
        },
        "pdf_audit_run_id": bid_audit["run"]["id"],
        "pdf_ruleset_version": bid_audit["run"].get("metadata", {}).get("ruleset_version"),
        "pdf_source_fingerprint": _bid_source_fingerprint(job),
        "pdf_totals": {
            "subtotal": bid_data["subtotal"],
            "tax_amount": bid_data["tax_amount"],
            "grand_total": bid_data["grand_total"],
            "total_cost": bid_data.get("total_cost", 0),
            "gpm_profit": bid_data.get("gpm_profit", 0),
            "markup_amount": bid_data.get("markup_amount", 0),
        },
    }
    job["bid_data"] = _json.dumps(bid_persist)
    save_job(job)

    # Generate PDF
    pdf_path = _job_pdf_path(job["id"], "bid")
    generate_bid_pdf(bid_data, pdf_path)
    _record_artifact(job["id"], pdf_path, "bid_pdf")

    bundle_count = len(bid_data.get("bundles", []))
    grand_total = bid_data.get("grand_total", 0)
    log_activity(job["id"], "bid_generated", f"Bid generated: {bundle_count} bundles, total ${grand_total:,.2f}", {"bundle_count": bundle_count, "grand_total": grand_total})

    bid_data["audit"] = bid_persist["audit"]
    return bid_data


@app.delete("/api/jobs/{job_id}/bid")
def api_clear_bid(job_id: str):
    """Clear saved bid data for a job."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job["bid_data"] = None
    save_job(job)
    # Clear bundles too
    save_bundles(job["id"], [])
    log_activity(job["id"], "bid_cleared", "Bid data cleared")
    return {"message": "Bid cleared"}


@app.get("/api/jobs/{job_id}/bid.pdf")
def api_download_bid_pdf(job_id: str):
    """Download the generated bid PDF."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _validate_bid_pdf_download_ready(job)
    pdf_path = _job_pdf_path(job["id"], "bid")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="PDF not found. Generate bid first.")
    _require_artifact_receipt(job["id"], pdf_path, "bid_pdf")
    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=f"bid_{job_id}.pdf",
    )


# ── Proposal Endpoints ─────────────────────────────────────────────────────

@app.post("/api/jobs/{job_id}/proposal/rewrite-descriptions")
async def api_rewrite_descriptions(job_id: str, request: Request):
    """Use AI to rewrite bundle descriptions in professional proposal style."""
    from description_agent import rewrite_bundle_descriptions

    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    body = await request.json()
    bundles = body.get("bundles", [])
    if not bundles:
        raise HTTPException(status_code=400, detail="No bundles provided")

    descriptions = rewrite_bundle_descriptions(bundles, job)
    return {"descriptions": descriptions}


@app.get("/api/jobs/{job_id}/proposal/bundles")
def api_get_proposal_bundles(job_id: str):
    """Load saved proposal editor state (bundles, notes, terms, etc.)."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    pd = job.get("proposal_data")
    if pd and isinstance(pd, dict) and pd.get("bundles"):
        return pd
    return {"bundles": [], "notes": [], "terms": [], "exclusions": []}


# ── Calculation Audit Endpoints ─────────────────────────────────────────────

@app.get("/api/jobs/{job_id}/audit/runs")
def api_get_calculation_runs(job_id: str, limit: int = 20):
    """List calculation audit runs for a job."""
    db_id = _resolve_job_id(job_id)
    return {"runs": list_calculation_runs(db_id, limit=limit)}


@app.get("/api/jobs/{job_id}/audit")
def api_get_latest_calculation_audit(job_id: str, limit: int = 1000, scope: Optional[str] = None):
    """Fetch the latest audit, optionally limited to accepted-proposal work."""
    db_id = _resolve_job_id(job_id)
    if scope not in (None, "", "proposal"):
        raise HTTPException(status_code=400, detail="Unsupported audit scope")
    if scope == "proposal":
        run = get_latest_completed_calculation_run(
            db_id,
            {"proposal_editor_save", "proposal_generation", "bid_calculation"},
        )
    else:
        runs = list_calculation_runs(db_id, limit=1)
        run = runs[0] if runs else None
    traces = []
    if run:
        traces = get_calculation_traces(db_id, run_id=run["id"], limit=limit)
        if run.get("run_type") in ("proposal_manual_save", "proposal_editor_save"):
            prior = get_latest_completed_calculation_run(
                db_id,
                {"proposal_generation", "bid_calculation"},
            )
            if prior:
                remaining = max(limit - len(traces), 0)
                if remaining:
                    traces.extend(get_calculation_traces(db_id, run_id=prior["id"], limit=remaining))
    return {
        "run": run,
        "traces": traces,
        "events": traces,
        "audit": run.get("summary", {}) if run else {},
    }


@app.get("/api/jobs/{job_id}/audit/trace")
def api_get_calculation_trace(
    job_id: str,
    run_id: Optional[int] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    entity_key: Optional[str] = None,
    limit: int = 1000,
):
    """Fetch calculation trace rows. Defaults to the latest run for the job."""
    db_id = _resolve_job_id(job_id)
    selected_run_id = run_id
    run = None
    if selected_run_id is None:
        runs = list_calculation_runs(db_id, limit=10)
        if runs:
            run = runs[0]
            selected_run_id = run["id"]
    elif selected_run_id is not None:
        runs = [r for r in list_calculation_runs(db_id, limit=100) if r["id"] == selected_run_id]
        run = runs[0] if runs else None

    traces = []
    if selected_run_id is not None:
        traces = get_calculation_traces(
            db_id,
            run_id=selected_run_id,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_key=entity_key,
            limit=limit,
        )
        if run and run.get("run_type") in ("proposal_manual_save", "proposal_editor_save") and not any([run_id, entity_type, entity_id, entity_key]):
            prior = get_latest_completed_calculation_run(
                db_id,
                {"proposal_generation", "bid_calculation"},
            )
            if prior:
                remaining = max(limit - len(traces), 0)
                if remaining:
                    traces.extend(get_calculation_traces(db_id, run_id=prior["id"], limit=remaining))
    return {"run": run, "traces": traces}


@app.get("/api/jobs/{job_id}/audit/runs/{run_id}/trace")
def api_get_calculation_run_trace(
    job_id: str,
    run_id: int,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    entity_key: Optional[str] = None,
    limit: int = 1000,
):
    """Fetch calculation trace rows for a specific run."""
    db_id = _resolve_job_id(job_id)
    traces = get_calculation_traces(
        db_id,
        run_id=run_id,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_key=entity_key,
        limit=limit,
    )
    return {"run_id": run_id, "traces": traces}


@app.post("/api/rules/audit-harness")
def api_rules_audit_harness_probe(body: Optional[dict] = Body(default=None)):
    """Small UI probe for rules/audit visibility; full harness lives in scripts/."""
    body = body or {}
    job_ref = body.get("job_id")
    stage = body.get("stage")
    category = body.get("category")
    field = body.get("field")
    rules = get_active_rules(stage=stage if stage not in ("", "all") else None,
                             category=category if category not in ("", "all") else None)
    response = {
        "status": "ok",
        "rule_count": len(rules),
        "rules": rules[:25],
        "note": "Full deployed harness: scripts/rules_audit_harness.py --base-url <fly-url>",
    }
    if field:
        response["field"] = field
    if job_ref:
        db_id = _resolve_job_id(str(job_ref))
        runs = list_calculation_runs(db_id, limit=1)
        response["latest_run"] = runs[0] if runs else None
        response["trace_count"] = 0
        if runs:
            traces = get_calculation_traces(db_id, run_id=runs[0]["id"], limit=1000)
            response["trace_count"] = len(traces)
            matching = [t for t in traces if not field or t.get("output_field") == field]
            response["field_trace"] = matching[-1] if matching else None
            response["field_trace_count"] = len(matching)
            response["sample_traces"] = (matching or traces)[:10]
            response["summary"] = {
                "run_type": runs[0].get("run_type"),
                "field": field,
                "field_found": bool(matching),
                "formula": (matching[-1].get("formula") if matching else None),
                "result": (matching[-1].get("result") if matching else None),
                "source": (matching[-1].get("source") if matching else None),
                "rule_id": (matching[-1].get("rule_id") if matching else None),
            }
    return response


def _public_golden_job(golden: dict | None) -> dict | None:
    if not golden:
        return None
    snapshot = golden.get("snapshot") or {}
    return {
        "id": golden.get("id"),
        "version_id": golden.get("version_id"),
        "version_number": golden.get("version_number"),
        "immutable": bool(golden.get("immutable")),
        "source_job_id": golden.get("source_job_id"),
        "name": golden.get("name"),
        "jr_quote_id": golden.get("jr_quote_id"),
        "target_totals": snapshot.get("target_totals") or golden.get("target_totals") or {},
        "accepted_totals": snapshot.get("accepted_totals") or {},
        "tolerance": golden.get("tolerance") or snapshot.get("tolerance") or DEFAULT_TOLERANCE,
        "ruleset_version": golden.get("ruleset_version"),
        "source_fingerprint": golden.get("source_fingerprint"),
        "engine_fingerprint": golden.get("engine_fingerprint") or (snapshot.get("build") or {}).get("engine_fingerprint"),
        "artifact_manifest": golden.get("artifact_manifest") or snapshot.get("artifact_manifest") or [],
        "reviewer_name": golden.get("reviewer_name", ""),
        "notes": golden.get("notes"),
        "status": golden.get("status"),
        "created_at": golden.get("created_at"),
        "updated_at": golden.get("updated_at"),
    }


def _public_replay(replay: dict | None, include_generated: bool = False) -> dict | None:
    if not replay:
        return None
    result = {
        "id": replay.get("id"),
        "golden_job_id": replay.get("golden_job_id"),
        "golden_version_id": replay.get("golden_version_id"),
        "source_job_id": replay.get("source_job_id"),
        "mode": replay.get("mode"),
        "status": replay.get("status"),
        "summary": replay.get("summary") or {},
        "diff": replay.get("diff") or {},
        "audit_run_id": replay.get("audit_run_id"),
        "created_at": replay.get("created_at"),
    }
    if include_generated:
        result["generated_proposal"] = replay.get("generated_proposal") or {}
    return result


def _golden_replay_statuses(
    job_id: int,
    current_source_fingerprint: str | None = None,
    current_engine_fingerprint: str | None = None,
) -> dict:
    golden = get_golden_job_for_source(job_id)
    if not golden:
        return {
            "overall": None,
            "verification": None,
            "current": None,
            "current_drift_classification": None,
            "source_matches": None,
        }
    active_version_id = golden.get("version_id")
    baseline = get_latest_golden_replay_for_version(job_id, active_version_id, "baseline") if active_version_id else None
    current = get_latest_golden_replay_for_version(job_id, active_version_id, "current") if active_version_id else None
    if not baseline:
        verification = "not_replayed"
    else:
        summary = baseline.get("summary") or {}
        if baseline.get("status") == "pass" and summary.get("raw_engine_status", summary.get("engine_status")) == "pass":
            verification = "golden_verified"
        elif baseline.get("status") == "incomparable" or summary.get("status") == "incomparable":
            verification = "incomparable"
        else:
            verification = "fail"

    baseline_source_fingerprint = ((golden.get("snapshot") or {}).get("proposal_source_fingerprint") or "").strip()
    baseline_engine_fingerprint = (
        golden.get("engine_fingerprint")
        or ((golden.get("snapshot") or {}).get("build") or {}).get("engine_fingerprint")
        or ""
    ).strip()
    source_matches = None
    if baseline_source_fingerprint and current_source_fingerprint:
        source_matches = baseline_source_fingerprint == current_source_fingerprint
        if not source_matches:
            verification = "stale"
    engine_incomparable = (
        not baseline_engine_fingerprint
        or (current_engine_fingerprint and baseline_engine_fingerprint != current_engine_fingerprint)
    )
    if verification in ("golden_verified", "stale") and engine_incomparable:
        verification = "incomparable"
    elif verification == "golden_verified" and not baseline_source_fingerprint:
        verification = "incomparable"

    current_status = current.get("status") if current else None
    current_drift_classification = ((current or {}).get("summary") or {}).get("drift_classification")
    if verification == "stale":
        overall = "drift"
    elif verification != "golden_verified":
        overall = verification
    elif current_status == "warn" and current_drift_classification == "metadata_only":
        overall = "metadata_changed"
    elif current_status in ("warn", "fail", "incomparable"):
        overall = "drift"
    else:
        overall = "golden_verified"
    return {
        "overall": overall,
        "verification": verification,
        "current": current_status,
        "current_drift_classification": current_drift_classification,
        "source_matches": source_matches,
    }


def _golden_readiness_status(job_id: int) -> str | None:
    """Backward-compatible combined golden status for job list consumers."""
    job = load_job(job_id)
    proposal = (job or {}).get("proposal_data") if isinstance((job or {}).get("proposal_data"), dict) else {}
    fingerprint = _proposal_source_fingerprint(job, proposal) if job else None
    return _golden_replay_statuses(job_id, fingerprint, get_build_info().get("engine_fingerprint"))["overall"]


def _active_golden_replays(job_id: int, golden: dict | None, limit: int = 50) -> list[dict]:
    """Return replay evidence only for the currently active immutable version."""
    if not golden:
        return []
    active_version_id = golden.get("version_id")
    if not active_version_id:
        return []
    return list_golden_replays_for_version(job_id, active_version_id, limit=limit)


def _readiness_trust_summary(
    job: dict,
    *,
    proposal: dict,
    latest_run: dict | None,
    build: dict,
    ruleset_version: int | None,
) -> dict:
    golden = get_golden_job_for_source(job["id"])
    replays = _active_golden_replays(job["id"], golden, limit=50)
    active_version_id = (golden or {}).get("version_id")
    current_replay = get_latest_golden_replay_for_version(job["id"], active_version_id, "current") if active_version_id else None
    latest_replay = current_replay or (replays[0] if replays else None)

    deleted_reasons = proposal.get("deleted_material_reasons")
    if not isinstance(deleted_reasons, dict):
        deleted_reasons = {}
    deleted_codes = {
        str(code)
        for code in (proposal.get("deleted_material_codes") or [])
        if code and str(deleted_reasons.get(str(code)) or "").strip()
    }
    active_materials = [
        material for material in (job.get("materials") or [])
        if isinstance(material, dict) and _job_material_key(material) not in deleted_codes
    ]
    unknown_count = sum(
        1 for material in active_materials
        if not is_valid_material_classification(material.get("material_type"))
    )
    low_confidence_count = 0
    for material in active_materials:
        confidence = _as_number(material.get("ai_confidence"))
        price_source = str(material.get("price_source") or "").strip().lower()
        if (
            (confidence is not None and confidence < 0.75)
            or price_source in {"ai_estimate", "vendor_history"}
            or (_as_number(material.get("unit_price")) or 0) > 0 and not price_source
        ):
            low_confidence_count += 1

    manual_override_count = sum(
        1 for material in active_materials
        if str(material.get("price_source") or "").strip().lower() == "manual"
    )
    for bundle in (proposal.get("bundles") or []):
        if not isinstance(bundle, dict):
            continue
        manual_override_count += int(bundle.get("price_override") is not None)
        manual_override_count += int(bundle.get("freight_override") is not None)
        manual_override_count += sum(
            1 for material in (bundle.get("materials") or [])
            if isinstance(material, dict) and material.get("freight_is_manual")
        )
        manual_override_count += sum(
            1 for sundry in (bundle.get("sundry_items") or [])
            if isinstance(sundry, dict) and sundry.get("is_manual_price")
        )
        manual_override_count += sum(
            1 for labor in (bundle.get("labor_items") or [])
            if isinstance(labor, dict) and (labor.get("is_manual") or labor.get("is_stair_labor"))
        )
        manual_override_count += len(bundle.get("deleted_labor_keys") or [])

    artifact_receipts = list_job_artifacts(job["id"])
    pdf_receipt = next(
        (item for item in artifact_receipts if item.get("artifact_kind") == "proposal_pdf"),
        None,
    )
    replay_summary = (latest_replay or {}).get("summary") or {}
    replay_diff = (latest_replay or {}).get("diff") or {}
    largest_deltas = [
        {
            "bundle_name": row.get("bundle_name"),
            "delta": row.get("delta"),
            "status": row.get("status"),
        }
        for row in sorted(
            [row for row in (replay_diff.get("bundles") or []) if isinstance(row, dict)],
            key=lambda row: abs(_as_number(row.get("delta")) or 0),
            reverse=True,
        )[:3]
    ]
    target_totals = ((golden or {}).get("snapshot") or {}).get("target_totals") or (golden or {}).get("target_totals") or {}

    return {
        "last_audit_at": (latest_run or {}).get("completed_at") or (latest_run or {}).get("started_at"),
        "last_pdf_at": proposal.get("pdf_generated_at") or (pdf_receipt or {}).get("created_at"),
        "ruleset_version": ruleset_version,
        "build_commit": build.get("commit"),
        "build_tag": build.get("tag"),
        "engine_fingerprint": build.get("engine_fingerprint"),
        "config_fingerprint": build.get("config_fingerprint"),
        "golden_baseline_version": (golden or {}).get("version_number"),
        "jr_target_total": _as_money_number(target_totals.get("grand_total")),
        "accepted_proposal_total": _as_money_number(proposal.get("grand_total")),
        "replay_total": _as_money_number((replay_summary.get("generated_totals") or {}).get("grand_total")),
        "replay_mode": (latest_replay or {}).get("mode"),
        "replay_status": (latest_replay or {}).get("status"),
        "largest_deltas": largest_deltas,
        "manual_override_count": manual_override_count,
        "unknown_material_count": unknown_count,
        "low_confidence_material_count": low_confidence_count,
        "artifact_count": len(artifact_receipts),
        "source_fingerprint": proposal.get("audit_source_fingerprint"),
    }


def _evaluate_job_readiness(job: dict) -> dict:
    proposal = job.get("proposal_data") if isinstance(job.get("proposal_data"), dict) else {}
    latest_run = _latest_completed_run(job["id"], {"proposal_editor_save", "proposal_generation"})
    pdf_ready = False
    pdf_message = None
    pdf_path = _job_pdf_path(job["id"], "proposal")
    if not os.path.exists(pdf_path):
        pdf_message = "The current proposal PDF is missing. Generate the proposal PDF first."
    else:
        try:
            _validate_proposal_pdf_download_ready(job)
            pdf_ready = True
        except HTTPException as exc:
            pdf_message = str(exc.detail)

    artifact_status, artifact_message, artifact_items = _artifact_readiness(
        job["id"],
        pdf_path,
        job.get("materials") or [],
    )
    proposal_source_ok = bool(proposal.get("bundles"))
    proposal_source_message = None
    if proposal_source_ok:
        try:
            _validate_proposal_body_matches_job_source(job, proposal)
        except HTTPException as exc:
            proposal_source_ok = False
            proposal_source_message = str(exc.detail)
    else:
        proposal_source_message = "No saved proposal exists to compare with the current material source."

    proposal_fingerprint = _proposal_source_fingerprint(job, proposal)
    build = get_build_info()
    ruleset_version = _current_ruleset_version()
    golden_statuses = _golden_replay_statuses(
        job["id"],
        proposal_fingerprint,
        build.get("engine_fingerprint"),
    )
    return evaluate_job_readiness(
        job,
        latest_run=latest_run,
        current_ruleset_version=ruleset_version,
        pdf_ready=pdf_ready,
        pdf_message=pdf_message,
        proposal_source_fingerprint=proposal_fingerprint,
        proposal_source_ok=proposal_source_ok,
        proposal_source_message=proposal_source_message,
        artifact_status=artifact_status,
        artifact_message=artifact_message,
        artifact_items=artifact_items,
        golden_status=golden_statuses["overall"],
        golden_verification_status=golden_statuses["verification"],
        current_replay_status=golden_statuses["current"],
        current_replay_drift_classification=golden_statuses["current_drift_classification"],
        build=build,
        trust_summary=_readiness_trust_summary(
            job,
            proposal=proposal,
            latest_run=latest_run,
            build=build,
            ruleset_version=ruleset_version,
        ),
    )


@app.get("/api/jobs/{job_id}/readiness")
def api_get_job_readiness(job_id: str):
    """Return non-mutating readiness checks for an estimator's current bid."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return _evaluate_job_readiness(job)


@app.get("/api/jobs/{job_id}/reproducibility")
def api_get_job_reproducibility(job_id: str):
    """Return golden baseline and recent replay status for a job."""
    db_id = _resolve_job_id(job_id)
    golden = get_golden_job_for_source(db_id)
    replays = list_golden_replays_for_job(db_id, limit=50)
    active_replays = _active_golden_replays(db_id, golden, limit=50)
    latest_active = active_replays[0] if active_replays else None
    active_version_id = (golden or {}).get("version_id")
    baseline_replay = get_latest_golden_replay_for_version(db_id, active_version_id, "baseline") if active_version_id else None
    current_replay = get_latest_golden_replay_for_version(db_id, active_version_id, "current") if active_version_id else None
    job = load_job(db_id)
    proposal = (job or {}).get("proposal_data") if isinstance((job or {}).get("proposal_data"), dict) else {}
    source_fingerprint = _proposal_source_fingerprint(job, proposal) if job else None
    statuses = _golden_replay_statuses(
        db_id,
        source_fingerprint,
        get_build_info().get("engine_fingerprint"),
    )
    return {
        "golden_job": _public_golden_job(golden),
        "latest_replay": _public_replay(latest_active),
        "baseline_replay": _public_replay(baseline_replay),
        "current_replay": _public_replay(current_replay),
        "comparison_status": statuses["overall"] or ("not_replayed" if golden else "not_captured"),
        "current_drift_classification": statuses["current_drift_classification"],
        "live_source_matches_baseline": statuses["source_matches"],
        "replays": [_public_replay(replay) for replay in replays[:10]],
        "active_version_replays": [_public_replay(replay) for replay in active_replays[:10]],
        "default_tolerance": DEFAULT_TOLERANCE,
    }


@app.post("/api/jobs/{job_id}/reproducibility/baseline")
def api_capture_golden_baseline(job_id: str, body: GoldenBaselineRequest):
    """Capture the current accepted proposal as this job's golden baseline."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if not str(body.jr_quote_id or "").strip():
        raise HTTPException(status_code=400, detail="Enter the Job Runner quote ID before capturing the golden baseline.")
    if not str(body.reviewer_name or "").strip():
        raise HTTPException(status_code=400, detail="Enter the reviewer name before capturing the golden baseline.")

    missing = _required_job_field_gaps(job)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot capture golden baseline until required job fields are filled: {', '.join(missing)}.",
        )

    proposal_data = job.get("proposal_data")
    if not isinstance(proposal_data, dict) or not proposal_data.get("bundles"):
        raise HTTPException(status_code=409, detail="Cannot capture golden baseline until the proposal has saved bundles.")
    nonpositive_bundles = [
        bundle.get("bundle_name") or "bundle"
        for bundle in proposal_data.get("bundles") or []
        if isinstance(bundle, dict) and (effective_bundle_total(bundle) <= 0)
    ]
    if nonpositive_bundles:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot capture golden baseline until every bundle has a positive accepted price: {', '.join(nonpositive_bundles[:5])}.",
        )
    arithmetic_errors = proposal_math_errors(proposal_data)
    if arithmetic_errors:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot capture golden baseline because proposal arithmetic is invalid: {' '.join(arithmetic_errors[:3])}",
        )
    try:
        _validate_proposal_body_matches_job_source(job, proposal_data)
    except HTTPException as exc:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot capture golden baseline because the accepted proposal is stale. {exc.detail}",
        ) from exc

    current_proposal_fingerprint = _proposal_source_fingerprint(job, proposal_data)
    if proposal_data.get("audit_source_fingerprint") != current_proposal_fingerprint:
        raise HTTPException(status_code=409, detail="Cannot capture golden baseline until the saved proposal audit matches the current job source. Save or regenerate the proposal first.")

    raw_deleted_codes = {str(code) for code in (proposal_data.get("deleted_material_codes") or []) if code}
    deleted_reasons = proposal_data.get("deleted_material_reasons")
    if not isinstance(deleted_reasons, dict):
        deleted_reasons = {}
    deleted_codes = {
        code for code in raw_deleted_codes
        if str(deleted_reasons.get(code) or "").strip()
    }
    missing_deletion_reasons = sorted(raw_deleted_codes - deleted_codes)
    if missing_deletion_reasons:
        raise HTTPException(
            status_code=409,
            detail="Cannot capture golden baseline until every deleted material has an explicit estimator reason.",
        )
    deleted_bundle_names = {str(name) for name in (proposal_data.get("deleted_bundles") or []) if name}
    deleted_bundle_reasons = proposal_data.get("deleted_bundle_reasons")
    if not isinstance(deleted_bundle_reasons, dict):
        deleted_bundle_reasons = {}
    if any(not str(deleted_bundle_reasons.get(name) or "").strip() for name in deleted_bundle_names):
        raise HTTPException(
            status_code=409,
            detail="Cannot capture golden baseline until every deleted bundle has an explicit estimator reason.",
        )
    accepted_material_codes = {
        _job_material_key(material)
        for bundle in (proposal_data.get("bundles") or [])
        if isinstance(bundle, dict)
        for material in (bundle.get("materials") or [])
        if isinstance(material, dict) and _job_material_key(material)
    }
    contradictory_deleted_codes = sorted(raw_deleted_codes & accepted_material_codes)
    if contradictory_deleted_codes:
        raise HTTPException(
            status_code=409,
            detail="Cannot capture golden baseline because a material is both deleted and still present in an accepted bundle.",
        )
    for bundle in proposal_data.get("bundles") or []:
        if not isinstance(bundle, dict):
            continue
        deleted_labor_reasons = bundle.get("deleted_labor_reasons")
        if not isinstance(deleted_labor_reasons, dict):
            deleted_labor_reasons = {}
        if any(not str(deleted_labor_reasons.get(key) or "").strip() for key in (bundle.get("deleted_labor_keys") or [])):
            raise HTTPException(
                status_code=409,
                detail="Cannot capture golden baseline until every deleted labor line has an explicit estimator reason.",
            )
    active_materials = [
        material for material in (job.get("materials") or [])
        if isinstance(material, dict) and _job_material_key(material) not in deleted_codes
    ]
    missing_from_proposal = [
        material.get("item_code") or material.get("description") or "material"
        for material in active_materials
        if not _job_material_key(material)
        or _job_material_key(material) not in accepted_material_codes
    ]
    if missing_from_proposal:
        raise HTTPException(
            status_code=409,
            detail=(
                "Cannot capture golden baseline until every active material is represented "
                f"in the accepted proposal: {', '.join(str(item) for item in missing_from_proposal[:5])}."
            ),
        )
    unknown = [
        m.get("item_code") or m.get("description") or "material"
        for m in active_materials
        if not is_valid_material_classification(m.get("material_type"))
    ]
    unpriced = [m.get("item_code") or m.get("description") or "material" for m in active_materials if _as_number(m.get("unit_price")) is None or _as_number(m.get("unit_price")) <= 0]
    if unknown or unpriced:
        details = []
        if unknown:
            details.append(f"unknown classifications: {', '.join(str(item) for item in unknown[:5])}")
        if unpriced:
            details.append(f"unpriced materials: {', '.join(str(item) for item in unpriced[:5])}")
        raise HTTPException(status_code=409, detail=f"Cannot capture golden baseline until the proposal is clean ({'; '.join(details)}).")

    run = _latest_completed_run(job["id"], {"proposal_manual_save", "proposal_editor_save", "proposal_generation"})
    if not run:
        raise HTTPException(status_code=409, detail="Cannot capture golden baseline until this proposal has a current audit trace. Save or regenerate first.")
    proposal_audit = proposal_data.get("audit") if isinstance(proposal_data.get("audit"), dict) else {}
    try:
        audit_receipt_matches = int(proposal_audit.get("run_id")) == int(run["id"])
    except (TypeError, ValueError):
        audit_receipt_matches = False
    if not audit_receipt_matches:
        raise HTTPException(
            status_code=409,
            detail="Cannot capture golden baseline because the proposal audit receipt is stale. Save or regenerate the proposal first.",
        )
    _ensure_audit_calculator_current(run, label="Golden baseline")

    target_totals = {}
    for key, value in (body.target_totals or {}).items():
        if value in (None, ""):
            continue
        number = _as_money_number(value)
        if number is None or number < 0:
            raise HTTPException(status_code=400, detail=f"Job Runner total '{key}' must be a non-negative number.")
        target_totals[key] = number
    if target_totals.get("grand_total") is None or target_totals["grand_total"] <= 0:
        raise HTTPException(status_code=400, detail="Enter a positive Job Runner grand total before capturing the golden baseline.")

    raw_tolerance = body.tolerance or {}
    unknown_tolerance_fields = sorted(set(raw_tolerance) - set(DEFAULT_TOLERANCE))
    if unknown_tolerance_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown tolerance field(s): {', '.join(unknown_tolerance_fields)}.",
        )
    tolerance = dict(DEFAULT_TOLERANCE)
    for key, value in raw_tolerance.items():
        number = _as_number(value)
        if number is None or number < 0:
            raise HTTPException(status_code=400, detail=f"Tolerance '{key}' must be a non-negative number.")
        tolerance[key] = number
    for scope in ("proposal", "bundle"):
        if tolerance[f"{scope}_warn_abs"] < tolerance[f"{scope}_pass_abs"]:
            raise HTTPException(
                status_code=400,
                detail=f"{scope.title()} warning tolerance must be at least its pass tolerance.",
            )

    ruleset = _ruleset_with_engine_contract(get_ruleset_version())
    config_snapshot = {
        "waste_factors": WASTE_FACTORS,
        "sundry_rules": SUNDRY_RULES,
        "freight_rates": FREIGHT_RATES,
        "labor_qty_rules": LABOR_QTY_RULES,
        "stair_sundry_kits": STAIR_SUNDRY_KITS,
    }
    snapshot, fingerprint = make_golden_snapshot(
        job=job,
        company_rates=get_all_company_rates(),
        labor_catalog=get_labor_catalog_entries(),
        ruleset=ruleset,
        target_totals=target_totals,
        tolerance=tolerance,
        build=build_manifest_for_snapshot(),
        artifact_manifest=_job_artifact_manifest(job["id"]),
        config_snapshot=config_snapshot,
        proposal_source_fingerprint=current_proposal_fingerprint,
    )
    golden = upsert_golden_job(
        source_job_id=job["id"],
        name=job.get("project_name") or f"Job {job['id']}",
        jr_quote_id=body.jr_quote_id or "",
        target_totals=target_totals,
        tolerance=snapshot.get("tolerance") or DEFAULT_TOLERANCE,
        snapshot=snapshot,
        ruleset_version=(ruleset or {}).get("version"),
        source_fingerprint=fingerprint,
        notes=body.notes or "",
        reviewer_name=body.reviewer_name or "",
        engine_fingerprint=build_manifest_for_snapshot().get("engine_fingerprint", ""),
        artifact_manifest=_job_artifact_manifest(job["id"]),
        rules_registry_snapshot=ruleset or {},
        config_snapshot=config_snapshot,
        status="active",
    )
    log_activity(
        job["id"],
        "golden_baseline_captured",
        "Captured golden reproducibility baseline",
        {"golden_job_id": golden["id"], "jr_quote_id": body.jr_quote_id, "source_fingerprint": fingerprint},
    )
    return {
        "golden_job": _public_golden_job(golden),
        "latest_replay": None,
        "baseline_replay": None,
        "current_replay": None,
        "comparison_status": "not_replayed",
        "current_drift_classification": None,
        "replays": [],
        "active_version_replays": [],
        "default_tolerance": DEFAULT_TOLERANCE,
    }


@app.post("/api/jobs/{job_id}/reproducibility/replay")
def api_replay_golden_job(job_id: str, body: GoldenReplayRequest):
    """Run a non-mutating golden replay and persist the replay report."""
    db_id = _resolve_job_id(job_id)
    golden = get_golden_job_for_source(db_id)
    if not golden:
        raise HTTPException(status_code=404, detail="No golden baseline exists for this job yet.")

    mode = (body.mode or "baseline").lower().strip()
    if mode not in ("baseline", "current"):
        raise HTTPException(status_code=400, detail="Replay mode must be 'baseline' or 'current'.")

    result, trace = replay_golden_job(
        golden_job=golden,
        mode=mode,
        current_company_rates=get_all_company_rates(),
        current_labor_catalog=get_labor_catalog_entries(),
        current_ruleset=_ruleset_with_engine_contract(get_ruleset_version()),
        current_config_snapshot={
            "waste_factors": WASTE_FACTORS,
            "sundry_rules": SUNDRY_RULES,
            "freight_rates": FREIGHT_RATES,
            "labor_qty_rules": LABOR_QTY_RULES,
            "stair_sundry_kits": STAIR_SUNDRY_KITS,
        },
    )
    run_id = create_calculation_run(
        db_id,
        "golden_replay",
        source="system",
        metadata=_audit_metadata({
            "endpoint": "reproducibility/replay",
            "golden_job_id": golden["id"],
            "golden_version_id": golden.get("version_id"),
            "mode": mode,
            "status": result["summary"]["status"],
            "source_fingerprint": golden.get("source_fingerprint"),
        }),
    )
    for record in trace._records:
        record["run_id"] = run_id
    trace_count = save_calculation_traces(db_id, run_id, trace.records)
    summary = dict(result["summary"])
    summary["trace_count"] = trace_count
    complete_calculation_run(run_id, summary=summary)

    replay = save_golden_replay(
        golden_job_id=golden["id"],
        source_job_id=db_id,
        mode=mode,
        status=summary["status"],
        summary=summary,
        diff=result["diff"],
        generated_proposal=result["proposal"],
        audit_run_id=run_id,
        golden_version_id=golden.get("version_id"),
    )
    log_activity(
        db_id,
        "golden_replay_ran",
        f"Golden replay {mode} finished: {summary['status'].upper()}",
        {"golden_job_id": golden["id"], "replay_id": replay["id"], "mode": mode, "status": summary["status"]},
    )
    return _public_replay(replay, include_generated=False)


@app.get("/api/reproducibility/replays/{replay_id}")
def api_get_golden_replay(replay_id: int):
    """Return a full golden replay report."""
    replay = get_golden_replay(replay_id)
    if not replay:
        raise HTTPException(status_code=404, detail="Replay not found")
    return _public_replay(replay, include_generated=True)


def _as_number(value):
    try:
        if value is None or value == "":
            return None
        number = float(value)
        if not math.isfinite(number):
            return None
        return round(number, 4)
    except (TypeError, ValueError):
        return None


def _as_money_number(value):
    if isinstance(value, str):
        value = value.replace("$", "").replace(",", "").strip()
    return _as_number(value)


def _job_material_key(material: dict) -> str:
    return str(
        material.get("item_code")
        or material.get("id")
        or material.get("material_id")
        or ""
    )


def _audit_metadata(extra: dict = None) -> dict:
    """Attach visible rules metadata and the calculator identity to an audit."""
    metadata = dict(extra or {})
    build = get_build_info()
    metadata.update({
        "engine_fingerprint": build.get("engine_fingerprint"),
        "config_fingerprint": build.get("config_fingerprint"),
        **_calculator_data_fingerprints(),
    })
    ruleset = get_ruleset_version()
    if ruleset:
        metadata.update({
            "ruleset_version": ruleset.get("version"),
            "ruleset_rule_count": ruleset.get("rule_count"),
            "ruleset_active_count": ruleset.get("active_count"),
            "ruleset_created_at": ruleset.get("created_at"),
        })
    return metadata


def _current_ruleset_version() -> int | None:
    ruleset = get_ruleset_version()
    return ruleset.get("version") if ruleset else None


def _ensure_audit_calculator_current(run: dict | None, *, label: str) -> None:
    """Require proof from the running calculator, not a metadata-only rules edit."""
    build = get_build_info()
    metadata = (run or {}).get("metadata") or {}
    expected_engine = build.get("engine_fingerprint")
    expected_config = build.get("config_fingerprint")
    run_engine = metadata.get("engine_fingerprint")
    run_config = metadata.get("config_fingerprint")
    if not run_engine or not run_config:
        raise HTTPException(
            status_code=409,
            detail=f"{label} audit predates calculator fingerprint evidence. Recalculate or save the proposal first.",
        )
    if run_engine != expected_engine or run_config != expected_config:
        raise HTTPException(
            status_code=409,
            detail=f"{label} audit was created by a different calculator build or config. Recalculate or save the proposal first.",
        )


def _fingerprint_payload(payload) -> str:
    raw = json.dumps(payload, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _materials_source_fingerprint(materials: list[dict]) -> str:
    """Fingerprint only durable material fields used by pricing and calculation."""
    fields = (
        "id", "item_code", "description", "material_type", "installed_qty", "unit",
        "waste_pct", "order_qty", "vendor", "unit_price", "extended_cost",
        "ai_confidence", "quote_status", "price_source", "freight_per_unit",
        "freight_source", "fixture_count", "labor_rate_lf", "labor_catalog",
        "tack_strip_lf", "seam_tape_lf", "pad_sy", "area_type", "is_mosaic",
        "is_penny_hex", "crack_isolation_sf", "weld_rod_lf",
    )
    return _fingerprint_payload([
        {field: material.get(field) for field in fields}
        for material in materials
        if isinstance(material, dict)
    ])


def _calculator_data_fingerprints() -> dict:
    """Fingerprint mutable rate inputs that live outside the source job."""
    return {
        "company_rates_fingerprint": _fingerprint_payload(get_all_company_rates()),
        "labor_catalog_fingerprint": _fingerprint_payload(get_labor_catalog_entries()),
    }


def _job_source_snapshot(job: dict) -> dict:
    return {
        "project_name": job.get("project_name"),
        "gc_name": job.get("gc_name"),
        "address": job.get("address"),
        "city": job.get("city"),
        "state": job.get("state"),
        "zip": job.get("zip"),
        "tax_rate": job.get("tax_rate", 0),
        "gpm_pct": job.get("gpm_pct", 0),
        "markup_pct": job.get("markup_pct", 0),
        "unit_count": job.get("unit_count", 0),
        "tub_shower_count": job.get("tub_shower_count", 0),
        "salesperson": job.get("salesperson"),
        "exclusions": job.get("exclusions"),
        "textura_fee": job.get("textura_fee", 0),
    }


def _line_snapshot(items: list[dict], fields: tuple[str, ...]) -> list[dict]:
    rows = []
    for item in items or []:
        if isinstance(item, dict):
            rows.append({field: item.get(field) for field in fields})
    return rows


def _bid_source_fingerprint(job: dict) -> str:
    build = get_build_info()
    data_fingerprints = _calculator_data_fingerprints()
    return _fingerprint_payload({
        "fingerprint_schema": 2,
        "job": _job_source_snapshot(job),
        "materials": _line_snapshot(job.get("materials", []), (
            "id", "item_code", "description", "material_type", "installed_qty",
            "unit", "waste_pct", "order_qty", "unit_price", "extended_cost",
            "vendor", "price_source", "quote_status", "ai_confidence",
            "freight_per_unit", "freight_source", "fixture_count", "labor_rate_lf",
            "labor_catalog", "tack_strip_lf", "seam_tape_lf", "pad_sy", "area_type",
            "is_mosaic", "is_penny_hex", "crack_isolation_sf", "weld_rod_lf",
        )),
        "sundries": _line_snapshot(job.get("sundries", []), (
            "id", "material_id", "sundry_name", "qty", "unit", "unit_price",
            "extended_cost", "freight_cost",
        )),
        "labor": _line_snapshot(job.get("labor", []), (
            "id", "material_id", "labor_description", "qty", "unit", "rate", "extended_cost",
        )),
        "engine_fingerprint": build.get("engine_fingerprint"),
        "config_fingerprint": build.get("config_fingerprint"),
        **data_fingerprints,
    })


def _proposal_source_fingerprint(job: dict, proposal_data: dict | None = None) -> str:
    proposal_data = proposal_data if isinstance(proposal_data, dict) else (job.get("proposal_data") if isinstance(job.get("proposal_data"), dict) else {})
    build = get_build_info()
    data_fingerprints = _calculator_data_fingerprints()
    return _fingerprint_payload({
        "fingerprint_schema": 2,
        "job": _job_source_snapshot(job),
        "materials": _line_snapshot(job.get("materials", []), (
            "id", "item_code", "description", "material_type", "installed_qty",
            "unit", "waste_pct", "order_qty", "unit_price", "extended_cost",
            "vendor", "price_source", "quote_status", "ai_confidence",
            "freight_per_unit", "freight_source", "fixture_count", "labor_rate_lf",
            "labor_catalog", "tack_strip_lf", "seam_tape_lf", "pad_sy", "area_type",
            "is_mosaic", "is_penny_hex", "crack_isolation_sf", "weld_rod_lf",
        )),
        "proposal": {
            "bundles": proposal_data.get("bundles", []),
            "notes": proposal_data.get("notes", []),
            "terms": proposal_data.get("terms", []),
            "exclusions": proposal_data.get("exclusions", []),
            "tax_rate": proposal_data.get("tax_rate", 0),
            "gpm_pct": proposal_data.get("gpm_pct", 0),
            "textura_fee": proposal_data.get("textura_fee", 0),
            "subtotal": proposal_data.get("subtotal", 0),
            "tax_amount": proposal_data.get("tax_amount", 0),
            "grand_total": proposal_data.get("grand_total", 0),
            "gpm_profit": proposal_data.get("gpm_profit", 0),
            "gpm_labor": proposal_data.get("gpm_labor", 0),
            "gpm_material": proposal_data.get("gpm_material", 0),
            "manual_adjustment": proposal_data.get("manual_adjustment", 0),
            "textura_amount": proposal_data.get("textura_amount", 0),
            "deleted_bundles": proposal_data.get("deleted_bundles", []),
            "deleted_bundle_reasons": proposal_data.get("deleted_bundle_reasons", {}),
            "deleted_material_codes": proposal_data.get("deleted_material_codes", []),
            "deleted_material_reasons": proposal_data.get("deleted_material_reasons", {}),
        },
        "sundries": _line_snapshot(job.get("sundries", []), (
            "id", "material_id", "sundry_name", "qty", "unit", "unit_price",
            "extended_cost", "freight_cost",
        )),
        "labor": _line_snapshot(job.get("labor", []), (
            "id", "material_id", "labor_description", "qty", "unit", "rate", "extended_cost",
        )),
        "engine_fingerprint": build.get("engine_fingerprint"),
        "config_fingerprint": build.get("config_fingerprint"),
        **data_fingerprints,
    })


def _latest_completed_run(job_id: int, run_types: set[str]) -> dict | None:
    return get_latest_completed_calculation_run(job_id, run_types)


def _required_job_field_gaps(job: dict) -> list[str]:
    missing = []
    for field in ("project_name", "gc_name", "salesperson"):
        if not str(job.get(field) or "").strip():
            missing.append(field)
    return missing


def _validate_bid_job_ready(job: dict) -> None:
    """Block bid/PDF generation when the bid cannot be trusted."""
    missing = _required_job_field_gaps(job)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot generate bid PDF until required job fields are filled: {', '.join(missing)}.",
        )

    materials = [m for m in (job.get("materials") or []) if isinstance(m, dict)]
    if not materials:
        raise HTTPException(status_code=400, detail="Cannot generate bid PDF until materials are loaded.")

    unpriced = [
        m.get("item_code") or m.get("description") or f"material {index + 1}"
        for index, m in enumerate(materials)
        if _as_number(m.get("unit_price")) is None or _as_number(m.get("unit_price")) <= 0
    ]
    if unpriced:
        sample = ", ".join(str(item) for item in unpriced[:5])
        suffix = "..." if len(unpriced) > 5 else ""
        raise HTTPException(
            status_code=409,
            detail=f"Cannot generate bid PDF until all materials have prices. Missing: {sample}{suffix}",
        )

    unknown = [
        m.get("item_code") or m.get("description") or f"material {index + 1}"
        for index, m in enumerate(materials)
        if not is_valid_material_classification(m.get("material_type"))
    ]
    if unknown:
        sample = ", ".join(str(item) for item in unknown[:5])
        suffix = "..." if len(unknown) > 5 else ""
        raise HTTPException(
            status_code=409,
            detail=f"Cannot generate bid PDF until unknown material types are classified. Unknown: {sample}{suffix}",
        )


def _validate_bid_pdf_download_ready(job: dict) -> None:
    """Reject old bid PDFs after the job or bid audit has changed."""
    _validate_bid_job_ready(job)
    bid_data = job.get("bid_data")
    if not isinstance(bid_data, dict) or not bid_data.get("pdf_audit_run_id"):
        raise HTTPException(status_code=409, detail="Bid PDF is missing its audit receipt. Regenerate the bid.")
    latest = _latest_completed_run(job["id"], {"bid_pdf_generation"})
    if not latest or int(latest["id"]) != int(bid_data.get("pdf_audit_run_id")):
        raise HTTPException(status_code=409, detail="Bid PDF is stale. Regenerate the bid before downloading.")
    _ensure_audit_calculator_current(latest, label="Bid PDF")
    if bid_data.get("pdf_source_fingerprint") != _bid_source_fingerprint(job):
        raise HTTPException(status_code=409, detail="Bid PDF is stale because the job source changed. Regenerate the bid before downloading.")
    traces = get_calculation_traces(job["id"], run_id=latest["id"], entity_type="bid", entity_key="bid", limit=200)
    by_field = {trace.get("output_field"): trace for trace in traces}
    totals = bid_data.get("pdf_totals") or {}
    for field in ("subtotal", "tax_amount", "grand_total", "total_cost", "gpm_profit", "markup_amount"):
        _trace_result_matches(by_field.get(field), totals.get(field), label=f"bid {field}")


def _record_bid_audit(job_id: int, bid_data: dict) -> dict:
    """Persist audit rows for every displayed bid total."""
    trace = AuditTraceBuilder(job_id, default_source="bid_assembler")

    def _money(value) -> float:
        try:
            return round(float(value or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    bundles = [b for b in (bid_data.get("bundles") or []) if isinstance(b, dict)]
    bundle_components = []
    for index, bundle in enumerate(bundles):
        bundle_name = bundle.get("bundle_name") or f"bundle:{index}"
        component = {
            "bundle_name": bundle_name,
            "material_cost": _money(bundle.get("material_cost")),
            "sundry_cost": _money(bundle.get("sundry_cost")),
            "labor_cost": _money(bundle.get("labor_cost")),
            "freight_cost": _money(bundle.get("freight_cost")),
            "gpm_labor_adder": _money(bundle.get("gpm_labor_adder")),
            "gpm_material_adder": _money(bundle.get("gpm_material_adder")),
            "gpm_adder": _money(bundle.get("gpm_adder")),
            "total_price": _money(bundle.get("total_price")),
        }
        bundle_components.append(component)
        for field, value in component.items():
            if field == "bundle_name":
                continue
            inputs = {"bundle_name": bundle_name, "bundle_index": index}
            if field == "material_cost":
                inputs.update({
                    "order_qty": _money(bundle.get("order_qty")),
                    "unit_price": _money(bundle.get("unit_price")),
                    "waste_pct": bundle.get("waste_pct"),
                })
                formula = "order_qty * unit_price"
            elif field == "freight_cost":
                inputs.update({
                    "order_qty": _money(bundle.get("order_qty")),
                    "freight_rate": _money(bundle.get("freight_rate")),
                })
                formula = "order_qty * freight_rate"
            elif field == "total_price":
                inputs.update({
                    "material_cost": component["material_cost"],
                    "sundry_cost": component["sundry_cost"],
                    "labor_cost": component["labor_cost"],
                    "freight_cost": component["freight_cost"],
                    "gpm_adder": component["gpm_adder"],
                })
                formula = "material_cost + sundry_cost + labor_cost + freight_cost + gpm_adder"
            elif field == "gpm_adder":
                inputs.update({
                    "gpm_labor_adder": component["gpm_labor_adder"],
                    "gpm_material_adder": component["gpm_material_adder"],
                })
                formula = "gpm_labor_adder + gpm_material_adder"
            else:
                inputs[field] = value
                formula = f"bundle.{field}"
            trace.record(
                entity_type="bid_bundle",
                entity_key=bundle_name,
                output_field=field,
                formula=formula,
                inputs=inputs,
                result=value,
                rule_id=f"bid_assembler:bundle:{field}",
                source="bid_assembler",
            )

    def _bundle_values(field: str) -> list[dict]:
        return [{"bundle_name": b["bundle_name"], "value": b.get(field, 0)} for b in bundle_components]

    total_material = _money(sum(b.get("material_cost", 0) for b in bundle_components))
    total_sundry = _money(sum(b.get("sundry_cost", 0) for b in bundle_components))
    total_labor = _money(sum(b.get("labor_cost", 0) for b in bundle_components))
    total_freight = _money(sum(b.get("freight_cost", 0) for b in bundle_components))
    total_gpm_labor = _money(sum(b.get("gpm_labor_adder", 0) for b in bundle_components))
    total_gpm_material = _money(sum(b.get("gpm_material_adder", 0) for b in bundle_components))
    subtotal = _money(bid_data.get("subtotal"))
    markup_amount = _money(bid_data.get("markup_amount"))
    tax_amount = _money(bid_data.get("tax_amount"))
    grand_total = _money(bid_data.get("grand_total"))

    proposal_specs = [
        ("material_cost", "sum(bundle.material_cost)", total_material, {"bundles": _bundle_values("material_cost")}),
        ("sundry_cost", "sum(bundle.sundry_cost)", total_sundry, {"bundles": _bundle_values("sundry_cost")}),
        ("labor_cost", "sum(bundle.labor_cost)", total_labor, {"bundles": _bundle_values("labor_cost")}),
        ("freight_cost", "sum(bundle.freight_cost)", total_freight, {"bundles": _bundle_values("freight_cost")}),
        ("total_cost", "sum(bundle.total_price before profit)", _money(bid_data.get("total_cost")), {
            "material_cost": total_material,
            "sundry_cost": total_sundry,
            "labor_cost": total_labor,
            "freight_cost": total_freight,
        }),
        ("gpm_profit", "total_cost / (1 - gpm_pct) - total_cost", _money(bid_data.get("gpm_profit")), {
            "total_cost": _money(bid_data.get("total_cost")),
            "gpm_pct": bid_data.get("gpm_pct", 0),
        }),
        ("gpm_labor", "sum(bundle.gpm_labor_adder)", total_gpm_labor, {"bundles": _bundle_values("gpm_labor_adder")}),
        ("gpm_material", "sum(bundle.gpm_material_adder)", total_gpm_material, {"bundles": _bundle_values("gpm_material_adder")}),
        ("subtotal", "sum(bundle.total_price)", subtotal, {"bundles": _bundle_values("total_price")}),
        ("markup_amount", "subtotal * markup_pct", markup_amount, {
            "subtotal": subtotal,
            "markup_pct": bid_data.get("markup_pct", 0),
        }),
        ("tax_amount", "taxable * tax_rate", tax_amount, {
            "tax_rate": bid_data.get("tax_rate", 0),
            "taxable_components": ["material_cost", "sundry_cost", "freight_cost", "gpm_material_adder"],
            "markup_amount": markup_amount,
        }),
        ("grand_total", "subtotal + markup_amount + tax_amount", grand_total, {
            "subtotal": subtotal,
            "markup_amount": markup_amount,
            "tax_amount": tax_amount,
        }),
    ]
    for field, formula, result, inputs in proposal_specs:
        trace.record(
            entity_type="bid",
            entity_id=job_id,
            entity_key="bid",
            output_field=field,
            formula=formula,
            inputs=inputs,
            result=result,
            rule_id=f"bid_assembler:{field}",
            source="bid_assembler",
        )

    run_id = create_calculation_run(
        job_id,
        "bid_pdf_generation",
        source="system",
        metadata=_audit_metadata({"endpoint": "generate-bid"}),
    )
    for record in trace._records:
        record["run_id"] = run_id
    trace_count = save_calculation_traces(job_id, run_id, trace.records)
    complete_calculation_run(run_id, summary=trace.summary())
    run = list_calculation_runs(job_id, limit=1)[0]
    return {"run": run, "trace_count": trace_count}


def _record_proposal_editor_audit(job_id: int, previous: dict, current: dict) -> dict:
    """Persist a complete audit receipt for the currently displayed proposal."""
    previous = previous or {}
    current = current or {}
    trace = AuditTraceBuilder(job_id, default_source="proposal_editor")

    def _money(value) -> float:
        try:
            return round(float(value or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    def _bundle_freight(bundle: dict) -> float:
        return _money(bundle.get("freight_override") if bundle.get("freight_override") is not None else bundle.get("freight_cost"))

    bundles = [b for b in (current.get("bundles") or []) if isinstance(b, dict)]
    bundle_components = []
    for index, bundle in enumerate(bundles):
        bundle_components.append({
            "bundle_name": bundle.get("bundle_name") or f"bundle:{index}",
            "material_cost": _money(bundle.get("material_cost")),
            "sundry_cost": _money(bundle.get("sundry_cost")),
            "labor_cost": _money(bundle.get("labor_cost")),
            "freight_cost": _bundle_freight(bundle),
            "gpm_labor_adder": _money(bundle.get("gpm_labor_adder")),
            "gpm_material_adder": _money(bundle.get("gpm_material_adder")),
            "gpm_adder": _money(bundle.get("gpm_adder")),
            "taxable": _money(bundle.get("taxable")),
            "tax_amount": _money(bundle.get("tax_amount")),
            "calculated_total": _money(bundle.get("total_price")),
            "total_price": _money(bundle.get("price_override") if bundle.get("price_override") is not None else bundle.get("total_price")),
            "price_override": bundle.get("price_override"),
        })
    totals = {
        "material_cost": _money(sum(_money(b.get("material_cost")) for b in bundles)),
        "sundry_cost": _money(sum(_money(b.get("sundry_cost")) for b in bundles)),
        "labor_cost": _money(sum(_money(b.get("labor_cost")) for b in bundles)),
        "freight_cost": _money(sum(_bundle_freight(b) for b in bundles)),
        "gpm_profit": _money(current.get("gpm_profit")),
        "gpm_labor": _money(current.get("gpm_labor")),
        "gpm_material": _money(current.get("gpm_material")),
        "manual_adjustment": _money(current.get("manual_adjustment")),
        "subtotal": _money(current.get("subtotal")),
        "tax_amount": _money(current.get("tax_amount")),
        "textura_amount": _money(current.get("textura_amount")),
        "grand_total": _money(current.get("grand_total")),
    }
    totals["total_cost"] = _money(
        totals["material_cost"] + totals["sundry_cost"] + totals["labor_cost"] + totals["freight_cost"]
    )

    component_values = {
        field: [{"bundle_name": b["bundle_name"], "value": b[field]} for b in bundle_components]
        for field in ("material_cost", "sundry_cost", "labor_cost", "freight_cost", "tax_amount", "total_price")
    }
    proposal_trace_specs = [
        ("material_cost", "sum(bundle.material_cost)", {"bundles": component_values["material_cost"]}),
        ("sundry_cost", "sum(bundle.sundry_cost)", {"bundles": component_values["sundry_cost"]}),
        ("labor_cost", "sum(bundle.labor_cost)", {"bundles": component_values["labor_cost"]}),
        ("freight_cost", "sum(bundle.freight_override ?? bundle.freight_cost)", {"bundles": component_values["freight_cost"]}),
        ("total_cost", "material_cost + sundry_cost + labor_cost + freight_cost", {
            "material_cost": totals["material_cost"],
            "sundry_cost": totals["sundry_cost"],
            "labor_cost": totals["labor_cost"],
            "freight_cost": totals["freight_cost"],
        }),
        ("gpm_profit", "total_cost / (1 - gpm_pct) - total_cost", {
            "total_cost": totals["total_cost"],
            "gpm_pct": current.get("gpm_pct", 0),
        }),
        ("gpm_labor", "gpm_profit * 0.9793", {"gpm_profit": totals["gpm_profit"], "split_pct": 0.9793}),
        ("gpm_material", "gpm_profit - gpm_labor", {
            "gpm_profit": totals["gpm_profit"],
            "gpm_labor": totals["gpm_labor"],
        }),
        ("manual_adjustment", "sum(effective bundle total - calculated bundle total)", {
            "bundles": [
                {"bundle_name": bundle["bundle_name"], "calculated": bundle["calculated_total"], "effective": bundle["total_price"]}
                for bundle in bundle_components
            ],
        }),
        ("subtotal", "sum(effective bundle totals) - tax_amount", {
            "manual_adjustment": totals["manual_adjustment"],
            "tax_amount": totals["tax_amount"],
        }),
        ("tax_amount", "sum(bundle.tax_amount)", {
            "tax_rate": current.get("tax_rate", 0),
            "bundles": component_values["tax_amount"],
        }),
        ("textura_amount", "min((subtotal + tax_amount) * 0.0022, 5000) when enabled else 0", {
            "textura_fee": current.get("textura_fee", 0),
            "subtotal": totals["subtotal"],
            "tax_amount": totals["tax_amount"],
            "cap": 5000,
            "rate": 0.0022,
        }),
        ("grand_total", "subtotal + tax_amount + textura_amount", {
            "subtotal": totals["subtotal"],
            "tax_amount": totals["tax_amount"],
            "textura_amount": totals["textura_amount"],
        }),
    ]

    for field, formula, inputs in proposal_trace_specs:
        trace.record(
            entity_type="proposal",
            entity_id=job_id,
            entity_key="proposal",
            output_field=field,
            formula=formula,
            inputs=inputs,
            result=totals[field],
            rule_id=f"proposal_editor:{field}",
            source="proposal_editor",
        )

    for index, bundle in enumerate(bundles):
        bundle_name = bundle.get("bundle_name") or f"bundle:{index}"
        freight = _bundle_freight(bundle)
        bundle_total = _money(bundle.get("price_override") if bundle.get("price_override") is not None else bundle.get("total_price"))
        component = bundle_components[index] if index < len(bundle_components) else {}
        bundle_values = {
            "material_cost": _money(bundle.get("material_cost")),
            "sundry_cost": _money(bundle.get("sundry_cost")),
            "labor_cost": _money(bundle.get("labor_cost")),
            "freight_cost": freight,
            "gpm_labor_adder": _money(bundle.get("gpm_labor_adder")),
            "gpm_material_adder": _money(bundle.get("gpm_material_adder")),
            "gpm_adder": _money(bundle.get("gpm_adder")),
            "taxable": _money(bundle.get("taxable")),
            "tax_amount": _money(bundle.get("tax_amount")),
            "total_price": bundle_total,
        }
        for field, value in bundle_values.items():
            inputs = {"bundle_name": bundle_name, "bundle_index": index}
            if field in ("gpm_adder", "total_price", "tax_amount"):
                inputs.update({
                    "material_cost": component.get("material_cost"),
                    "sundry_cost": component.get("sundry_cost"),
                    "labor_cost": component.get("labor_cost"),
                    "freight_cost": component.get("freight_cost"),
                    "gpm_labor_adder": component.get("gpm_labor_adder"),
                    "gpm_material_adder": component.get("gpm_material_adder"),
                    "gpm_adder": component.get("gpm_adder"),
                    "taxable": component.get("taxable"),
                    "tax_amount": component.get("tax_amount"),
                    "price_override": component.get("price_override"),
                })
            elif field == "freight_cost":
                inputs.update({
                    "freight_cost": _money(bundle.get("freight_cost")),
                    "freight_override": bundle.get("freight_override"),
                })
            else:
                inputs[field] = value
            trace.record(
                entity_type="bundle",
                entity_key=bundle_name,
                output_field=field,
                formula={
                    "freight_cost": "freight_override if present else freight_cost",
                    "total_price": "price_override if present else material_cost + sundry_cost + labor_cost + freight_cost + gpm_adder + tax_amount",
                    "gpm_adder": "gpm_labor_adder + gpm_material_adder",
                    "tax_amount": "taxable * tax_rate",
                }.get(field, f"bundle.{field}"),
                inputs=inputs,
                result=value,
                rule_id=f"proposal_editor:bundle:{field}",
                source="proposal_editor",
            )

        for material_index, material in enumerate(bundle.get("materials") or []):
            if not isinstance(material, dict):
                continue
            order_qty = _money(material.get("order_qty") or material.get("installed_qty"))
            unit_price = _money(material.get("unit_price"))
            trace.record(
                entity_type="material",
                entity_id=material.get("id") or material.get("material_id"),
                entity_key=material.get("item_code") or material.get("description"),
                output_field="extended_cost",
                formula="order_qty * unit_price",
                inputs={
                    "bundle_name": bundle_name,
                    "bundle_index": index,
                    "line_index": material_index,
                    "order_qty": order_qty,
                    "unit_price": unit_price,
                },
                result=_money(material.get("extended_cost")),
                rule_id=f"proposal_editor:material:{material.get('material_type', '')}:extended_cost",
                source=material.get("price_source") or "proposal_editor",
            )

        for sundry_index, sundry in enumerate(bundle.get("sundry_items") or []):
            if not isinstance(sundry, dict):
                continue
            qty = _money(sundry.get("qty"))
            unit_price = _money(sundry.get("unit_price"))
            material_id = sundry.get("material_id")
            sundry_name = sundry.get("sundry_name") or "sundry"
            trace.record(
                entity_type="sundry",
                entity_id=material_id,
                entity_key=f"{material_id}:{sundry_name}",
                output_field="extended_cost",
                formula="qty * unit_price",
                inputs={
                    "bundle_name": bundle_name,
                    "bundle_index": index,
                    "line_index": sundry_index,
                    "qty": qty,
                    "unit_price": unit_price,
                    "unit": sundry.get("unit"),
                },
                result=_money(sundry.get("extended_cost")),
                rule_id=f"proposal_editor:sundry:{sundry_name}",
                source="proposal_editor",
            )

        for labor_index, labor in enumerate(bundle.get("labor_items") or []):
            if not isinstance(labor, dict):
                continue
            qty = _money(labor.get("qty"))
            rate = _money(labor.get("rate"))
            material_id = labor.get("material_id")
            labor_description = labor.get("labor_description") or "labor"
            trace.record(
                entity_type="labor",
                entity_id=material_id,
                entity_key=f"{material_id}:{labor_description}",
                output_field="extended_cost",
                formula="qty * rate",
                inputs={
                    "bundle_name": bundle_name,
                    "bundle_index": index,
                    "line_index": labor_index,
                    "qty": qty,
                    "rate": rate,
                    "unit": labor.get("unit"),
                },
                result=_money(labor.get("extended_cost")),
                rule_id=f"proposal_editor:labor:{labor.get('unit', '')}",
                source="proposal_editor",
            )

    manual_trace_count = 0
    for field in ("tax_rate", "gpm_pct", "textura_fee"):
        old = _as_number(previous.get(field))
        new = _as_number(current.get(field))
        if old != new:
            trace.manual_override(
                entity_type="proposal",
                entity_id=job_id,
                entity_key="proposal",
                output_field=field,
                prior_value=previous.get(field),
                value=current.get(field),
                note="Estimator changed a proposal calculation input.",
            )
            manual_trace_count += 1

    previous_bundles = previous.get("bundles") or []
    previous_by_name = {
        b.get("bundle_name"): b for b in previous_bundles
        if isinstance(b, dict) and b.get("bundle_name")
    }
    for index, bundle in enumerate(current.get("bundles") or []):
        if not isinstance(bundle, dict):
            continue
        bundle_name = bundle.get("bundle_name") or f"bundle:{index}"
        prior = previous_by_name.get(bundle.get("bundle_name"))
        if prior is None and index < len(previous_bundles):
            prior = previous_bundles[index] if isinstance(previous_bundles[index], dict) else {}
        prior = prior or {}

        for field in ("price_override", "freight_override"):
            prior_value = prior.get(field)
            value = bundle.get(field)
            if value is not None or prior_value is not None:
                trace.manual_override(
                    entity_type="bundle",
                    entity_key=bundle_name,
                    output_field=field,
                    prior_value=prior_value,
                    value=value,
                    note="Accepted bundle override is active." if value is not None else "Estimator cleared the accepted bundle override.",
                )
                manual_trace_count += 1

        prior_materials = {
            str(item.get("item_code") or item.get("id") or item.get("material_id") or ""): item
            for item in (prior.get("materials") or []) if isinstance(item, dict)
        }
        for material in (bundle.get("materials") or []):
            if not isinstance(material, dict) or not material.get("freight_is_manual"):
                continue
            material_key = str(material.get("item_code") or material.get("id") or material.get("material_id") or "")
            prior_material = prior_materials.get(material_key) or {}
            trace.manual_override(
                entity_type="material",
                entity_id=material.get("id") or material.get("material_id"),
                entity_key=material.get("item_code") or material.get("description"),
                output_field="freight_per_unit",
                prior_value=prior_material.get("freight_per_unit"),
                value=material.get("freight_per_unit"),
                note="Accepted material freight override is active.",
            )
            manual_trace_count += 1

        prior_sundries = {
            f"{item.get('material_id')}:{item.get('sundry_name') or 'sundry'}": item
            for item in (prior.get("sundry_items") or []) if isinstance(item, dict)
        }
        for sundry in (bundle.get("sundry_items") or []):
            if not isinstance(sundry, dict) or not sundry.get("is_manual_price"):
                continue
            sundry_key = f"{sundry.get('material_id')}:{sundry.get('sundry_name') or 'sundry'}"
            prior_sundry = prior_sundries.get(sundry_key) or {}
            trace.manual_override(
                entity_type="sundry",
                entity_id=sundry.get("material_id"),
                entity_key=sundry_key,
                output_field="extended_cost",
                prior_value=prior_sundry.get("extended_cost"),
                value=sundry.get("extended_cost"),
                note="Accepted sundry price override is active.",
            )
            manual_trace_count += 1

        prior_labor = {
            str(item.get("manual_source_key") or f"{item.get('material_id')}:{item.get('labor_description') or 'labor'}"): item
            for item in (prior.get("labor_items") or []) if isinstance(item, dict)
        }
        for labor in (bundle.get("labor_items") or []):
            if not isinstance(labor, dict) or not (labor.get("is_manual") or labor.get("is_stair_labor")):
                continue
            source_key = str(labor.get("manual_source_key") or f"{labor.get('material_id')}:{labor.get('labor_description') or 'labor'}")
            prior_line = prior_labor.get(source_key) or {}
            labor_description = labor.get("labor_description") or "labor"
            trace.manual_override(
                entity_type="labor",
                entity_id=labor.get("material_id"),
                entity_key=f"{labor.get('material_id')}:{labor_description}",
                output_field="extended_cost",
                prior_value=prior_line.get("extended_cost"),
                value=labor.get("extended_cost"),
                note="Accepted manual labor line is active.",
            )
            manual_trace_count += 1

        for deleted_key in (bundle.get("deleted_labor_keys") or []):
            trace.manual_override(
                entity_type="bundle",
                entity_key=bundle_name,
                output_field="deleted_labor",
                prior_value=(prior.get("deleted_labor_reasons") or {}).get(deleted_key),
                value=(bundle.get("deleted_labor_reasons") or {}).get(deleted_key),
                note=f"Generated labor line remains deleted: {deleted_key}",
            )
            manual_trace_count += 1

    if not trace.records:
        return {"trace_count": 0, "manual_trace_count": 0, "audit_trace": None}

    run_id = create_calculation_run(
        job_id,
        "proposal_editor_save",
        source="user",
        metadata=_audit_metadata({"endpoint": "proposal/bundles/save"}),
    )
    for record in trace._records:
        record["run_id"] = run_id
    trace_count = save_calculation_traces(job_id, run_id, trace.records)
    complete_calculation_run(run_id, summary=trace.summary())
    run = list_calculation_runs(job_id, limit=1)[0]
    return {
        "trace_count": trace_count,
        "manual_trace_count": manual_trace_count,
        "audit_trace": {"run": run, "traces": get_calculation_traces(job_id, run_id=run_id, limit=2000), "events": trace.records, "audit": trace.summary()},
    }


def _append_proposal_totals_snapshot(trace: AuditTraceBuilder, job_id: int, proposal: dict) -> None:
    """Append final editor-style proposal totals to a generation audit run."""
    if not isinstance(proposal, dict):
        return

    normalize_proposal_totals(proposal)
    if not trace:
        return

    def _money(value) -> float:
        try:
            return round(float(value or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    def _float(value) -> float:
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    def _freight(bundle: dict) -> float:
        return _money(bundle.get("freight_override") if bundle.get("freight_override") is not None else bundle.get("freight_cost"))

    bundles = [b for b in (proposal.get("bundles") or []) if isinstance(b, dict)]
    tax_rate = _float(proposal.get("tax_rate"))
    gpm_pct = _float(proposal.get("gpm_pct"))

    material_total = _money(sum(_money(b.get("material_cost")) for b in bundles))
    sundry_total = _money(sum(_money(b.get("sundry_cost")) for b in bundles))
    labor_total = _money(sum(_money(b.get("labor_cost")) for b in bundles))
    freight_total = _money(sum(_freight(b) for b in bundles))
    total_cost = _money(material_total + sundry_total + labor_total + freight_total)

    gpm_profit = _money(proposal.get("gpm_profit"))
    gpm_labor = _money(proposal.get("gpm_labor"))
    gpm_material = _money(proposal.get("gpm_material"))

    bundle_rows = []
    for index, bundle in enumerate(bundles):
        bundle_name = bundle.get("bundle_name") or f"bundle:{index}"
        freight = _freight(bundle)
        row = {
            "bundle_name": bundle_name,
            "bundle_index": index,
            "material_cost": _money(bundle.get("material_cost")),
            "sundry_cost": _money(bundle.get("sundry_cost")),
            "labor_cost": _money(bundle.get("labor_cost")),
            "freight_cost": freight,
            "gpm_labor_adder": bundle["gpm_labor_adder"],
            "gpm_material_adder": bundle["gpm_material_adder"],
            "gpm_adder": bundle["gpm_adder"],
            "taxable": bundle["taxable"],
            "tax_amount": bundle["tax_amount"],
            "calculated_total": _money(bundle.get("total_price")),
            "total_price": effective_bundle_total(bundle),
            "price_override": bundle.get("price_override"),
        }
        bundle_rows.append(row)

    tax_amount = _money(proposal.get("tax_amount"))
    subtotal = _money(proposal.get("subtotal"))
    manual_adjustment = _money(proposal.get("manual_adjustment"))
    textura_enabled = 1 if proposal.get("textura_fee") else 0
    textura_amount = _money(proposal.get("textura_amount"))
    grand_total = _money(proposal.get("grand_total"))

    def _bundle_values(field: str) -> list[dict]:
        return [{"bundle_name": row["bundle_name"], "value": row[field]} for row in bundle_rows]

    proposal_specs = [
        ("material_cost", "sum(bundle.material_cost)", {"bundles": _bundle_values("material_cost")}, material_total),
        ("sundry_cost", "sum(bundle.sundry_cost)", {"bundles": _bundle_values("sundry_cost")}, sundry_total),
        ("labor_cost", "sum(bundle.labor_cost)", {"bundles": _bundle_values("labor_cost")}, labor_total),
        ("freight_cost", "sum(bundle.freight_override ?? bundle.freight_cost)", {"bundles": _bundle_values("freight_cost")}, freight_total),
        ("total_cost", "material_cost + sundry_cost + labor_cost + freight_cost", {
            "material_cost": material_total,
            "sundry_cost": sundry_total,
            "labor_cost": labor_total,
            "freight_cost": freight_total,
        }, total_cost),
        ("gpm_profit", "total_cost / (1 - gpm_pct) - total_cost", {"total_cost": total_cost, "gpm_pct": gpm_pct}, gpm_profit),
        ("gpm_labor", "gpm_profit * 0.9793", {"gpm_profit": gpm_profit, "split_pct": 0.9793}, gpm_labor),
        ("gpm_material", "gpm_profit - gpm_labor", {"gpm_profit": gpm_profit, "gpm_labor": gpm_labor}, gpm_material),
        ("manual_adjustment", "sum(effective bundle total - calculated bundle total)", {
            "bundles": [
                {"bundle_name": row["bundle_name"], "calculated": row["calculated_total"], "effective": row["total_price"]}
                for row in bundle_rows
            ],
        }, manual_adjustment),
        ("subtotal", "sum(effective bundle totals) - tax_amount", {"manual_adjustment": manual_adjustment, "tax_amount": tax_amount}, subtotal),
        ("tax_amount", "sum(bundle.tax_amount)", {"tax_rate": tax_rate, "bundles": _bundle_values("tax_amount")}, tax_amount),
        ("textura_amount", "min((subtotal + tax_amount) * 0.0022, 5000) when enabled else 0", {
            "textura_fee": textura_enabled,
            "subtotal": subtotal,
            "tax_amount": tax_amount,
            "rate": 0.0022,
            "cap": 5000,
        }, textura_amount),
        ("grand_total", "subtotal + tax_amount + textura_amount", {
            "subtotal": subtotal,
            "tax_amount": tax_amount,
            "textura_amount": textura_amount,
        }, grand_total),
    ]
    for field, formula, inputs, result in proposal_specs:
        trace.record(
            entity_type="proposal",
            entity_id=job_id,
            entity_key="proposal",
            output_field=field,
            formula=formula,
            inputs=inputs,
            result=result,
            rule_id=f"proposal_generation:{field}",
            source="proposal_generation",
        )

    for row in bundle_rows:
        for field in (
            "material_cost", "sundry_cost", "labor_cost", "freight_cost",
            "gpm_labor_adder", "gpm_material_adder", "gpm_adder",
            "taxable", "tax_amount", "total_price",
        ):
            inputs = {"bundle_name": row["bundle_name"], "bundle_index": row["bundle_index"]}
            if field in ("gpm_adder", "tax_amount", "total_price"):
                inputs.update(row)
            elif field == "freight_cost":
                inputs.update({"freight_cost": row["freight_cost"], "price_override": row.get("price_override")})
            else:
                inputs[field] = row[field]
            trace.record(
                entity_type="bundle",
                entity_key=row["bundle_name"],
                output_field=field,
                formula={
                    "freight_cost": "freight_override if present else freight_cost",
                    "gpm_adder": "gpm_labor_adder + gpm_material_adder",
                    "tax_amount": "taxable * tax_rate",
                    "total_price": "price_override if present else material_cost + sundry_cost + labor_cost + freight_cost + gpm_adder + tax_amount",
                }.get(field, f"bundle.{field}"),
                inputs=inputs,
                result=row[field],
                rule_id=f"proposal_generation:bundle:{field}",
                source="proposal_generation",
            )


@app.put("/api/jobs/{job_id}/proposal/bundles")
@app.post("/api/jobs/{job_id}/proposal/bundles/save")
async def api_save_proposal_bundles(job_id: str, request: Request):
    """Auto-save proposal editor state (bundles, notes, terms, GPM, etc.)."""
    import json as _json
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    body = await request.json()
    previous_proposal_data = job.get("proposal_data") or {}
    incoming_session_id = str(body.get("client_session_id") or "").strip()
    try:
        incoming_edit_version = max(0, int(body.get("client_edit_version") or 0))
    except (TypeError, ValueError):
        incoming_edit_version = 0
    try:
        incoming_save_sequence = max(0, int(body.get("client_save_sequence") or 0))
    except (TypeError, ValueError):
        incoming_save_sequence = 0
    stored_session_id = str(previous_proposal_data.get("_client_session_id") or "").strip()
    try:
        stored_edit_version = max(0, int(previous_proposal_data.get("_client_edit_version") or 0))
    except (TypeError, ValueError):
        stored_edit_version = 0
    try:
        stored_save_sequence = max(0, int(previous_proposal_data.get("_client_save_sequence") or 0))
    except (TypeError, ValueError):
        stored_save_sequence = 0
    try:
        stored_server_revision = max(0, int(previous_proposal_data.get("_server_revision") or 0))
    except (TypeError, ValueError):
        stored_server_revision = 0
    incoming_base_revision_present = "base_server_revision" in body
    try:
        incoming_base_revision = max(0, int(body.get("base_server_revision") or 0))
    except (TypeError, ValueError):
        incoming_base_revision = -1
    same_client_session = bool(incoming_session_id and incoming_session_id == stored_session_id)
    if (
        same_client_session
        and (
            incoming_save_sequence < stored_save_sequence
            or incoming_edit_version < stored_edit_version
        )
    ):
        return {
            "status": "stale_ignored",
            "stale_save_ignored": True,
            "manual_trace_count": 0,
            "trace_count": 0,
            "audit_trace": None,
            "audit": None,
            "proposal_data": previous_proposal_data,
        }
    if (
        incoming_session_id
        and not same_client_session
        and incoming_base_revision_present
        and incoming_base_revision != stored_server_revision
    ):
        raise HTTPException(
            status_code=409,
            detail=(
                "This proposal changed in another tab or session. Your stale copy was not saved. "
                "Reload the job and review the newer accepted proposal before editing again."
            ),
        )
    proposal_data = {
        "bundles": body.get("bundles", []),
        "notes": body.get("notes", []),
        "terms": body.get("terms", []),
        "exclusions": body.get("exclusions", []),
        "tax_rate": body.get("tax_rate", 0),
        "gpm_pct": body.get("gpm_pct", 0),
        "textura_fee": body.get("textura_fee", 0),
        "subtotal": body.get("subtotal", 0),
        "tax_amount": body.get("tax_amount", 0),
        "grand_total": body.get("grand_total", 0),
        "gpm_profit": body.get("gpm_profit", 0),
        "gpm_labor": body.get("gpm_labor", 0),
        "gpm_material": body.get("gpm_material", 0),
        "manual_adjustment": body.get("manual_adjustment", 0),
        "textura_amount": body.get("textura_amount", 0),
        "deleted_bundles": body.get("deleted_bundles", []),
        "deleted_bundle_reasons": body.get("deleted_bundle_reasons", {}),
        "deleted_material_codes": body.get("deleted_material_codes", []),
        "deleted_material_reasons": body.get("deleted_material_reasons", {}),
        "audit": body.get("audit", {}),
        "_client_session_id": incoming_session_id,
        "_client_edit_version": incoming_edit_version,
        "_client_save_sequence": incoming_save_sequence,
        "_server_revision": stored_server_revision + 1,
    }
    normalize_proposal_totals(proposal_data)
    proposal_data["audit_source_fingerprint"] = _proposal_source_fingerprint(job, proposal_data)
    audit_result = _record_proposal_editor_audit(job["id"], previous_proposal_data, proposal_data)
    if audit_result.get("audit_trace"):
        proposal_data["audit"] = {
            "run_id": audit_result["audit_trace"]["run"]["id"],
            "trace_count": audit_result["trace_count"],
            "summary": audit_result["audit_trace"].get("audit", {}),
        }
    job["proposal_data"] = proposal_data
    save_job(job)
    return {
        "status": "ok",
        "manual_trace_count": audit_result.get("manual_trace_count", 0),
        "trace_count": audit_result.get("trace_count", 0),
        "audit_trace": audit_result.get("audit_trace"),
        "audit": audit_result.get("audit_trace"),
        "proposal_data": proposal_data,
    }


def _validate_proposal_pdf_ready(job: dict, body: dict) -> None:
    """Reject PDF generation when required header data or current audit is missing."""
    missing = _required_job_field_gaps(job)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot generate PDF until required job fields are filled: {', '.join(missing)}.",
        )

    saved_proposal = job.get("proposal_data") if isinstance(job.get("proposal_data"), dict) else {}
    body_fingerprint = _proposal_source_fingerprint(job, body)
    if not saved_proposal.get("audit_source_fingerprint") or saved_proposal.get("audit_source_fingerprint") != body_fingerprint:
        raise HTTPException(
            status_code=409,
            detail="Cannot generate PDF because this is not the exact saved and audited proposal. Save it and try again.",
        )

    raw_deleted_codes = {str(code) for code in (body.get("deleted_material_codes") or []) if code}
    deleted_material_reasons = body.get("deleted_material_reasons")
    if not isinstance(deleted_material_reasons, dict):
        deleted_material_reasons = {}
    missing_material_reasons = [
        code for code in sorted(raw_deleted_codes)
        if not str(deleted_material_reasons.get(code) or "").strip()
    ]
    if missing_material_reasons:
        raise HTTPException(status_code=409, detail="Cannot generate PDF until every deleted material has an estimator reason.")

    deleted_bundle_names = {str(name) for name in (body.get("deleted_bundles") or []) if name}
    deleted_bundle_reasons = body.get("deleted_bundle_reasons")
    if not isinstance(deleted_bundle_reasons, dict):
        deleted_bundle_reasons = {}
    if any(not str(deleted_bundle_reasons.get(name) or "").strip() for name in deleted_bundle_names):
        raise HTTPException(status_code=409, detail="Cannot generate PDF until every deleted bundle has an estimator reason.")

    accepted_codes = {
        _job_material_key(material)
        for bundle in (body.get("bundles") or [])
        if isinstance(bundle, dict)
        for material in (bundle.get("materials") or [])
        if isinstance(material, dict) and _job_material_key(material)
    }
    if raw_deleted_codes & accepted_codes:
        raise HTTPException(status_code=409, detail="Cannot generate PDF because a material is both deleted and present in an accepted bundle.")

    for bundle in body.get("bundles") or []:
        if not isinstance(bundle, dict):
            continue
        reasons = bundle.get("deleted_labor_reasons")
        if not isinstance(reasons, dict):
            reasons = {}
        if any(not str(reasons.get(key) or "").strip() for key in (bundle.get("deleted_labor_keys") or [])):
            raise HTTPException(status_code=409, detail="Cannot generate PDF until every deleted labor line has an estimator reason.")

    active_materials = [
        material for material in (job.get("materials") or [])
        if isinstance(material, dict) and _job_material_key(material) not in raw_deleted_codes
    ]
    unknown = [
        material.get("item_code") or material.get("description") or "material"
        for material in active_materials
        if not is_valid_material_classification(material.get("material_type"))
    ]
    unpriced = [
        material.get("item_code") or material.get("description") or "material"
        for material in active_materials
        if _as_number(material.get("unit_price")) is None or _as_number(material.get("unit_price")) <= 0
    ]
    if unknown or unpriced:
        issues = []
        if unknown:
            issues.append(f"classify {', '.join(str(item) for item in unknown[:5])}")
        if unpriced:
            issues.append(f"price {', '.join(str(item) for item in unpriced[:5])}")
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF until active materials are ready: {'; '.join(issues)}.")

    arithmetic_errors = proposal_math_errors(body)
    if arithmetic_errors:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot generate PDF because proposal arithmetic is invalid: {' '.join(arithmetic_errors[:3])}",
        )

    run = _latest_completed_run(job["id"], {"proposal_editor_save", "proposal_generation"})
    if not run:
        raise HTTPException(status_code=409, detail="Cannot generate PDF until this proposal has a current audit trace. Save or regenerate first.")
    _ensure_audit_calculator_current(run, label="Proposal PDF")
    _validate_proposal_body_matches_job_source(job, body)
    traces = get_calculation_traces(job["id"], run_id=run["id"], limit=5000)
    proposal_traces = [
        trace for trace in traces
        if trace.get("entity_type") == "proposal" and trace.get("entity_key") == "proposal"
    ]
    by_field = {}
    for trace in proposal_traces:
        by_field[trace.get("output_field")] = trace
    required = ["subtotal", "tax_amount", "grand_total", "gpm_profit", "gpm_labor", "gpm_material", "manual_adjustment", "textura_amount"]
    missing_traces = [field for field in required if field not in by_field]
    if missing_traces:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot generate PDF because audit is missing: {', '.join(missing_traces)}.",
        )
    for field in required:
        _trace_result_matches(by_field[field], body.get(field), label=f"proposal {field}")
    _validate_proposal_body_against_trace(body, traces)


def _trace_result_matches(trace: dict | None, expected, *, label: str) -> None:
    if not trace:
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because audit is missing for {label}.")
    expected_value = _as_number(expected)
    trace_value = _as_number(trace.get("result_value"))
    if expected_value is None:
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because {label} is not a finite number.")
    if trace_value is None:
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because audit has no numeric result for {label}.")
    if abs(expected_value - trace_value) > 0.02:
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because audit is stale for {label}.")


def _find_trace(
    traces: list[dict],
    *,
    entity_type: str,
    output_field: str,
    entity_key: str = None,
    entity_id=None,
    bundle_index: int = None,
    line_index: int = None,
) -> dict | None:
    matches = [
        trace for trace in traces
        if trace.get("entity_type") == entity_type and trace.get("output_field") == output_field
    ]
    if entity_key is not None:
        matches = [trace for trace in matches if str(trace.get("entity_key") or "") == str(entity_key)]
    if entity_id is not None:
        matches = [trace for trace in matches if str(trace.get("entity_id") or "") == str(entity_id)]
    if bundle_index is not None:
        traces_with_index = [trace for trace in matches if "bundle_index" in (trace.get("inputs") or {})]
        if traces_with_index:
            matches = [
                trace for trace in traces_with_index
                if (trace.get("inputs") or {}).get("bundle_index") == bundle_index
            ]
    if line_index is not None:
        traces_with_index = [trace for trace in matches if "line_index" in (trace.get("inputs") or {})]
        if traces_with_index:
            matches = [
                trace for trace in traces_with_index
                if (trace.get("inputs") or {}).get("line_index") == line_index
            ]
    return matches[-1] if matches else None


def _validate_proposal_body_against_trace(body: dict, traces: list[dict]) -> None:
    """Make sure every bundle/line number sent to the PDF has a matching trace."""
    bundles = [b for b in (body.get("bundles") or []) if isinstance(b, dict)]
    for bundle_index, bundle in enumerate(bundles):
        bundle_name = bundle.get("bundle_name") or f"bundle:{bundle_index}"
        bundle_fields = {
            "material_cost": bundle.get("material_cost"),
            "sundry_cost": bundle.get("sundry_cost"),
            "labor_cost": bundle.get("labor_cost"),
            "freight_cost": bundle.get("freight_override") if bundle.get("freight_override") is not None else bundle.get("freight_cost"),
            "gpm_labor_adder": bundle.get("gpm_labor_adder"),
            "gpm_material_adder": bundle.get("gpm_material_adder"),
            "gpm_adder": bundle.get("gpm_adder"),
            "taxable": bundle.get("taxable"),
            "tax_amount": bundle.get("tax_amount"),
            "total_price": bundle.get("price_override") if bundle.get("price_override") is not None else bundle.get("total_price"),
        }
        for field, expected in bundle_fields.items():
            trace = _find_trace(
                traces,
                entity_type="bundle",
                output_field=field,
                entity_key=bundle_name,
                bundle_index=bundle_index,
            )
            _trace_result_matches(trace, expected, label=f"{bundle_name} {field}")

        for line_index, material in enumerate(bundle.get("materials") or []):
            if not isinstance(material, dict):
                continue
            trace = _find_trace(
                traces,
                entity_type="material",
                output_field="extended_cost",
                entity_key=material.get("item_code") or material.get("description"),
                entity_id=material.get("id") or material.get("material_id"),
                bundle_index=bundle_index,
                line_index=line_index,
            )
            _trace_result_matches(trace, material.get("extended_cost"), label=f"{bundle_name} material line {line_index + 1}")

        for line_index, sundry in enumerate(bundle.get("sundry_items") or []):
            if not isinstance(sundry, dict):
                continue
            material_id = sundry.get("material_id")
            sundry_name = sundry.get("sundry_name") or "sundry"
            trace = _find_trace(
                traces,
                entity_type="sundry",
                output_field="extended_cost",
                entity_key=f"{material_id}:{sundry_name}",
                entity_id=material_id,
                bundle_index=bundle_index,
                line_index=line_index,
            )
            _trace_result_matches(trace, sundry.get("extended_cost"), label=f"{bundle_name} sundry line {line_index + 1}")

        for line_index, labor in enumerate(bundle.get("labor_items") or []):
            if not isinstance(labor, dict):
                continue
            material_id = labor.get("material_id")
            labor_description = labor.get("labor_description") or "labor"
            trace = _find_trace(
                traces,
                entity_type="labor",
                output_field="extended_cost",
                entity_key=f"{material_id}:{labor_description}",
                entity_id=material_id,
                bundle_index=bundle_index,
                line_index=line_index,
            )
            _trace_result_matches(trace, labor.get("extended_cost"), label=f"{bundle_name} labor line {line_index + 1}")


def _validate_proposal_body_matches_job_source(job: dict, body: dict) -> None:
    """Reject stale proposal bodies after material source edits."""
    def _is_synthetic_material(material: dict) -> bool:
        identifiers = (
            material.get("item_code"),
            material.get("id"),
            material.get("material_id"),
        )
        return any(str(value or "").startswith("synthetic_") for value in identifiers)

    deleted_codes = {str(code) for code in (body.get("deleted_material_codes") or []) if code}
    current_by_id = {}
    current_by_code = {}
    for material in job.get("materials", []) or []:
        if not isinstance(material, dict):
            continue
        if _is_synthetic_material(material):
            continue
        if material.get("id") is not None:
            current_by_id[str(material.get("id"))] = material
        if material.get("item_code"):
            current_by_code[str(material.get("item_code"))] = material

    seen = set()
    compare_fields = (
        "item_code", "description", "material_type", "installed_qty", "unit", "waste_pct", "order_qty",
        "vendor", "unit_price", "extended_cost", "price_source", "quote_status",
        "ai_confidence", "freight_per_unit", "freight_source", "fixture_count",
        "labor_rate_lf", "labor_catalog", "tack_strip_lf", "seam_tape_lf",
        "pad_sy", "area_type", "is_mosaic", "is_penny_hex",
        "crack_isolation_sf", "weld_rod_lf",
    )
    text_fields = {
        "item_code", "description", "material_type", "unit", "vendor",
        "price_source", "quote_status", "freight_source", "labor_catalog", "area_type",
    }

    for bundle in body.get("bundles") or []:
        if not isinstance(bundle, dict):
            continue
        for material in bundle.get("materials") or []:
            if not isinstance(material, dict):
                continue
            if _is_synthetic_material(material):
                continue
            item_code = str(material.get("item_code") or "")
            current = None
            if material.get("id") is not None:
                current = current_by_id.get(str(material.get("id")))
            if current is None and material.get("material_id") is not None:
                current = current_by_id.get(str(material.get("material_id")))
            if current is None and item_code:
                current = current_by_code.get(item_code)
            if current is None:
                raise HTTPException(status_code=409, detail=f"Cannot generate PDF because proposal material {item_code or 'line'} no longer exists. Regenerate the proposal.")

            seen.add(str(current.get("id")))
            label = item_code or current.get("description") or f"material {current.get('id')}"
            for field in compare_fields:
                # Per-material freight may be an accepted proposal edit. It is
                # fingerprinted and audited with the proposal, so comparing it
                # to the raw job-material value would incorrectly block a valid
                # PDF after the estimator explicitly changed freight.
                if field == "freight_per_unit" and material.get("freight_is_manual"):
                    continue
                current_value = current.get(field)
                body_value = material.get(field)
                if field in text_fields:
                    if str(current_value or "") != str(body_value or ""):
                        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because proposal material {label} is stale for {field}. Regenerate the proposal.")
                    continue
                current_number = _as_number(current_value)
                body_number = _as_number(body_value)
                if current_number is not None or body_number is not None:
                    if abs((current_number or 0) - (body_number or 0)) > 0.02:
                        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because proposal material {label} is stale for {field}. Regenerate the proposal.")
                elif str(current_value or "") != str(body_value or ""):
                    raise HTTPException(status_code=409, detail=f"Cannot generate PDF because proposal material {label} is stale for {field}. Regenerate the proposal.")

    missing = []
    for material in job.get("materials", []) or []:
        if not isinstance(material, dict):
            continue
        if _is_synthetic_material(material):
            continue
        item_code = str(material.get("item_code") or "")
        if _job_material_key(material) in deleted_codes:
            continue
        material_id = str(material.get("id"))
        if material_id and material_id not in seen:
            missing.append(item_code or material.get("description") or material_id)
    if missing:
        sample = ", ".join(str(item) for item in missing[:5])
        suffix = "..." if len(missing) > 5 else ""
        raise HTTPException(status_code=409, detail=f"Cannot generate PDF because proposal is missing current materials: {sample}{suffix}. Regenerate the proposal.")


def _validate_proposal_pdf_download_ready(job: dict) -> None:
    """Reject old proposal PDFs after the proposal audit or required job fields change."""
    missing = _required_job_field_gaps(job)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot download proposal PDF until required job fields are filled: {', '.join(missing)}.",
        )
    proposal_data = job.get("proposal_data")
    if not isinstance(proposal_data, dict) or not proposal_data.get("pdf_audit_run_id"):
        raise HTTPException(status_code=409, detail="Proposal PDF is missing its audit receipt. Regenerate the proposal PDF.")
    latest = _latest_completed_run(job["id"], {"proposal_editor_save", "proposal_generation"})
    if not latest or int(latest["id"]) != int(proposal_data.get("pdf_audit_run_id")):
        raise HTTPException(status_code=409, detail="Proposal PDF is stale. Regenerate the proposal PDF before downloading.")
    _ensure_audit_calculator_current(latest, label="Proposal PDF")
    if proposal_data.get("pdf_source_fingerprint") != _proposal_source_fingerprint(job, proposal_data):
        raise HTTPException(status_code=409, detail="Proposal PDF is stale because the job or proposal source changed. Regenerate the proposal PDF before downloading.")
    traces = get_calculation_traces(
        job["id"],
        run_id=latest["id"],
        entity_type="proposal",
        entity_key="proposal",
        limit=200,
    )
    by_field = {trace.get("output_field"): trace for trace in traces}
    totals = proposal_data.get("pdf_totals") or {}
    for field in ("subtotal", "tax_amount", "grand_total", "gpm_profit", "gpm_labor", "gpm_material", "manual_adjustment", "textura_amount"):
        _trace_result_matches(by_field.get(field), totals.get(field), label=f"proposal {field}")


@app.post("/api/jobs/{job_id}/proposal/generate")
def api_generate_proposal(job_id: str):
    """Auto-bundle materials into proposal line items.

    Preserves AI-rewritten bundle names and descriptions from prior runs
    by snapshotting them keyed on material item_code before regeneration,
    then re-applying after.
    """
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    trace = AuditTraceBuilder(job["id"])

    # ── Snapshot existing rewrites so Regenerate doesn't destroy them ──────
    # Key by the first material's item_code per bundle.
    existing_rewrites: dict[str, dict] = {}
    existing_by_codes: dict[tuple[str, ...], list[dict]] = {}
    existing_pd = job.get("proposal_data") or {}
    has_saved_accepted_proposal = bool(existing_pd.get("bundles"))
    for b in (existing_pd.get("bundles") or []):
        mats = b.get("materials") or []
        if not mats:
            continue
        codes = tuple(sorted(
            str(m.get("item_code") or m.get("id") or m.get("material_id"))
            for m in mats
            if m.get("item_code") or m.get("id") or m.get("material_id")
        ))
        if codes:
            existing_by_codes.setdefault(codes, []).append(b)
        key = str(mats[0].get("item_code") or mats[0].get("id") or mats[0].get("material_id") or "").strip()
        if not key:
            continue
        existing_rewrites[key] = {
            "bundle_name": b.get("bundle_name"),
            "description_text": b.get("description_text"),
        }

    # ── Carry forward deleted-bundle / deleted-material-code flags ─────────
    # Users mark these via the UI's delete action; they must survive regenerate.
    deleted_bundle_names = set(existing_pd.get("deleted_bundles") or [])
    raw_deleted_bundle_reasons = existing_pd.get("deleted_bundle_reasons")
    deleted_bundle_reasons = dict(raw_deleted_bundle_reasons) if isinstance(raw_deleted_bundle_reasons, dict) else {}
    deleted_material_codes = set(existing_pd.get("deleted_material_codes") or [])
    raw_deleted_reasons = existing_pd.get("deleted_material_reasons")
    deleted_material_reasons = dict(raw_deleted_reasons) if isinstance(raw_deleted_reasons, dict) else {}

    # Always recalculate sundries and labor to reflect latest rules/flags
    materials = job.get("materials", [])
    unit_count = job.get("unit_count", 0) or 0
    tub_shower_count = job.get("tub_shower_count", 0) or 0

    # Re-apply current waste rules to materials so rule changes propagate.
    # Skip piece-priced EA items (Schluter sticks) whose order_qty isn't waste-based.
    import json as _json
    _waste_data = get_company_rate("waste_factors")
    _waste_factors = _json.loads(_waste_data) if _waste_data else WASTE_FACTORS
    waste_touched = False
    for mat in materials:
        mtype = mat.get("material_type", "")
        new_waste = _waste_factors.get(mtype)
        if new_waste is None:
            continue
        mat_unit = (mat.get("unit") or "").upper()
        if mat_unit == "EA" and mtype != "sound_mat":
            continue  # Piece-counted materials — waste doesn't apply to order_qty
        old_waste = mat.get("waste_pct", 0) or 0
        if abs(new_waste - old_waste) < 1e-6:
            continue
        installed_qty = mat.get("installed_qty", 0) or 0
        mat["waste_pct"] = new_waste
        mat["order_qty"] = round(installed_qty * (1 + new_waste), 2)
        mat["extended_cost"] = round(mat["order_qty"] * (mat.get("unit_price", 0) or 0), 2)
        trace.record(
            entity_type="material",
            entity_id=mat.get("id"),
            entity_key=mat.get("item_code"),
            output_field="order_qty",
            formula="installed_qty * (1 + waste_pct)",
            inputs={"installed_qty": installed_qty, "waste_pct": new_waste, "old_waste_pct": old_waste},
            result=mat["order_qty"],
            rule_id=f"waste_factor:{mtype}",
            source="waste_factors",
        )
        trace.record(
            entity_type="material",
            entity_id=mat.get("id"),
            entity_key=mat.get("item_code"),
            output_field="extended_cost",
            formula="order_qty * unit_price",
            inputs={"order_qty": mat["order_qty"], "unit_price": mat.get("unit_price", 0) or 0},
            result=mat["extended_cost"],
            rule_id=f"material:{mtype}:extended_cost",
            source=mat.get("price_source") or "waste_factors",
        )
        waste_touched = True
    if waste_touched:
        save_materials(job["id"], materials)

    # Stamp job-level counts onto materials so sundry_calc can use them
    for mat in materials:
        mtype = mat.get("material_type", "")
        if mtype == "backsplash":
            mat["unit_count"] = unit_count
        if mtype == "tub_shower_surround":
            mat["tub_shower_total"] = tub_shower_count  # total tubs/showers on job

    if materials:
        sundries = calculate_sundries_for_materials(materials, trace=trace)
        save_sundries(job["id"], sundries)
        labor_items = calculate_labor_for_materials(materials, trace=trace)
        save_labor(job["id"], labor_items)
        # Reload job with freshly calculated sundries/labor
        job = load_job(db_id)

    current_company_rates = get_all_company_rates()
    proposal = generate_proposal_data(
        job["id"],
        job,
        trace=trace,
        freight_rates_override=current_company_rates.get("freight_rates") or FREIGHT_RATES,
    )

    # ── Re-apply snapshotted rewrites where item_codes match ───────────────
    for b in proposal.get("bundles", []):
        mats = b.get("materials") or []
        if not mats:
            continue
        codes = tuple(sorted(
            str(m.get("item_code") or m.get("id") or m.get("material_id"))
            for m in mats
            if m.get("item_code") or m.get("id") or m.get("material_id")
        ))
        candidates = existing_by_codes.get(codes, [])
        named_candidates = [candidate for candidate in candidates if candidate.get("bundle_name") == b.get("bundle_name")]
        exact = named_candidates[0] if len(named_candidates) == 1 else (candidates[0] if len(candidates) == 1 else None)
        key = str(mats[0].get("item_code") or mats[0].get("id") or mats[0].get("material_id") or "").strip()
        rw = exact if codes else existing_rewrites.get(key)
        if rw:
            if rw.get("bundle_name"):
                b["bundle_name"] = rw["bundle_name"]
            if rw.get("description_text"):
                b["description_text"] = rw["description_text"]
        if exact:
            for field in ("price_override", "freight_override", "stair_count", "stair_labor_type"):
                if exact.get(field) is not None:
                    b[field] = exact[field]

    # Carry deletion lists back so the FE save can persist them.
    # The shared normalizer below recalculates totals once after these flags and
    # any exact-match accepted overrides have been restored.
    if deleted_bundle_names or deleted_material_codes:
        proposal["deleted_bundles"] = sorted(deleted_bundle_names)
        proposal["deleted_bundle_reasons"] = deleted_bundle_reasons
        proposal["deleted_material_codes"] = sorted(deleted_material_codes)
        proposal["deleted_material_reasons"] = deleted_material_reasons
        trace.record(
            entity_type="proposal",
            entity_id=job["id"],
            entity_key="proposal",
            output_field="bundle_count",
            formula="filter generated bundles by deleted_bundles and deleted_material_codes before totals",
            inputs={
                "deleted_bundles": sorted(deleted_bundle_names),
                "deleted_material_codes": sorted(deleted_material_codes),
            },
            result=len(proposal.get("bundles") or []),
            rule_id="proposal:deleted_bundle_filter",
            source="proposal_bundler",
        )

    for field in ("notes", "terms", "exclusions"):
        if field in existing_pd:
            proposal[field] = list(existing_pd.get(field) or [])
    accepted_edit_count = apply_accepted_numeric_edits(proposal, existing_pd) if has_saved_accepted_proposal else 0
    if accepted_edit_count:
        trace.record(
            entity_type="proposal",
            entity_id=job["id"],
            entity_key="proposal",
            output_field="accepted_numeric_edits",
            formula="reapply explicit estimator inputs and overrides after raw bundle generation",
            inputs={"accepted_proposal_present": True},
            result=accepted_edit_count,
            rule_id="proposal:accepted_numeric_edits",
            source="proposal_regeneration",
        )

    _append_proposal_totals_snapshot(trace, job["id"], proposal)

    run_id = create_calculation_run(
        job["id"],
        "proposal_generation",
        metadata=_audit_metadata({"endpoint": "proposal/generate"}),
    )
    trace_count = save_calculation_traces(job["id"], run_id, trace.records)
    summary = trace.summary()
    complete_calculation_run(run_id, summary=summary)
    proposal["audit"] = {
        "run_id": run_id,
        "trace_count": trace_count,
        "summary": summary,
    }
    proposal["audit_source_fingerprint"] = _proposal_source_fingerprint(job, proposal)
    # A regenerate must never overwrite the accepted proposal before the UI has
    # merged its full manual structure and saved it. Initial generation is safe
    # to persist immediately because there is no accepted proposal yet.
    if not has_saved_accepted_proposal:
        try:
            prior_revision = max(0, int(existing_pd.get("_server_revision") or 0))
        except (TypeError, ValueError):
            prior_revision = 0
        proposal["_server_revision"] = prior_revision + 1
        saved_job = load_job(db_id) or job
        saved_job["proposal_data"] = proposal
        save_job(saved_job)
    return proposal


@app.post("/api/jobs/{job_id}/proposal/pdf")
async def api_generate_proposal_pdf(job_id: str, request: Request):
    """Generate proposal PDF from edited bundle data."""
    import json as _json
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    body = await request.json()
    _validate_proposal_pdf_ready(job, body)
    # Build proposal data from the edited bundles sent by frontend
    job_info = {
        "id": job["id"],
        "project_name": job["project_name"],
        "gc_name": job.get("gc_name"),
        "address": job.get("address"),
        "city": job.get("city"),
        "state": job.get("state"),
        "zip": job.get("zip"),
        "tax_rate": job.get("tax_rate", 0),
        "unit_count": job.get("unit_count", 0),
        "salesperson": job.get("salesperson") or "Standard Interiors",
    }

    proposal_data = {
        "job_info": job_info,
        "bundles": body.get("bundles", []),
        "subtotal": body.get("subtotal", 0),
        "tax_rate": body.get("tax_rate", 0),
        "tax_amount": body.get("tax_amount", 0),
        "grand_total": body.get("grand_total", 0),
        "gpm_pct": body.get("gpm_pct", 0),
        "gpm_profit": body.get("gpm_profit", 0),
        "gpm_labor": body.get("gpm_labor", 0),
        "gpm_material": body.get("gpm_material", 0),
        "manual_adjustment": body.get("manual_adjustment", 0),
        "textura_fee": body.get("textura_fee", 0),
        "textura_amount": body.get("textura_amount", 0),
        "notes": body.get("notes", []),
        "terms": body.get("terms", []),
        "exclusions": body.get("exclusions", []),
        "deleted_bundles": body.get("deleted_bundles", []),
        "deleted_bundle_reasons": body.get("deleted_bundle_reasons", {}),
        "deleted_material_codes": body.get("deleted_material_codes", []),
        "deleted_material_reasons": body.get("deleted_material_reasons", {}),
        "audit": body.get("audit", (job.get("proposal_data") or {}).get("audit", {})),
        "_client_session_id": body.get("_client_session_id") or (job.get("proposal_data") or {}).get("_client_session_id", ""),
        "_client_edit_version": body.get("_client_edit_version") or (job.get("proposal_data") or {}).get("_client_edit_version", 0),
        "_client_save_sequence": body.get("_client_save_sequence") or (job.get("proposal_data") or {}).get("_client_save_sequence", 0),
        "_server_revision": (job.get("proposal_data") or {}).get("_server_revision", 0),
        "pdf_generated_at": datetime.now(timezone.utc).isoformat(),
    }
    proposal_data["audit_source_fingerprint"] = _proposal_source_fingerprint(job, proposal_data)
    pdf_run = _latest_completed_run(job["id"], {"proposal_editor_save", "proposal_generation"})
    if pdf_run:
        proposal_data["pdf_audit_run_id"] = pdf_run["id"]
        proposal_data["pdf_ruleset_version"] = (pdf_run.get("metadata") or {}).get("ruleset_version")
        proposal_data["pdf_source_fingerprint"] = _proposal_source_fingerprint(job, proposal_data)
        proposal_data["pdf_totals"] = {
            "subtotal": body.get("subtotal", 0),
            "tax_amount": body.get("tax_amount", 0),
            "grand_total": body.get("grand_total", 0),
            "gpm_profit": body.get("gpm_profit", 0),
            "gpm_labor": body.get("gpm_labor", 0),
            "gpm_material": body.get("gpm_material", 0),
            "manual_adjustment": body.get("manual_adjustment", 0),
            "textura_amount": body.get("textura_amount", 0),
        }

    pdf_path = _job_pdf_path(job["id"], "proposal")
    generate_proposal_pdf(proposal_data, pdf_path)
    _record_artifact(job["id"], pdf_path, "proposal_pdf")

    # Save proposal data to job
    job["proposal_data"] = _json.dumps(proposal_data)
    save_job(job)

    log_activity(job["id"], "proposal_generated",
                 f"Proposal generated: {len(proposal_data['bundles'])} bundles, total ${proposal_data['grand_total']:,.2f}",
                 {"bundle_count": len(proposal_data["bundles"]), "grand_total": proposal_data["grand_total"]})

    return {"status": "ok", "pdf_url": f"/api/jobs/{job_id}/proposal.pdf"}


@app.get("/api/jobs/{job_id}/proposal.pdf")
def api_download_proposal_pdf(job_id: str):
    """Download the generated proposal PDF."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    _validate_proposal_pdf_download_ready(job)
    pdf_path = _job_pdf_path(job["id"], "proposal")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="PDF not found. Generate proposal first.")
    _require_artifact_receipt(job["id"], pdf_path, "proposal_pdf")
    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=f"{job['project_name']} Proposal.pdf",
    )


@app.post("/api/labor-catalog/upload")
async def api_upload_labor_catalog(file: UploadFile = File(...)):
    """Upload labor catalog (Excel or PDF)."""
    file_path = os.path.join(UPLOAD_DIR, f"labor_catalog_{file.filename}")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = os.path.splitext(file.filename)[1].lower()
    try:
        if ext == ".pdf":
            settings = get_settings()
            api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
            model = settings.get("openai_model", "gpt-5-mini")
            catalog = load_labor_catalog_from_pdf(file_path, api_key=api_key, model=model)
        elif ext in (".xlsx", ".xls"):
            catalog = load_labor_catalog(file_path)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}. Upload .pdf or .xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse labor catalog: {e}")

    return {"message": "Labor catalog loaded", "entries": len(catalog)}


# ── Exclusions ────────────────────────────────────────────────────────────────

class ExclusionsUpdate(BaseModel):
    exclusions: list[str]


@app.put("/api/jobs/{job_id}/exclusions")
def api_update_exclusions(job_id: str, body: ExclusionsUpdate):
    """Update job-specific exclusions list."""
    import json as _json
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job["exclusions"] = _json.dumps(body.exclusions)
    save_job(job)
    excl_count = len(body.exclusions) if body.exclusions else 0
    log_activity(job["id"], "exclusions_updated", f"Updated exclusions ({excl_count} items)")
    return {"message": "Exclusions saved", "exclusions": body.exclusions}


@app.get("/api/jobs/{job_id}/exclusions")
def api_get_exclusions(job_id: str):
    """Get job exclusions (custom or defaults)."""
    import json as _json
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    raw = job.get("exclusions")
    if raw:
        try:
            return {"exclusions": _json.loads(raw), "is_custom": True}
        except (ValueError, TypeError):
            pass
    return {"exclusions": EXCLUSIONS_TEMPLATE, "is_custom": False}


# ── Materials Export ──────────────────────────────────────────────────────────

@app.get("/api/jobs/{job_id}/materials/export")
def api_export_materials(job_id: str):
    """Export materials as CSV download."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    materials = job.get("materials", [])
    if not materials:
        raise HTTPException(status_code=400, detail="No materials to export")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Item Code", "Description", "Material Type", "Install Qty", "Unit",
        "Waste %", "Order Qty", "Unit Price", "Extended Cost"
    ])
    for m in materials:
        writer.writerow([
            m.get("item_code", ""),
            m.get("description", ""),
            m.get("material_type", ""),
            m.get("installed_qty", 0),
            m.get("unit", ""),
            f"{(m.get('waste_pct', 0) * 100):.0f}%",
            m.get("order_qty", 0),
            m.get("unit_price", 0),
            m.get("extended_cost", 0),
        ])

    output.seek(0)
    slug = job.get("slug", f"job-{job['id']}")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{slug}-materials.csv"'}
    )


# ── Labor Catalog ────────────────────────────────────────────────────────────

@app.get("/api/labor-catalog")
def api_get_labor_catalog():
    """Get the currently loaded labor catalog entries."""
    catalog = get_labor_catalog()
    return {"entries": catalog, "count": len(catalog)}


@app.get("/api/labor-catalog/stairs")
def api_get_stair_labor():
    """Get stair-related labor catalog entries."""
    catalog = get_labor_catalog()
    stair_entries = [e for e in catalog if 'stair' in (e.get('description') or '').lower()]
    return {"entries": stair_entries}


@app.post("/api/labor-catalog/entry")
def api_insert_labor_catalog_entry(body: dict):
    """Insert a single labor catalog entry; returns the new entry id."""
    from models import insert_labor_catalog_entry
    new_id = insert_labor_catalog_entry(body)
    return {"id": new_id}


@app.put("/api/labor-catalog/{entry_id}")
def api_update_labor_catalog_entry(entry_id: int, body: dict):
    """Update a single labor catalog entry."""
    if not update_labor_catalog_entry(entry_id, body):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry updated"}


@app.delete("/api/labor-catalog/{entry_id}")
def api_delete_labor_catalog_entry(entry_id: int):
    """Delete a single labor catalog entry."""
    if not delete_labor_catalog_entry(entry_id):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}


@app.delete("/api/labor-catalog")
def api_clear_labor_catalog():
    """Clear all labor catalog entries."""
    clear_labor_catalog()
    return {"message": "Labor catalog cleared"}


@app.get("/api/stair-sundry-kits")
def api_get_stair_sundry_kits():
    """Get stair sundry kit definitions (ratios per stair)."""
    return {"kits": STAIR_SUNDRY_KITS}


@app.get("/api/search")
def api_search(q: str = ""):
    """Global search across jobs and materials."""
    if not q or len(q) < 2:
        return {"jobs": [], "materials": []}
    return search_all(q)


# ── Company Rates ────────────────────────────────────────────────────────────

def _seed_company_rates():
    """Seed company rates from config.py defaults if not yet in DB."""
    import json as _json
    for rate_type, default_data in [
        ("sundry_rules", SUNDRY_RULES),
        ("waste_factors", WASTE_FACTORS),
        ("freight_rates", FREIGHT_RATES),
    ]:
        existing = get_company_rate(rate_type)
        if existing is None:
            save_company_rate(rate_type, _json.dumps(default_data))
            print(f"[seed] Seeded {rate_type} from config.py defaults")


@app.get("/api/company-rates")
def api_get_all_company_rates():
    """Get all company rates."""
    return get_all_company_rates()


@app.get("/api/company-rates/{rate_type}")
def api_get_company_rate(rate_type: str):
    """Get a specific company rate (sundry_rules, waste_factors, freight_rates)."""
    import json as _json
    data = get_company_rate(rate_type)
    if data is None:
        raise HTTPException(status_code=404, detail=f"Rate type '{rate_type}' not found")
    return {"rate_type": rate_type, "data": _json.loads(data)}


class CompanyRateUpdate(BaseModel):
    data: dict


@app.put("/api/company-rates/{rate_type}")
def api_update_company_rate(rate_type: str, body: CompanyRateUpdate):
    """Update a company rate."""
    import json as _json
    if rate_type not in ("sundry_rules", "waste_factors", "freight_rates", "sundry_prices"):
        raise HTTPException(status_code=400, detail=f"Invalid rate type: {rate_type}")
    save_company_rate(rate_type, _json.dumps(body.data))
    return {"message": f"{rate_type} updated"}


# ── Price List ───────────────────────────────────────────────────────────────

@app.get("/api/price-list")
def api_get_price_list():
    """Get all price list entries."""
    entries = get_price_list_entries()
    return {"entries": entries, "count": len(entries)}


class PriceListEntry(BaseModel):
    product_name: str
    material_type: Optional[str] = ""
    unit: Optional[str] = ""
    unit_price: Optional[float] = 0
    vendor: Optional[str] = ""
    notes: Optional[str] = ""


@app.post("/api/price-list")
def api_add_price_list_entry(body: PriceListEntry):
    """Add a single price list entry."""
    entry_id = add_price_list_entry(body.model_dump())
    return {"id": entry_id, "message": "Entry added"}


@app.put("/api/price-list/{entry_id}")
def api_update_price_list_entry(entry_id: int, body: PriceListEntry):
    """Update a price list entry."""
    if not update_price_list_entry(entry_id, body.model_dump()):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry updated"}


@app.delete("/api/price-list/{entry_id}")
def api_delete_price_list_entry(entry_id: int):
    """Delete a price list entry."""
    if not delete_price_list_entry(entry_id):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}


class PriceListBulkUpload(BaseModel):
    entries: list[dict]


@app.post("/api/price-list/bulk")
def api_bulk_upload_price_list(body: PriceListBulkUpload):
    """Replace all price list entries (bulk upload)."""
    save_price_list_entries(body.entries)
    return {"message": "Price list updated", "count": len(body.entries)}


@app.delete("/api/price-list")
def api_clear_price_list():
    """Clear all price list entries."""
    clear_price_list()
    return {"message": "Price list cleared"}


@app.post("/api/price-list/upload")
async def api_upload_price_list(file: UploadFile = File(...)):
    """Upload price list from CSV or Excel file."""
    import json as _json
    file_path = os.path.join(UPLOAD_DIR, f"price_list_{file.filename}")
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = os.path.splitext(file.filename)[1].lower()
    entries = []

    if ext == ".csv":
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                entries.append({
                    "product_name": row.get("product_name", row.get("Product Name", row.get("name", ""))),
                    "material_type": row.get("material_type", row.get("Material Type", row.get("type", ""))),
                    "unit": row.get("unit", row.get("Unit", "")),
                    "unit_price": float(row.get("unit_price", row.get("Unit Price", row.get("price", 0))) or 0),
                    "vendor": row.get("vendor", row.get("Vendor", "")),
                    "notes": row.get("notes", row.get("Notes", "")),
                })
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))
            entries.append({
                "product_name": str(row_dict.get("product_name", row_dict.get("product name", row_dict.get("name", ""))) or ""),
                "material_type": str(row_dict.get("material_type", row_dict.get("material type", row_dict.get("type", ""))) or ""),
                "unit": str(row_dict.get("unit", "") or ""),
                "unit_price": float(row_dict.get("unit_price", row_dict.get("unit price", row_dict.get("price", 0))) or 0),
                "vendor": str(row_dict.get("vendor", "") or ""),
                "notes": str(row_dict.get("notes", "") or ""),
            })
        wb.close()
    elif ext == ".pdf":
        # Parse price list PDF using AI
        settings = get_settings()
        api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
        model = settings.get("openai_model", "gpt-5-mini")
        entries = _parse_price_list_pdf(file_path, api_key, model)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}. Upload .csv, .xlsx, or .pdf")

    if not entries:
        raise HTTPException(status_code=400, detail="No entries found in uploaded file")

    save_price_list_entries(entries)
    return {"message": "Price list uploaded", "count": len(entries)}


def _parse_price_list_pdf(file_path: str, api_key: str = None, model: str = None) -> list[dict]:
    """Parse a price list PDF using AI."""
    if model is None:
        settings = get_settings()
        model = settings.get("openai_model", "gpt-5-mini")
    import pdfplumber

    text_parts = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    text = "\n".join(text_parts)
    if not text.strip():
        return []

    import json as _json

    system_msg = """You are parsing a material price list for a flooring/interiors company.
Extract every product/material entry into a JSON array.

For each entry, extract:
- product_name: the product name or description
- material_type: the flooring type if identifiable (e.g. "unit_lvt", "unit_carpet_no_pattern", "floor_tile", etc.)
- unit: the unit of measure (SF, SY, LF, EA, etc.)
- unit_price: the price per unit as a number
- vendor: the vendor/manufacturer if shown
- notes: any additional notes

Return JSON: {"entries": [{"product_name": "...", "material_type": "...", "unit": "...", "unit_price": 0.00, "vendor": "...", "notes": "..."}, ...]}"""

    raw = chat_complete(
        system=system_msg,
        user=text,
        api_key=api_key,
        model=model,
        json_mode=True,
    )
    parsed = _json.loads(raw)
    return parsed.get("entries", [])


# ── Settings ─────────────────────────────────────────────────────────────────

@app.get("/api/settings")
def api_get_settings():
    """Get app settings (API key is masked)."""
    settings = get_settings()
    # Mask API keys for display
    raw_key = settings.get("openai_api_key", "")
    if raw_key and len(raw_key) > 8:
        masked = raw_key[:4] + "•" * (len(raw_key) - 8) + raw_key[-4:]
    elif raw_key:
        masked = "•" * len(raw_key)
    else:
        masked = ""

    anthropic_key = settings.get("anthropic_api_key", "") or os.environ.get("ANTHROPIC_API_KEY", "")
    if anthropic_key and len(anthropic_key) > 8:
        anthropic_masked = anthropic_key[:4] + "•" * (len(anthropic_key) - 8) + anthropic_key[-4:]
    elif anthropic_key:
        anthropic_masked = "•" * len(anthropic_key)
    else:
        anthropic_masked = ""

    provider = get_provider_info(raw_key)
    return {
        "openai_api_key_set": bool(raw_key),
        "openai_api_key_masked": masked,
        "anthropic_api_key_set": bool(anthropic_key),
        "anthropic_api_key_masked": anthropic_masked,
        "openai_model": settings.get("openai_model", "gpt-5-mini"),
        "multi_pass_count": int(settings.get("multi_pass_count", "2")),
        "email_automation_enabled": settings.get("email_automation_enabled", "false"),
        "email_config": settings.get("email_config", ""),
        "bid_folder_path": settings.get("bid_folder_path", ""),
        "ai_provider": provider["provider"],
        "ai_available": provider["available"],
        "vendor_quote_test_mode": settings.get("vendor_quote_test_mode", "false"),
    }


@app.post("/api/settings")
def api_update_settings(body: SettingsUpdate):
    """Update app settings."""
    updates = {}
    if body.openai_api_key is not None:
        updates["openai_api_key"] = body.openai_api_key
    if body.anthropic_api_key is not None:
        updates["anthropic_api_key"] = body.anthropic_api_key
    if body.openai_model is not None:
        if body.openai_model not in ("gpt-5-mini", "gpt-5.4"):
            raise HTTPException(status_code=400, detail="Invalid model. Choose gpt-5-mini or gpt-5.4")
        updates["openai_model"] = body.openai_model
    if body.multi_pass_count is not None:
        if body.multi_pass_count < 1 or body.multi_pass_count > 5:
            raise HTTPException(status_code=400, detail="Multi-pass count must be between 1 and 5")
        updates["multi_pass_count"] = str(body.multi_pass_count)
    # Email automation settings
    if body.email_automation_enabled is not None:
        updates["email_automation_enabled"] = body.email_automation_enabled
    if body.email_config is not None:
        updates["email_config"] = body.email_config
    if body.bid_folder_path is not None:
        updates["bid_folder_path"] = body.bid_folder_path
    if body.vendor_quote_test_mode is not None:
        updates["vendor_quote_test_mode"] = body.vendor_quote_test_mode
    if updates:
        save_settings(updates)
        # Apply API key and model to quote parser
        settings = get_settings()
        _apply_openai_config(settings)
        # Restart inbox monitor if email settings changed
        if body.email_automation_enabled is not None or body.email_config is not None:
            _start_inbox_monitor()
        # Restart sim watcher if test mode changed
        if body.vendor_quote_test_mode is not None:
            _start_sim_watcher()
            _start_inbox_monitor()  # Re-evaluate: stop real monitor if test mode on
    return {"message": "Settings updated", **api_get_settings()}


def _apply_openai_config(settings: dict = None):
    """Apply stored AI settings to the quote parser and ai_client."""
    if settings is None:
        settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")
    passes = int(settings.get("multi_pass_count", "2"))

    # If Anthropic key is stored in settings, set it in the environment
    # so ai_client.py can detect it as a fallback
    anthropic_key = settings.get("anthropic_api_key") or os.environ.get("ANTHROPIC_API_KEY")
    if anthropic_key and not os.environ.get("ANTHROPIC_API_KEY"):
        os.environ["ANTHROPIC_API_KEY"] = anthropic_key

    provider = get_provider_info(api_key)
    print(f"[ai_config] provider={provider['provider']}, openai_key={'set' if api_key else 'MISSING'}, "
          f"anthropic_key={'set' if anthropic_key else 'MISSING'}, model={model}, passes={passes}")
    set_openai_config(api_key=api_key, model=model, num_passes=passes)


# ── Vendor Pricing Intelligence ───────────────────────────────────────────────

@app.get("/api/vendors")
async def api_list_vendors():
    return list_vendors()


@app.get("/api/vendors/{vendor_id}")
async def api_get_vendor(vendor_id: int):
    v = get_vendor(vendor_id)
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return v


@app.post("/api/vendors")
async def api_create_vendor(request: Request):
    data = await request.json()
    try:
        vendor = create_vendor(data)
        return vendor
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/vendors/{vendor_id}")
async def api_update_vendor(vendor_id: int, request: Request):
    data = await request.json()
    update_vendor(vendor_id, data)
    return {"ok": True}


@app.delete("/api/vendors/{vendor_id}")
async def api_delete_vendor(vendor_id: int):
    if delete_vendor(vendor_id):
        return {"ok": True}
    raise HTTPException(status_code=404, detail="Vendor not found")


@app.post("/api/vendors/merge")
async def api_merge_vendors(body: dict):
    keep_id = body.get("keep_id")
    merge_ids = body.get("merge_ids", [])
    if not keep_id or not merge_ids:
        raise HTTPException(status_code=400, detail="keep_id and merge_ids required")
    from models import merge_vendors
    merge_vendors(keep_id, merge_ids)
    return {"ok": True}


@app.get("/api/vendor-prices")
async def api_search_vendor_prices(vendor: str = None, product: str = None, limit: int = 50):
    return search_vendor_prices(vendor, product, limit)


@app.post("/api/vendor-prices/import")
async def api_import_vendor_prices(file: UploadFile = File(...)):
    """Bulk import vendor prices from CSV. Accepts partial data - only product_name and unit_price required."""
    contents = await file.read()
    text = contents.decode('utf-8-sig')  # handle BOM
    result = import_vendor_prices_csv(text)
    return result


@app.get("/api/materials/price-history")
async def api_price_history(item_code: str = None, product: str = None, exclude_job: int = None):
    return get_price_history(item_code, product, exclude_job)


# ── Quote Requests ───────────────────────────────────────────────────────────

@app.post("/api/jobs/{job_id}/quote-requests")
async def api_create_quote_request(job_id: str, request: Request):
    db_id = _resolve_job_id(job_id)
    data = await request.json()
    vendor_name = data.get("vendor_name", "").strip()
    if not vendor_name:
        raise HTTPException(status_code=400, detail="vendor_name is required")
    status = data.get("status", "draft")
    sent_at = data.get("sent_at")
    qr = create_quote_request(
        job_id=db_id,
        vendor_name=vendor_name,
        material_ids=data.get("material_ids", []),
        request_text=data.get("request_text", ""),
        vendor_id=data.get("vendor_id"),
        status=status,
        sent_at=sent_at,
    )
    return qr


@app.get("/api/jobs/{job_id}/quote-requests")
async def api_list_quote_requests(job_id: str):
    db_id = _resolve_job_id(job_id)
    return list_quote_requests(db_id)


@app.put("/api/quote-requests/{request_id}")
async def api_update_quote_request(request_id: int, request: Request):
    data = await request.json()
    update_quote_request(request_id, **data)
    return {"ok": True}


@app.delete("/api/quote-requests/{request_id}")
async def api_delete_quote_request(request_id: int):
    if delete_quote_request(request_id):
        return {"ok": True}
    raise HTTPException(status_code=404, detail="Quote request not found")


# ── AI: Vendor Detection & Quote Text ────────────────────────────────────────

def _build_vendor_memory() -> dict:
    """Build vendor knowledge from DB history — no hardcoded lists.

    Returns:
      - known_names: vendor names from vendors table
      - aliases: lowercase name → canonical name (shortest/cleanest wins)
      - product_hints: product_normalized → vendor_name
    """
    conn = _get_conn()
    try:
        vendors = conn.execute("SELECT id, name FROM vendors").fetchall()
        vendor_id_map = {v["id"]: v["name"] for v in vendors}

        # Build alias map: prefer shorter names as canonical
        # e.g., "Daltile" (7 chars) beats "Dal-Tile" (8 chars)
        aliases = {}
        def _add_alias(name_str: str, canonical: str):
            """Register a name and its cleaned variants, don't overwrite existing."""
            for v in {name_str.lower(), name_str.lower().replace("-", "").replace(" ", "")}:
                if v not in aliases:
                    aliases[v] = canonical

        # Shorter names first → they become canonical
        for v in sorted(vendors, key=lambda v: (len(v["name"]), v["name"])):
            name = v["name"]
            # If cleaned form already maps somewhere, use THAT as canonical
            # (so "Dal-Tile" → cleaned "daltile" → already maps to "Daltile")
            canonical = aliases.get(name.lower().replace("-", "").replace(" ", ""), name)
            _add_alias(name, canonical)

        # Learn from vendor_prices (same vendor_id, different name strings)
        vp_names = conn.execute("""
            SELECT DISTINCT vendor_id, vendor_name FROM vendor_prices
            WHERE vendor_id IS NOT NULL AND vendor_name IS NOT NULL
        """).fetchall()
        for vp in vp_names:
            raw_canonical = vendor_id_map.get(vp["vendor_id"])
            if raw_canonical:
                # Resolve through existing aliases
                canonical = aliases.get(raw_canonical.lower().replace("-", "").replace(" ", ""), raw_canonical)
                _add_alias(vp["vendor_name"], canonical)

        # Product hints: which vendor historically supplies which products
        product_hints = {}
        for row in conn.execute("""
            SELECT product_normalized, vendor_name, COUNT(*) as cnt
            FROM vendor_prices
            WHERE product_normalized IS NOT NULL AND vendor_name IS NOT NULL
            GROUP BY product_normalized, vendor_name ORDER BY cnt DESC
        """).fetchall():
            pn = row["product_normalized"]
            if pn and pn not in product_hints:
                product_hints[pn] = row["vendor_name"]

        # Vendor catalog: what products each vendor sells (for AI context)
        vendor_catalog = {}
        for row in conn.execute("""
            SELECT vendor_name, product_name
            FROM vendor_prices
            WHERE vendor_name IS NOT NULL AND product_name IS NOT NULL
            ORDER BY vendor_name, product_name
        """).fetchall():
            vn = row["vendor_name"]
            # Resolve through aliases
            canonical = aliases.get(vn.lower().replace("-", "").replace(" ", ""), vn)
            vendor_catalog.setdefault(canonical, []).append(row["product_name"])

        return {
            "known_names": [v["name"] for v in vendors],
            "aliases": aliases,
            "product_hints": product_hints,
            "vendor_catalog": vendor_catalog,
        }
    finally:
        conn.close()


def _fuzzy_match_vendor(text: str, aliases: dict, threshold: int = 2) -> str | None:
    """Match text against known vendor aliases. Handles case, punctuation, typos, substrings."""
    if not text or len(text) < 2:
        return None
    t = text.lower().strip()

    # 1. Exact / cleaned match
    if t in aliases:
        return aliases[t]
    t_clean = t.replace("-", "").replace(".", "").replace(",", "").replace(" ", "")
    for alias, canonical in aliases.items():
        if t_clean == alias.replace("-", "").replace(".", "").replace(",", "").replace(" ", ""):
            return canonical

    # 2. Substring (one contains the other, min 4 chars)
    if len(t) >= 4:
        for alias, canonical in aliases.items():
            if len(alias) >= 4 and (alias in t or t in alias):
                return canonical

    # 3. Levenshtein for typos (e.g., "Dalitle" ↔ "daltile")
    if len(t) >= 5:
        best, best_dist = None, threshold + 1
        for alias, canonical in aliases.items():
            if len(alias) < 4 or abs(len(alias) - len(t)) > threshold:
                continue
            # Inline Levenshtein
            s1, s2 = (t, alias) if len(t) >= len(alias) else (alias, t)
            prev = list(range(len(s2) + 1))
            for i, c1 in enumerate(s1):
                curr = [i + 1]
                for j, c2 in enumerate(s2):
                    curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (c1 != c2)))
                prev = curr
            if prev[-1] <= threshold and prev[-1] < best_dist:
                best, best_dist = canonical, prev[-1]
        if best:
            return best
    return None


def _quick_regex_extract(description: str) -> list[str]:
    """Extract vendor candidates from dash-separated material descriptions.
    No hardcoded vendor names — pure structural parsing.

    "F109 - Interface - Breakout - ..." → ["Interface"]
    "(Scheme A) Daltile - Modern Hearth - ..." → ["Daltile"]
    "Schluter - Dilex-AHKA Cove - ..." → ["Schluter"]
    """
    import re
    if not description:
        return []
    desc = re.sub(r'^\(Scheme\s+[^)]+\)\s*', '', description.strip())
    segments = [s.strip() for s in desc.split(' - ') if s.strip()]
    candidates = []

    # "ItemCode - Vendor - Product - ..." (3+ segments, first is a code)
    if len(segments) >= 3 and re.match(r'^[A-Za-z0-9/]{1,20}$', segments[0]):
        candidates.append(segments[1])

    # "Vendor - Product - ..." (first segment IS the vendor)
    if len(segments) >= 2:
        first = segments[0]
        if (2 <= len(first) <= 50
            and not re.match(r'^[A-Z]\d+$', first)
            and not re.match(r'^[\d"\'x\s.]+$', first)
            and not first.lower().startswith(('transition', 'horizontal', 'vertical'))
            and first not in candidates):
            candidates.append(first)

    return candidates


@app.post("/api/jobs/{job_id}/detect-vendors")
async def api_detect_vendors(job_id: str):
    """Hybrid vendor detection: memory + regex + AI.

    1. MEMORY: Check vendor_prices history and vendors table for known names/aliases
    2. REGEX: Extract candidate from description structure (no hardcoded names)
    3. FUZZY MATCH: Match regex candidate against memory (handles typos)
    4. AI: For unresolved items, ask AI with evidence requirement
    5. VALIDATE: Cross-check AI results — reject if evidence doesn't appear in description
    6. RETRY: Focused single-item AI call for any conflicts
    """
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    materials = job.get("materials", [])
    if not materials:
        return {"vendors": {}, "materials": []}

    import json

    # ── Build vendor memory from DB history ──
    memory = _build_vendor_memory()
    aliases = memory["aliases"]
    known_names = memory["known_names"]
    product_hints = memory["product_hints"]
    vendor_catalog = memory["vendor_catalog"]

    vendor_groups = {}
    needs_ai = []  # indices still unresolved after regex+memory
    method_log = {}  # track how each was resolved for debugging

    for i, m in enumerate(materials):
        desc = m.get("description", "")

        # Step 1: Try regex extraction + fuzzy match against memory
        candidates = _quick_regex_extract(desc)
        resolved = False
        best_candidate = ""
        for candidate in candidates:
            matched = _fuzzy_match_vendor(candidate, aliases)
            if matched:
                materials[i]["vendor"] = matched
                vendor_groups.setdefault(matched, []).append(i)
                method_log[i] = "regex+memory"
                resolved = True
                break
            if not best_candidate:
                best_candidate = candidate
        if resolved:
            continue
        if best_candidate:
            # Candidate extracted but not in memory — send to AI for confirmation
            materials[i]["_regex_candidate"] = best_candidate

        # Step 2: Check product hints from vendor_prices history
        item_code = (m.get("item_code") or "").strip()
        normalized = _normalize_product(item_code or desc)
        if normalized and normalized in product_hints:
            hint_vendor = product_hints[normalized]
            canonical = _fuzzy_match_vendor(hint_vendor, aliases) or hint_vendor
            # Only use product hint if no regex candidate contradicts it
            if not best_candidate or best_candidate.lower() in canonical.lower() or canonical.lower() in best_candidate.lower():
                materials[i]["vendor"] = canonical
                vendor_groups.setdefault(canonical, []).append(i)
                method_log[i] = "product_history"
                continue

        # Step 3: If material already has a vendor set, normalize it through memory
        if m.get("vendor"):
            normalized_vendor = _fuzzy_match_vendor(m["vendor"], aliases) or m["vendor"]
            materials[i]["vendor"] = normalized_vendor
            vendor_groups.setdefault(normalized_vendor, []).append(i)
            method_log[i] = "existing"
            continue

        # Unresolved — need AI
        needs_ai.append({
            "index": i,
            "item_code": item_code,
            "description": desc,
            "material_type": m.get("material_type", ""),
            "regex_candidate": materials[i].pop("_regex_candidate", ""),
        })

    # ── AI pass: verify fast-pass results + resolve unknowns ──
    ai_resolved = 0
    ai_corrected = 0

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if provider["available"]:
        # Send ALL materials to AI — fast-pass results shown as "pre-assigned"
        # AI's job: confirm correct ones, correct wrong ones, fill in unknowns
        mat_lines = []
        for i, m in enumerate(materials):
            desc = m.get("description", "")
            item_code = (m.get("item_code") or "").strip()
            vendor = m.get("vendor", "")
            method = method_log.get(i, "")
            regex_cand = m.pop("_regex_candidate", "")
            line = f'{i}. [{item_code}] {desc}'
            if vendor:
                line += f'  [pre-assigned: {vendor} via {method}]'
            elif regex_cand:
                line += f'  [regex candidate: {regex_cand}]'
            else:
                line += '  [UNASSIGNED]'
            mat_lines.append(line)

        # Build vendor catalog context for AI
        catalog_lines = []
        for vname, products in sorted(vendor_catalog.items()):
            # Deduplicate and limit to 10 products per vendor
            unique = list(dict.fromkeys(products))[:10]
            catalog_lines.append(f"  {vname}: {', '.join(unique)}")
        catalog_ctx = "\n".join(catalog_lines) if catalog_lines else "  (no price history yet)"

        prompt = (
            f"You are a commercial flooring industry expert reviewing vendor assignments for a material list.\n\n"
            f"KNOWN VENDORS AND WHAT THEY SELL:\n{catalog_ctx}\n\n"
            f"Other known vendors (no price history yet): "
            f"{', '.join(n for n in known_names if n not in vendor_catalog) or 'None'}\n\n"
            f"MATERIALS TO REVIEW:\n"
            f"{chr(10).join(mat_lines)}\n\n"
            f"YOUR JOB:\n"
            f"1. VERIFY pre-assigned vendors — if the description clearly says a different vendor, CORRECT it\n"
            f"2. FILL IN unassigned items — identify vendor from the description\n"
            f"3. CHECK CONSISTENCY — if 5 items are 'Interface - Woven Gradience' and 1 similar item got assigned differently, flag it\n"
            f"4. USE PRODUCT KNOWLEDGE — match product names/styles to known vendor catalogs above\n"
            f"5. SKIP genuinely generic items (transitions, trims with no vendor name, TBD specs)\n\n"
            f"RULES:\n"
            f"- Description is the source of truth. Format is usually 'ItemCode - VendorName - Product - ...'\n"
            f"- Only OVERRIDE a pre-assigned vendor if you have clear evidence it's wrong\n"
            f"- Include 'evidence': exact text from the description proving the vendor\n"
            f"- For corrections, set 'correction': true and explain WHY in 'reason'\n\n"
            f'Return JSON: {{"results": [{{"index": 0, "vendor": "Name", "evidence": "exact text", "correction": false, "reason": ""}}]}}\n'
            f"Only include items where you're adding a vendor OR correcting a wrong one. Skip confirmed-correct items."
        )

        try:
            raw = chat_complete(
                system="You are a commercial flooring industry expert.",
                user=prompt,
                api_key=api_key,
                model=model,
                json_mode=True,
            )
            result = json.loads(raw)

            detections = result if isinstance(result, list) else next(
                (result[k] for k in ("results", "vendors", "detections", "materials", "data")
                 if k in result and isinstance(result[k], list)), []
            )

            for det in detections:
                idx = det.get("index", -1)
                vendor = (det.get("vendor") or "").strip()
                evidence = (det.get("evidence") or "").strip().lower()
                is_correction = det.get("correction", False)

                if not vendor or idx < 0 or idx >= len(materials):
                    continue

                # Validate: evidence must appear in description
                desc_lower = (materials[idx].get("description") or "").lower()
                vl = vendor.lower()
                evidence_ok = (not evidence or evidence in desc_lower
                               or vl in desc_lower or vl.replace("-", "") in desc_lower.replace("-", ""))

                if not evidence_ok:
                    continue  # AI claim doesn't check out — keep fast-pass result

                canonical = _fuzzy_match_vendor(vendor, aliases) or vendor
                old_vendor = materials[idx].get("vendor", "")

                if is_correction and old_vendor:
                    # AI is overriding fast-pass — remove from old group
                    if old_vendor in vendor_groups and idx in vendor_groups[old_vendor]:
                        vendor_groups[old_vendor].remove(idx)
                        if not vendor_groups[old_vendor]:
                            del vendor_groups[old_vendor]
                    materials[idx]["vendor"] = canonical
                    vendor_groups.setdefault(canonical, []).append(idx)
                    method_log[idx] = "ai_corrected"
                    ai_corrected += 1
                elif not old_vendor:
                    # Filling in an unassigned item
                    materials[idx]["vendor"] = canonical
                    vendor_groups.setdefault(canonical, []).append(idx)
                    method_log[idx] = "ai"
                    ai_resolved += 1

        except Exception as e:
            print(f"AI vendor detection failed: {e}")

    # Clean up any leftover regex candidates
    for m in materials:
        m.pop("_regex_candidate", None)

    # ── Learning: save new vendor-product associations for future jobs ──
    # When AI fills in or corrects a vendor, learn the association
    conn = _get_conn()
    try:
        for i, m in enumerate(materials):
            if method_log.get(i) in ("ai", "ai_corrected") and m.get("vendor"):
                item_code = (m.get("item_code") or "").strip()
                desc = m.get("description", "")
                product_name = desc[:100] if desc else item_code
                normalized = _normalize_product(item_code or desc)
                if normalized:
                    # Check if this product-vendor pair already exists
                    existing = conn.execute(
                        "SELECT id FROM vendor_prices WHERE product_normalized=? AND vendor_name=? LIMIT 1",
                        (normalized, m["vendor"])
                    ).fetchone()
                    if not existing:
                        vendor_id = conn.execute(
                            "SELECT id FROM vendors WHERE name=? LIMIT 1", (m["vendor"],)
                        ).fetchone()
                        conn.execute("""
                            INSERT INTO vendor_prices (product_name, product_normalized, vendor_name, vendor_id, job_id, unit_price, unit, quote_date, notes, created_at)
                            VALUES (?, ?, ?, ?, ?, 0, '', datetime('now'), 'Auto-learned from AI vendor detection', datetime('now'))
                        """, (product_name, normalized, m["vendor"], vendor_id["id"] if vendor_id else None, db_id))
        conn.commit()
    except Exception as e:
        print(f"Learning save failed: {e}")
    finally:
        conn.close()

    # Save updated vendor fields
    save_materials(db_id, materials)

    return {
        "vendor_groups": vendor_groups,
        "total_detected": sum(1 for m in materials if m.get("vendor")),
        "resolved_by": {
            "memory": sum(1 for v in method_log.values() if v in ("regex+memory", "product_history")),
            "existing": sum(1 for v in method_log.values() if v == "existing"),
            "ai": ai_resolved,
            "ai_corrected": ai_corrected,
        },
    }


@app.post("/api/jobs/{job_id}/generate-quote-text")
async def api_generate_quote_text(job_id: str, request: Request):
    """Use AI to generate professional quote request text for a vendor."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    data = await request.json()
    vendor_name = data.get("vendor_name", "")
    material_indices = data.get("material_indices", [])
    is_follow_up = data.get("follow_up", False)
    days_since_sent = data.get("days_since_sent", 0)

    materials = job.get("materials", [])
    selected = [materials[i] for i in material_indices if 0 <= i < len(materials)]

    # For follow-ups, we don't need materials selected
    if not selected and not is_follow_up:
        raise HTTPException(status_code=400, detail="No materials selected")

    # Get past vendor prices for context
    conn = _get_conn()
    try:
        rows = conn.execute(
            """SELECT product_name, unit_price, unit, created_at
               FROM vendor_prices WHERE vendor_name LIKE ? AND unit_price > 0
               ORDER BY created_at DESC LIMIT 20""",
            (f"%{vendor_name}%",)
        ).fetchall()
        past_prices = [dict(r) for r in rows]
    finally:
        conn.close()

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if not provider["available"]:
        raise HTTPException(status_code=400, detail="No AI API key configured (set OpenAI or ANTHROPIC_API_KEY)")

    mat_lines = []
    for i, m in enumerate(selected, 1):
        qty = round((m.get("order_qty") or m.get("installed_qty") or 0) * 100) / 100
        desc = m.get("description", m.get("item_code", "Unknown"))
        item_code = m.get("item_code", "")
        unit = m.get("unit", "")
        mat_lines.append(f"{i}. {desc}\n   Item: {item_code} | Qty: {qty:,.2f} {unit}")

    location_parts = [job.get("address"), job.get("city"), job.get("state"), job.get("zip")]
    location = ", ".join(p for p in location_parts if p)

    past_context = ""
    if past_prices:
        past_context = f"\n\nPast pricing from {vendor_name} (for reference, DO NOT include in the email):\n"
        for pp in past_prices[:5]:
            past_context += f"  - {pp['product_name']}: ${pp['unit_price']}/{pp.get('unit', 'unit')} ({pp.get('created_at', '')[:10]})\n"

    if is_follow_up:
        prompt = f"""Write a polite but firm follow-up email to {vendor_name} regarding a Request for Pricing we sent {days_since_sent} days ago.

Project: {job.get('project_name', '')}
Location: {location}
GC: {job.get('gc_name', '')}

Requirements:
- Reference the original request sent {days_since_sent} days ago
- Politely ask for a status update on pricing
- Mention we need pricing to complete our bid
- Be professional, concise, and not passive-aggressive
- Do NOT include subject line, greeting name, or signature — just the body text
- Keep it to 3-4 sentences"""
    else:
        prompt = f"""Write a professional Request for Pricing email body for a commercial flooring project.
This is being sent to {vendor_name}.

Project: {job.get('project_name', '')}
Architect: {job.get('architect', '')}
Designer: {job.get('designer', '')}
Location: {location}
GC: {job.get('gc_name', '')}

Materials requiring pricing ({len(selected)} items):
{chr(10).join(mat_lines)}
{past_context}

Write a clean, professional email body. Requirements:
- Number each material for easy reference
- Include quantities and units
- Ask for: unit pricing, freight/delivery, lead times, and quote validity period
- Be professional but not overly formal — this is a normal vendor relationship
- Use proper flooring industry terminology
- Keep it concise — vendors receive many of these
- Do NOT include subject line, greeting name, or signature — just the body text
- Start with a brief intro about the project"""

    try:
        text = chat_complete(
            system="You are a professional flooring estimator composing vendor emails.",
            user=prompt,
            api_key=api_key,
            model=model,
        ).strip()
        return {"text": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI quote text generation failed: {str(e)}")


@app.post("/api/jobs/{job_id}/suggest-vendors")
async def api_suggest_vendors(job_id: str, request: Request):
    """AI suggests which vendor to contact for unassigned materials based on history."""
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    data = await request.json()
    material_indices = data.get("material_indices", [])
    materials = job.get("materials", [])
    unassigned = []
    for i in material_indices:
        if 0 <= i < len(materials):
            m = materials[i]
            unassigned.append({
                "index": i,
                "item_code": m.get("item_code", ""),
                "description": m.get("description", ""),
                "material_type": m.get("material_type", ""),
            })

    if not unassigned:
        return {"suggestions": []}

    # Build vendor history context from vendor_prices
    conn = _get_conn()
    try:
        rows = conn.execute(
            """SELECT vendor_name, product_name, product_normalized, unit_price, unit
               FROM vendor_prices
               WHERE unit_price > 0
               ORDER BY created_at DESC LIMIT 200"""
        ).fetchall()
        vendor_history = {}
        for r in rows:
            vn = r["vendor_name"]
            if vn not in vendor_history:
                vendor_history[vn] = []
            if len(vendor_history[vn]) < 10:
                vendor_history[vn].append(r["product_name"])
    finally:
        conn.close()

    known_vendors = list_vendors()
    vendor_summary = []
    for v in known_vendors:
        products = vendor_history.get(v["name"], [])
        vendor_summary.append(f"- {v['name']}: {', '.join(products[:5]) if products else 'no quote history'}")

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if not provider["available"]:
        raise HTTPException(status_code=400, detail="No AI API key configured (set OpenAI or ANTHROPIC_API_KEY)")

    prompt = f"""You are a commercial flooring industry expert. Given these unassigned materials and our vendor history, suggest which vendor to contact for each material.

Unassigned materials:
{chr(10).join(f'{m["index"]}. [{m["item_code"]}] {m["description"]} (type: {m["material_type"]})' for m in unassigned)}

Known vendors and what they've quoted before:
{chr(10).join(vendor_summary) if vendor_summary else 'No vendor history yet.'}

For each material, suggest the most likely vendor based on:
1. Product name/brand recognition (e.g., "Johnsonite" in the name = Johnsonite vendor)
2. Material type (carpet tile → Interface/Shaw, LVT → Shaw/Mannington, rubber base → Johnsonite)
3. Past vendor history matches

Return ONLY a JSON array: [{{"material_index": 0, "suggested_vendor": "Vendor Name", "reason": "brief reason"}}]
If you cannot suggest a vendor, omit that material from the array."""

    try:
        raw = chat_complete(
            system="You are a commercial flooring industry expert.",
            user=prompt,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        result = json.loads(raw)
        if isinstance(result, list):
            suggestions = result
        else:
            suggestions = []
            for key in ("suggestions", "results", "data", "result", "materials"):
                if key in result and isinstance(result[key], list):
                    suggestions = result[key]
                    break
            if not suggestions and all(k.isdigit() for k in result.keys()):
                suggestions = [v for _, v in sorted(result.items(), key=lambda x: int(x[0]))]

        return {"suggestions": suggestions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI vendor suggestion failed: {str(e)}")


@app.post("/api/vendors/suggest-contacts")
async def api_suggest_vendor_contacts(request: Request):
    """Use AI to suggest contact info for vendors."""
    data = await request.json()
    vendor_names = data.get("vendor_names", [])
    if not vendor_names:
        return {"suggestions": []}

    settings = get_settings()
    api_key = settings.get("openai_api_key") or os.environ.get("OPENAI_API_KEY")
    model = settings.get("openai_model", "gpt-5-mini")

    provider = get_provider_info(api_key)
    if not provider["available"]:
        raise HTTPException(status_code=400, detail="No AI API key configured (set OpenAI or ANTHROPIC_API_KEY)")

    prompt = f"""You are a commercial flooring industry expert. For each flooring manufacturer/vendor listed below,
provide helpful contact information and suggestions for finding a sales rep.

Vendors:
{chr(10).join(f'- {name}' for name in vendor_names)}

For each vendor, provide:
1. Their website URL (if a well-known manufacturer)
2. How to find a local sales rep (e.g., "visit interface.com/find-a-rep")
3. A general contact email if publicly known
4. What products they're known for (e.g., "carpet tile", "LVT", "rubber base")
5. Any helpful notes for a flooring estimator

Return ONLY a JSON array:
[{{"vendor": "Name", "website": "url", "find_rep_url": "url or instruction", "general_email": "email or empty", "products": "what they sell", "notes": "helpful tip"}}]"""

    try:
        raw = chat_complete(
            system="You are a commercial flooring industry expert.",
            user=prompt,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        result = json.loads(raw)
        suggestions = result if isinstance(result, list) else result.get("suggestions", result.get("vendors", []))
        return {"suggestions": suggestions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI vendor suggestion failed: {str(e)}")


# ── Notifications ────────────────────────────────────────────────────────────

@app.get("/api/notifications")
async def api_get_notifications(unread_only: bool = True):
    return get_notifications(unread_only)


@app.put("/api/notifications/{notification_id}/read")
async def api_mark_notification_read(notification_id: int):
    mark_notification_read(notification_id)
    return {"ok": True}


# ── Activity & Comments ──────────────────────────────────────────────────────

@app.get("/api/jobs/{job_id}/activity")
def api_get_activity(job_id: str):
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return get_activity(job["id"])


@app.get("/api/jobs/{job_id}/comments")
def api_get_comments(job_id: str):
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return get_comments(job["id"])


class CommentCreate(BaseModel):
    text: str

@app.post("/api/jobs/{job_id}/comments")
def api_add_comment(job_id: str, body: CommentCreate):
    db_id = _resolve_job_id(job_id)
    job = load_job(db_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    comment = add_comment(job["id"], body.text)
    log_activity(job["id"], "comment_added", "Comment added", {"text": body.text})
    return comment


# ── Price Book Endpoints ─────────────────────────────────────────────────────


@app.get("/api/price-book")
def api_price_book_summary():
    """Get summary of all imported price books."""
    return get_price_book_summary()


@app.post("/api/price-book/import")
def api_import_price_book(body: dict = Body(...)):
    """Import a vendor price book from JSON data.
    Body: {vendor, discount_pct, items: [{product_line, item_no, ...}]}
    """
    vendor = body.get("vendor")
    discount_pct = body.get("discount_pct", 0)
    items = body.get("items", [])
    category = body.get("category", "")
    if not vendor or not items:
        raise HTTPException(status_code=400, detail="vendor and items required")
    count = import_price_book(vendor, items, discount_pct, category)
    return {"imported": count, "vendor": vendor}


@app.get("/api/price-book/search")
def api_search_price_book(q: str = "", vendor: str = None):
    """Search price book items."""
    if not q:
        return []
    return search_price_book(q, vendor)


@app.post("/api/price-book/import-schluter")
def api_import_schluter():
    """Import the pre-parsed Schluter price book (schluter_prices.json)."""
    json_path = os.path.join(os.path.dirname(__file__), "schluter_prices.json")
    if not os.path.exists(json_path):
        raise HTTPException(status_code=404, detail="schluter_prices.json not found. Run parse_schluter.py first.")
    import json as _json
    with open(json_path) as f:
        items = _json.load(f)
    count = import_price_book("Schluter", items, discount_pct=0.55, category="transitions")
    return {"imported": count, "vendor": "Schluter", "discount": "45% of list (55% off)"}


# ── Vendor Quote Test Mode / Simulation ──────────────────────────────────────

@app.get("/api/sim/status")
def api_sim_status():
    """Check if vendor quote test mode is active."""
    settings = get_settings()
    test_mode = str(settings.get("vendor_quote_test_mode", "false")).lower() == "true"
    watcher_active = _sim_watcher is not None and _sim_watcher.is_running if _sim_watcher else False
    return {
        "test_mode": test_mode,
        "watcher_active": watcher_active,
    }


@app.post("/api/jobs/{job_id}/send-quote-email")
async def api_send_quote_email(job_id: str, request: Request):
    """Send a vendor quote request email via SMTP.

    In test mode: routes to localhost:2525 (PowerShell relay → Vendor Simulator)
    In production: routes to real SMTP server → real vendor
    """
    db_id = _resolve_job_id(job_id)
    if not db_id:
        raise HTTPException(status_code=404, detail="Job not found")

    body = await request.json()
    vendor_name = body.get("vendor_name", "").strip()
    vendor_email = body.get("vendor_email", "").strip()
    subject = body.get("subject", "").strip()
    email_body = body.get("body", "").strip()
    material_ids = body.get("material_ids", [])
    vendor_id = body.get("vendor_id")

    if not vendor_name or not vendor_email:
        raise HTTPException(status_code=400, detail="vendor_name and vendor_email are required")
    if not subject:
        job = load_job(db_id)
        subject = f"Request for Pricing — {job.get('project_name', 'Project')}"
    job_tag = f"[SI Job {db_id}]"
    if job_tag.lower() not in subject.lower():
        subject = f"{subject} {job_tag}"

    # Get SMTP config based on test mode
    settings = get_settings()
    from sim_email import get_smtp_config, sim_send_quote
    smtp_config = get_smtp_config(settings)

    try:
        sim_send_quote(
            job_id=db_id,
            vendor_name=vendor_name,
            vendor_email=vendor_email,
            subject=subject,
            body=email_body,
            smtp_config=smtp_config,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {e}")

    # Create quote_request record (same as existing Mark Sent flow)
    qr = create_quote_request(
        job_id=db_id,
        vendor_name=vendor_name,
        material_ids=material_ids,
        request_text=email_body,
        vendor_id=vendor_id,
        status="sent",
        sent_at=body.get("sent_at") or __import__("datetime").datetime.utcnow().isoformat(),
    )

    test_mode = str(settings.get("vendor_quote_test_mode", "false")).lower() == "true"
    log_activity(db_id, "quote_email_sent",
                 f"{'[SIM] ' if test_mode else ''}Quote email sent to {vendor_name} ({vendor_email})",
                 {"vendor": vendor_name, "vendor_email": vendor_email, "test_mode": test_mode})

    return {"status": "sent", "quote_request": qr, "test_mode": test_mode}


# ── Static Files (React frontend) ────────────────────────────────────────────
# Check both dev path (../frontend/dist) and Docker path (./static)
frontend_dist = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
static_dir = os.path.join(os.path.dirname(__file__), "static")
_static_root = frontend_dist if os.path.isdir(frontend_dist) else static_dir if os.path.isdir(static_dir) else None

if _static_root:
    # Serve static assets (JS, CSS, images)
    app.mount("/assets", StaticFiles(directory=os.path.join(_static_root, "assets")), name="assets")

    # SPA catch-all: serve index.html for any non-API route
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        # Return 404 for undefined API routes instead of SPA HTML
        if full_path.startswith("api/"):
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=404, content={"detail": "Not found"})
        # If the file exists in static dir, serve it directly
        file_path = os.path.join(_static_root, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        # Otherwise serve index.html for client-side routing
        return FileResponse(os.path.join(_static_root, "index.html"), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
