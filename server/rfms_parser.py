"""
Parse RFMS Measure pivot table Excel files (.xlsx).
Reads the "By Item" sheet for material summaries and "Customer" sheet for job info.

The "By Item" sheet layout:
  Rows 1-6: Work order header (company name, address, date)
  Row 7:    Column headers — DESCRIPTION | QUANTITY | LINE TOTAL
  Row 8+:   Material/sundry/labor line items
  Last row:  "Grand Total"

Strategy:
  1. Read all rows from "By Item" sheet.
  2. Separate into: material lines, install lines, sundry lines.
  3. Send material + install lines to OpenAI GPT to classify each material
     by matching it to its corresponding install line.
  4. If AI is unavailable, materials stay "unknown" for manual review.
"""

import json
import os
import re
from typing import Optional

import openpyxl

# ── Valid material types for the bid tool ─────────────────────────────────────
VALID_MATERIAL_TYPES = [
    "unit_carpet_no_pattern",
    "unit_carpet_pattern",
    "unit_lvt",
    "cpt_tile",
    "corridor_broadloom",
    "floor_tile",
    "wall_tile",
    "backsplash",
    "tub_shower_surround",
    "rubber_base",
    "vct",
    "rubber_tile",
    "rubber_sheet",
    "wood",
    "tread_riser",
    "transitions",
    "waterproofing",
]

# ── Sundry / skip patterns ───────────────────────────────────────────────────
SUNDRY_PREFIXES = [
    "adhesive",
    "thin set",
    "thinset",
    "crack isolation",
    "carpet pad",
    "tack strip",
    "seam seal",
    "isolation strip",
    "weld rod",
    "primer",
    "grout ",
    "caulk",
]


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=3, values_only=False):
        a_val = _safe_str(row[0].value).lower()
        if "description" in a_val:
            return row[0].row
    return 7


def _is_sundry(desc: str) -> bool:
    desc_lower = desc.lower().strip()
    if not desc_lower or desc_lower in ("none", "none:"):
        return True
    # Strip scheme prefix before checking e.g. "(Scheme A & B) - Carpet Pad"
    clean = re.sub(r'^\(scheme\s+[a-z](?:\s*&\s*[a-z])?\)\s*-?\s*', '', desc_lower)
    for prefix in SUNDRY_PREFIXES:
        if desc_lower.startswith(prefix) or clean.startswith(prefix):
            return True
    return False


def _is_install(desc: str) -> bool:
    return desc.lower().strip().startswith("install")


def _extract_install_code(desc: str) -> str:
    """Extract item code from an install line description.

    Examples:
        "Install CPT-2 Broadloom"                          -> "CPT-2"
        "Install T-3 Floor Tile"                            -> "T-3"
        "Install F102 - Carpet Tile"                        -> "F102"
        "Install BR-1 Wall Tile"                            -> "BR-1"
        "Install LVP-1 Vinyl Plank"                         -> "LVP-1"
        "Install - (Scheme A) Metroflor - Performer..."     -> "(SCHEME A) METROFLOR"
        "Install (Standard) - CPT-200 - Broadloom"          -> "(STANDARD) CPT-200"
        "Install (Premium) -  T-203 - Wall Tile..."          -> "(PREMIUM) T-203"
    """
    # Strip "Install " prefix
    clean = re.sub(r'^install\s+', '', desc.strip(), flags=re.IGNORECASE)
    # Strip leading dashes/whitespace
    clean = re.sub(r'^[\s\-]+', '', clean)
    # Extract scheme prefix if present (e.g. "(Scheme A)") to include in key
    scheme_m = re.match(r'\(Scheme\s+([A-Z](?:\s*&\s*[A-Z])?)\)\s*', clean, re.IGNORECASE)
    scheme_prefix = ""
    if scheme_m:
        scheme_prefix = f"(SCHEME {scheme_m.group(1).upper()}) "
        clean_no_scheme = clean[scheme_m.end():]
    else:
        clean_no_scheme = clean
    # Extract option prefixes if present (e.g. "(Standard)", "(Premium)")
    # These work like scheme prefixes — may be stacked: "(Alternate) (Standard)"
    _OPTION_RE = re.compile(r'^\((Standard|Premium|Alternate|Budget|Base|Upgrade)\)\s*[-–—]?\s*', re.IGNORECASE)
    option_parts = []
    while True:
        option_m = _OPTION_RE.match(clean_no_scheme)
        if not option_m:
            break
        option_parts.append(option_m.group(1).upper())
        clean_no_scheme = clean_no_scheme[option_m.end():]
    option_prefix = " ".join(f"({p})" for p in option_parts) + " " if option_parts else ""
    prefix = scheme_prefix or option_prefix
    # Match item codes like CPT-2, T-3, F102, BR-1, LVP-1, W131, SCH-1, RB-1, T-200.1
    m = re.match(r'([A-Z]{1,4}-?\d{1,4}(?:\.\d+)?(?:/[A-Z]{1,4}-?\d{1,4})*)', clean_no_scheme, re.IGNORECASE)
    if m:
        code = m.group(1).upper()
        return (prefix + code) if prefix else code
    # Fallback: take the first part before " - " (e.g. "Metroflor")
    # Prefix with scheme/option to distinguish variants
    parts = clean_no_scheme.split(' - ')
    if parts and parts[0].strip():
        label = parts[0].strip()[:30]
        if re.search(r'[A-Za-z]', label):
            return (prefix + label).upper() if prefix else label.upper()
    return ""


def _extract_unit(desc: str, material_type: str) -> str:
    if material_type in ("rubber_base", "transitions"):
        return "LF"
    if material_type in ("corridor_broadloom", "unit_carpet_no_pattern", "unit_carpet_pattern", "cpt_tile"):
        return "SY"
    return "SF"


def _extract_item_label(desc: str) -> str:
    # Check for scheme prefix (e.g. "(Scheme A)")
    scheme_m = re.match(r'^\(Scheme\s+([A-Z](?:\s*&\s*[A-Z])?)\)\s*', desc, re.IGNORECASE)
    scheme_prefix = ""
    if scheme_m:
        scheme_prefix = f"(Scheme {scheme_m.group(1).upper()}) "
        clean = desc[scheme_m.end():]
    else:
        clean = desc
    # Check for option prefixes (e.g. "(Standard)", "(Premium)", "(Alternate)")
    # These work like scheme prefixes — they designate separate product choices
    # May be stacked: "(Alternate) (Standard) - T-200.1 -..."
    _OPTION_RE = re.compile(r'^\((Standard|Premium|Alternate|Budget|Base|Upgrade)\)\s*[-–—]?\s*', re.IGNORECASE)
    option_parts = []
    while True:
        option_m = _OPTION_RE.match(clean)
        if not option_m:
            break
        option_parts.append(option_m.group(1).title())
        clean = clean[option_m.end():]
    option_prefix = " ".join(f"({p})" for p in option_parts) + " " if option_parts else ""
    prefix = scheme_prefix or option_prefix
    # Match codes like F102, W131/W132/W133, B102, CPT-200, T-202, T-200.1
    m = re.match(r'^([A-Z]{1,4}-?\d{1,4}(?:\.\d+)?(?:/[A-Z]{1,4}-?\d{1,4})*)\b', clean, re.IGNORECASE)
    if m:
        code = m.group(1)
        return (prefix + code).strip() if prefix else code
    parts = clean.split(' - ')
    if parts:
        label = parts[0].strip()[:30]
        if prefix and label:
            return prefix + label
        return label
    return desc[:30]


# ── AI Classification ────────────────────────────────────────────────────────

AI_CLASSIFICATION_PROMPT = """You are a commercial flooring estimator assistant. I need you to classify each material line item from an RFMS takeoff pivot table.

You will receive two lists:
1. MATERIAL LINES — the actual materials that need to be classified
2. INSTALL LINES — these tell you what each material IS (e.g. "Install - F102 - Carpet Tile" means F102 is carpet tile)

Use the install lines to identify each material's type. The install line's description after the code tells you the material category. The @location after the dash tells you WHERE it goes, which sometimes changes the classification (e.g. "Wall Tile @Kitchen Backsplash" = backsplash, "Wall Tile @Shower Surrounds" = tub_shower_surround).

VALID MATERIAL TYPES (you must use exactly one of these):
- unit_carpet_no_pattern: Unit carpet without pattern (stretch-in over pad)
- unit_carpet_pattern: Unit carpet with pattern repeat
- unit_lvt: Luxury vinyl plank/tile (LVP/LVT)
- cpt_tile: Carpet tile (modular carpet squares)
- corridor_broadloom: Broadloom carpet for corridors/stairs (direct glue)
- floor_tile: Floor tile (porcelain, ceramic, mosaic floor)
- wall_tile: Wall tile (deco, mosaic, ceramic wall applications)
- backsplash: Kitchen backsplash tile
- tub_shower_surround: Tile for tub/shower surround walls
- rubber_base: Vinyl/rubber base (cove base)
- vct: Vinyl composition tile
- rubber_tile: Rubber floor tile
- rubber_sheet: Rubber sheet flooring
- wood: Hardwood flooring
- tread_riser: Stair tread and riser
- transitions: Transitions, edge trims, Schluter profiles
- waterproofing: Waterproofing membrane

CLASSIFICATION RULES:
- Match material codes (F102, W131, B102, etc.) to their install line counterparts
- "Wall Tile @Kitchen Backsplash" → backsplash (location overrides)
- "Wall Tile @Shower Surrounds" or "@Tub Surrounds" → tub_shower_surround (location overrides)
- "Deco Wall Tile @Dog Wash Tub Surrounds" → wall_tile (it says "Deco Wall Tile", the tub surround is the location name, not the material application)
- Transitions, Schluter profiles, edge trims → transitions
- Vinyl Base, Rubber Base → rubber_base
- Carpet with "Broadloom" and in common areas/stairwells → corridor_broadloom
- Carpet with "Broadloom" in units/bedrooms → unit_carpet_no_pattern (or unit_carpet_pattern if pattern repeat is mentioned)
- Lines about carpet pad, isolation strips, adhesive → these are sundries, classify as "sundry"

Return a JSON object with a "classifications" key containing an array. Each element must have:
  {"index": <number>, "material_type": "<type>", "confidence": <0.0-1.0>}
The confidence score indicates how certain you are about the classification (1.0 = very certain, 0.5 = unsure).

You MUST classify EVERY material line — one entry per material. Return ALL of them.
Example response format: {"classifications": [{"index": 0, "material_type": "floor_tile", "confidence": 0.95}, {"index": 1, "material_type": "wall_tile", "confidence": 0.8}]}"""


def _classify_with_ai(material_lines: list[tuple[int, str]],
                       install_lines: list[str]) -> dict[int, dict]:
    """
    Use OpenAI to classify material lines by matching with install lines.
    Returns: {index: material_type}
    """
    try:
        from ai_client import chat_complete, get_provider_info
        from quote_parser import _openai_config

        api_key = _openai_config.get("api_key") or os.environ.get("OPENAI_API_KEY")
        provider = get_provider_info(api_key)
        if not provider["available"]:
            print("[rfms_parser] No AI API key available for classification")
            return {}

        model = _openai_config.get("model", "gpt-5-mini")
        print(f"[rfms_parser] Classifying {len(material_lines)} materials with provider={provider['provider']}, model={model}")

        # Build the user message
        mat_text = "\n".join(
            f"  [{idx}] {desc}" for idx, desc in material_lines
        )
        inst_text = "\n".join(f"  - {desc}" for desc in install_lines)

        user_msg = f"""MATERIAL LINES (classify each one):
{mat_text}

INSTALL LINES (use these to identify material types):
{inst_text}"""

        raw = chat_complete(
            system=AI_CLASSIFICATION_PROMPT,
            user=user_msg,
            api_key=api_key,
            model=model,
            json_mode=True,
        )
        print(f"[rfms_parser] AI classification raw response: {raw[:500]}")
        parsed = json.loads(raw)

        # Handle various response formats
        if isinstance(parsed, dict):
            # Check for known wrapper keys
            items = parsed.get("classifications", parsed.get("results", []))
            if not items:
                # Try any list value in the dict
                for v in parsed.values():
                    if isinstance(v, list):
                        items = v
                        break
            if not items and "index" in parsed and "material_type" in parsed:
                # Single item response — wrap it
                items = [parsed]
        elif isinstance(parsed, list):
            items = parsed
        else:
            return {}

        result = {}
        for item in items:
            idx = item.get("index")
            mtype = item.get("material_type", "unknown")
            confidence = item.get("confidence", 0.5)
            if idx is not None and mtype in VALID_MATERIAL_TYPES:
                result[idx] = {"type": mtype, "confidence": confidence}
            elif idx is not None and mtype == "sundry":
                result[idx] = {"type": "sundry", "confidence": confidence}

        print(f"[rfms_parser] Classification result: {len(result)} classified out of {len(material_lines)}")
        return result

    except Exception as e:
        print(f"[rfms_parser] AI classification failed, materials will be 'unknown': {e}")
        return {}


# ── Main parser ──────────────────────────────────────────────────────────────

def parse_rfms(file_path: str) -> dict:
    """
    Parse an RFMS pivot table Excel file.

    Returns:
        {
            "job_info": {project_name, address, city, state, zip, gc_name},
            "materials": [
                {item_code, qty, unit, description, material_type},
                ...
            ]
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # ── Parse Customer sheet ─────────────────────────────────────────────────
    job_info = {}
    customer_sheet = None
    for name in wb.sheetnames:
        if "customer" in name.lower():
            customer_sheet = wb[name]
            break

    if customer_sheet:
        for row in customer_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            key = _safe_str(row[0]).lower() if row[0] else ""
            val = _safe_str(row[1]) if len(row) > 1 and row[1] else ""
            if "project" in key or "jobname" in key or key == "name":
                job_info["project_name"] = val
            elif "jobaddress" in key or ("address" in key and "project" not in key):
                job_info["address"] = val
            elif "city" in key:
                job_info["city"] = val
            elif "state" in key:
                job_info["state"] = val
            elif "zip" in key:
                job_info["zip"] = val
            elif "contractor" in key or "gc" in key or "builder" in key:
                job_info["gc_name"] = val

    # ── Find By Item sheet ───────────────────────────────────────────────────
    by_item_sheet = None
    for name in wb.sheetnames:
        if "by item" in name.lower():
            by_item_sheet = wb[name]
            break
    if by_item_sheet is None and len(wb.sheetnames) >= 3:
        by_item_sheet = wb[wb.sheetnames[2]]

    if not by_item_sheet:
        wb.close()
        return {"job_info": job_info, "materials": []}

    header_row = _find_header_row(by_item_sheet)

    # ── Read all rows ────────────────────────────────────────────────────────
    all_rows = []
    for row in by_item_sheet.iter_rows(
        min_row=header_row + 1, max_col=3, values_only=True
    ):
        desc_raw = _safe_str(row[0]).rstrip(":")
        qty = _safe_float(row[1]) if len(row) > 1 else 0.0
        if not desc_raw or desc_raw.lower() in ("none", "") or qty == 0:
            continue
        if desc_raw.lower().strip() in ("grand total", "total"):
            continue
        all_rows.append((desc_raw, qty))

    # ── Separate into categories ─────────────────────────────────────────────
    material_lines = []   # (index, description, qty)
    install_lines = []    # descriptions only
    sundry_qtys = {}      # measured sundry quantities from RFMS
    idx = 0

    install_qtys = {}     # item_code -> install qty (net installed area)

    for desc, qty in all_rows:
        if _is_install(desc):
            install_lines.append(desc)
            # Extract item code and qty from install lines
            # e.g. "Install CPT-2 Broadloom" -> code "CPT-2", qty 34.6
            install_code = _extract_install_code(desc)
            if install_code:
                install_qtys[install_code] = install_qtys.get(install_code, 0) + qty
        elif _is_sundry(desc):
            # Capture measured sundry quantities for carpet materials
            desc_lower = desc.lower().strip()
            if desc_lower.startswith("tack strip"):
                sundry_qtys["tack_strip_lf"] = sundry_qtys.get("tack_strip_lf", 0) + qty
            elif desc_lower.startswith("seam seal"):
                sundry_qtys["seam_tape_lf"] = sundry_qtys.get("seam_tape_lf", 0) + qty
            elif desc_lower.startswith("carpet pad"):
                sundry_qtys["pad_sy"] = sundry_qtys.get("pad_sy", 0) + qty
            elif "crack isolation" in desc_lower:
                sundry_qtys["crack_isolation_sf"] = sundry_qtys.get("crack_isolation_sf", 0) + qty
            elif desc_lower.startswith("weld rod"):
                sundry_qtys["weld_rod_lf"] = sundry_qtys.get("weld_rod_lf", 0) + qty
        else:
            material_lines.append((idx, desc, qty))
            idx += 1

    # ── Classify with AI ─────────────────────────────────────────────────────
    ai_input = [(i, desc) for i, desc, qty in material_lines]
    ai_results = _classify_with_ai(ai_input, install_lines)

    # ── Build materials list ─────────────────────────────────────────────────
    materials = []
    for i, desc, qty in material_lines:
        # Use AI result — no fallback, stays "unknown" if AI didn't classify
        ai_result = ai_results.get(i, {"type": "unknown", "confidence": None})
        if isinstance(ai_result, str):
            # Backwards compat: old format returned just a string
            material_type = ai_result
            ai_confidence = None
        else:
            material_type = ai_result.get("type", "unknown")
            ai_confidence = ai_result.get("confidence")

        # Skip if AI classified as sundry
        if material_type == "sundry":
            continue

        unit = _extract_unit(desc, material_type)
        item_code = _extract_item_label(desc)

        # Use install line qty (net installed area) when available.
        # The material line qty from RFMS includes RFMS-calculated waste,
        # but we apply our own waste factors — so we need the net qty.
        install_qty = install_qtys.get(item_code.upper())
        if install_qty and install_qty > 0:
            base_qty = install_qty
        else:
            base_qty = qty

        materials.append({
            "item_code": item_code,
            "qty": base_qty,
            "unit": unit,
            "description": desc,
            "material_type": material_type,
            "ai_confidence": ai_confidence,
        })

    # Detect mosaic tiles from description and tile dimensions
    _MOSAIC_KEYWORDS = {"mosaic"}
    _PENNY_HEX_KEYWORDS = {"penny round", "hex", "hexagon"}
    _TILE_TYPES = {"floor_tile", "wall_tile", "backsplash", "tub_shower_surround"}

    for m in materials:
        if m.get("material_type") not in _TILE_TYPES:
            continue
        desc_lower = m.get("description", "").lower()
        # Check keywords
        is_mosaic = any(kw in desc_lower for kw in _MOSAIC_KEYWORDS | _PENNY_HEX_KEYWORDS)
        # Check small tile dimensions (both dims ≤ 6") — pattern like 1" x 6", 2x4, etc.
        if not is_mosaic:
            dim_match = re.findall(r'(\d+(?:\.\d+)?)\s*["\u201d]?\s*x\s*(\d+(?:\.\d+)?)\s*["\u201d]?', desc_lower)
            if dim_match:
                w, h = float(dim_match[0][0]), float(dim_match[0][1])
                if w <= 6 and h <= 6:
                    is_mosaic = True
        if is_mosaic:
            m["is_mosaic"] = True
            m["is_penny_hex"] = any(kw in desc_lower for kw in _PENNY_HEX_KEYWORDS)

    # Attach RFMS-measured sundry quantities to stretch-in carpet materials
    if sundry_qtys:
        carpet_types = ("unit_carpet_no_pattern", "unit_carpet_pattern")
        carpet_mats = [m for m in materials if m.get("material_type") in carpet_types]
        if carpet_mats:
            # Distribute sundry quantities proportionally by carpet SY
            total_carpet_sy = sum(m["qty"] for m in carpet_mats)
            for m in carpet_mats:
                share = m["qty"] / total_carpet_sy if total_carpet_sy > 0 else 1.0
                m["tack_strip_lf"] = round(sundry_qtys.get("tack_strip_lf", 0) * share, 2)
                m["seam_tape_lf"] = round(sundry_qtys.get("seam_tape_lf", 0) * share, 2)
                m["pad_sy"] = round(sundry_qtys.get("pad_sy", 0) * share, 2)

        # Attach weld rod LF to rubber_sheet materials
        weld_rod_lf = sundry_qtys.get("weld_rod_lf", 0)
        if weld_rod_lf > 0:
            sheet_mats = [m for m in materials if m.get("material_type") == "rubber_sheet"]
            if sheet_mats:
                total_sheet_sf = sum(m["qty"] for m in sheet_mats)
                for m in sheet_mats:
                    share = m["qty"] / total_sheet_sf if total_sheet_sf > 0 else 1.0
                    m["weld_rod_lf"] = round(weld_rod_lf * share, 2)

        # Attach crack isolation SF to floor_tile materials
        crack_sf = sundry_qtys.get("crack_isolation_sf", 0)
        if crack_sf > 0:
            floor_mats = [m for m in materials if m.get("material_type") == "floor_tile"]
            if floor_mats:
                total_floor_sf = sum(m["qty"] for m in floor_mats)
                for m in floor_mats:
                    share = m["qty"] / total_floor_sf if total_floor_sf > 0 else 1.0
                    m["crack_isolation_sf"] = round(crack_sf * share, 2)

    wb.close()
    return {"job_info": job_info, "materials": materials}


# ── AI Merge Prompt ───────────────────────────────────────────────────────────

AI_MERGE_PROMPT = """You are a commercial flooring estimator. You are given two lists of materials for a job:

1. EXISTING MATERIALS — already saved on the job (may have pricing set by the estimator)
2. NEW MATERIALS — just parsed from a newly uploaded RFMS takeoff file

Your task: Return the CORRECT FINAL materials list as a JSON array.

Use your expertise to:
- Identify duplicates (same material uploaded again) and keep only one copy
- Identify materials that exist in both lists and should be combined (sum their quantities)
- Keep materials that are genuinely different (different areas, different products)
- Preserve any pricing data (vendor, unit_price) from existing materials — the estimator already set these
- Review and correct all material_type values — fix any "unknown" types using context from the full list

VALID MATERIAL TYPES:
unit_carpet_no_pattern, unit_carpet_pattern, unit_lvt, cpt_tile, corridor_broadloom,
floor_tile, wall_tile, backsplash, tub_shower_surround, rubber_base, vct,
rubber_tile, rubber_sheet, wood, tread_riser, transitions, waterproofing

Return ONLY a JSON object: {"materials": [...]}
Each material must have: item_code, description, material_type, installed_qty, unit
If preserving pricing from existing: also include vendor, unit_price"""


AI_VERIFY_PROMPT = """You are a commercial flooring estimator reviewing a material merge.

ORIGINAL EXISTING MATERIALS (count: {existing_count}, total qty: {existing_total}):
{existing_summary}

ORIGINAL NEW MATERIALS (count: {new_count}, total qty: {new_total}):
{new_summary}

PROPOSED MERGED RESULT (count: {merged_count}, total qty: {merged_total}):
{merged_json}

Review this merge carefully:
- Are any materials missing that should be in the final list?
- Are any materials incorrectly duplicated?
- Are any quantities wrong (e.g. should have been summed or shouldn't have been)?
- Are all material_type values correct?

If the merge is correct, return: {{"correct": true}}
If there are issues, return: {{"correct": false, "materials": [<the corrected full list>]}}
Each material must have: item_code, description, material_type, installed_qty, unit, and optionally vendor, unit_price"""


def ai_merge_materials(existing: list[dict], new_parsed: list[dict]) -> list[dict]:
    """
    Use AI to intelligently merge existing materials with newly parsed ones.
    Returns the correct final materials list.
    Falls back to appending new to existing if AI is unavailable.
    """
    try:
        from ai_client import chat_complete, get_provider_info
        from quote_parser import _openai_config

        api_key = _openai_config.get("api_key") or os.environ.get("OPENAI_API_KEY")
        provider = get_provider_info(api_key)
        if not provider["available"]:
            print("[ai_merge] No AI provider available, falling back to append")
            return _fallback_merge(existing, new_parsed)

        model = _openai_config.get("model", "gpt-5-mini")

        # Prepare existing materials for AI (strip DB-only fields)
        existing_for_ai = []
        for m in existing:
            existing_for_ai.append({
                "item_code": m.get("item_code"),
                "description": m.get("description"),
                "material_type": m.get("material_type"),
                "installed_qty": m.get("installed_qty", 0),
                "unit": m.get("unit"),
                "vendor": m.get("vendor", ""),
                "unit_price": m.get("unit_price", 0),
                "area_type": m.get("area_type", "unit"),
            })

        # Prepare new materials (rename qty → installed_qty for consistency)
        new_for_ai = []
        for m in new_parsed:
            new_for_ai.append({
                "item_code": m.get("item_code"),
                "description": m.get("description"),
                "material_type": m.get("material_type"),
                "installed_qty": m.get("qty", m.get("installed_qty", 0)),
                "unit": m.get("unit"),
                "area_type": m.get("area_type", "unit"),
            })

        user_msg = f"""EXISTING MATERIALS ({len(existing_for_ai)} items):
{json.dumps(existing_for_ai, indent=2)}

NEW MATERIALS ({len(new_for_ai)} items):
{json.dumps(new_for_ai, indent=2)}"""

        # Pass 1: Merge
        print(f"[ai_merge] Pass 1: merging {len(existing_for_ai)} existing + {len(new_for_ai)} new materials via {provider['provider']}")
        raw = chat_complete(
            system=AI_MERGE_PROMPT,
            user=user_msg,
            api_key=api_key,
            model=model,
            json_mode=True,
        )

        parsed = json.loads(raw)
        merged = parsed.get("materials", [])
        print(f"[ai_merge] Pass 1 result: {len(merged)} materials")

        if not merged:
            print("[ai_merge] Empty merge result, falling back")
            return _fallback_merge(existing, new_parsed)

        # Pass 2: Verify
        existing_total = sum(m.get("installed_qty", 0) for m in existing_for_ai)
        new_total = sum(m.get("installed_qty", 0) for m in new_for_ai)
        merged_total = sum(m.get("installed_qty", 0) for m in merged)

        verify_msg = AI_VERIFY_PROMPT.format(
            existing_count=len(existing_for_ai),
            existing_total=round(existing_total, 2),
            existing_summary=json.dumps(existing_for_ai, indent=2),
            new_count=len(new_for_ai),
            new_total=round(new_total, 2),
            new_summary=json.dumps(new_for_ai, indent=2),
            merged_count=len(merged),
            merged_total=round(merged_total, 2),
            merged_json=json.dumps(merged, indent=2),
        )

        print(f"[ai_merge] Pass 2: verifying merge")
        verify_raw = chat_complete(
            system="You are a commercial flooring estimator verifying a material list merge. Return your response as JSON.",
            user=verify_msg,
            api_key=api_key,
            model=model,
            json_mode=True,
        )

        verify_parsed = json.loads(verify_raw)

        if verify_parsed.get("correct"):
            print("[ai_merge] Pass 2: merge verified correct")
            return merged
        elif verify_parsed.get("materials"):
            corrected = verify_parsed["materials"]
            print(f"[ai_merge] Pass 2: corrected to {len(corrected)} materials")
            return corrected
        else:
            print("[ai_merge] Pass 2: unclear response, using pass 1 result")
            return merged

    except Exception as e:
        print(f"[ai_merge] AI merge failed: {e}")
        return _fallback_merge(existing, new_parsed)


def _fallback_merge(existing: list[dict], new_parsed: list[dict]) -> list[dict]:
    """Simple fallback: append new materials to existing."""
    result = []
    for m in existing:
        result.append({
            "item_code": m.get("item_code"),
            "description": m.get("description"),
            "material_type": m.get("material_type"),
            "installed_qty": m.get("installed_qty", 0),
            "unit": m.get("unit"),
            "vendor": m.get("vendor", ""),
            "unit_price": m.get("unit_price", 0),
            "area_type": m.get("area_type", "unit"),
        })
    for m in new_parsed:
        result.append({
            "item_code": m.get("item_code"),
            "description": m.get("description"),
            "material_type": m.get("material_type"),
            "installed_qty": m.get("qty", m.get("installed_qty", 0)),
            "unit": m.get("unit"),
            "area_type": m.get("area_type", "unit"),
        })
    return result
