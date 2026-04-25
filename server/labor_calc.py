"""
Labor calculator: match materials to labor catalog entries,
determine quantities based on LABOR_QTY_RULES, and compute costs.
"""

import json
import os
import re
from typing import Optional

import openpyxl
import pdfplumber

from ai_client import chat_complete, get_provider_info
from config import LABOR_QTY_RULES, WASTE_FACTORS
from models import save_labor_catalog_entries, get_labor_catalog_entries

LABOR_CATALOG_PROMPT = """You are parsing a labor rate catalog for a flooring/interiors contractor.
Extract every labor line item from this document into a JSON array.

For each entry, extract:
- labor_type: the category/type (e.g. "Carpet Stretch-In", "LVT Install", "Floor Tile", etc.)
- description: any additional description or notes
- cost: the cost/rate as a number (the contractor's cost per unit)
- retail_display: the retail/display price if shown, otherwise empty string
- unit: the unit of measure (e.g. "SY", "SF", "LF", "EA", "HR")
- gpm_markup: the GPM/markup percentage as a decimal if shown (e.g. 0.35 for 35%), otherwise 0

Return JSON in this exact format:
{"entries": [{"labor_type": "...", "description": "...", "cost": 0.00, "retail_display": "...", "unit": "...", "gpm_markup": 0.00}, ...]}

Important:
- Extract ALL labor entries, not just a sample
- Cost must be a number, not a string
- If a field is missing, use empty string for text or 0 for numbers
"""

# ── Labor matching rules ─────────────────────────────────────────────────────
# Each material type maps to a base labor entry search + optional add-ons.
# "base" is a list of description substrings that ALL must match (AND logic).
# "size_tiers" defines qty thresholds to pick the right size tier.
# "addons" is a list of description substrings for X ADD entries to include.
LABOR_RULES: dict[str, dict] = {
    "unit_carpet_no_pattern": {
        "labor_type": "Project Carpet",
        "base": ["broadloom", "stretch", "over pad"],
        "size_tiers": [
            (2000, "more than 2000"),
            (0, "2000sy or less"),
        ],
        "size_field": "installed_qty",  # SY
        "addons": [],
    },
    "unit_carpet_pattern": {
        "labor_type": "Project Carpet",
        "base": ["broadloom", "stretch", "over pad"],
        "size_tiers": [
            (2000, "more than 2000"),
            (0, "2000sy or less"),
        ],
        "size_field": "installed_qty",
        "addons": ["x add for pattern"],
    },
    "cpt_tile": {
        "labor_type": "Project Carpet",
        "base": ["carpet tile"],
        "size_tiers": [],
        "addons": [],
    },
    "corridor_broadloom": {
        "labor_type": "Project Carpet",
        "base": ["broadloom", "direct glue"],
        "size_tiers": [],
        "addons": [],
    },
    "unit_lvt": {
        "labor_type": "Project Resilient Tile",
        "base": ["plank or lvt", "glue down"],
        "size_tiers": [
            (1000, "more than 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": [],
    },
    "floor_tile": {
        "labor_type": "Project Tile",
        "base": ["porcelain set monolithic"],
        "size_tiers": [
            (2500, "more than 2500"),
            (1000, "between 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": [],
    },
    "wall_tile": {
        "labor_type": "Project Tile",
        "base": ["porcelain set monolithic"],
        "size_tiers": [
            (2500, "more than 2500"),
            (1000, "between 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": [],
    },
    "backsplash": {
        "labor_type": "Project Tile",
        "base": ["porcelain set monolithic"],
        "size_tiers": [
            (2500, "more than 2500"),
            (1000, "between 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": ["x add for backsplash"],
    },
    "tub_shower_surround": {
        "labor_type": "Project Tile",
        "base": ["porcelain set monolithic"],
        "size_tiers": [
            (2500, "more than 2500"),
            (1000, "between 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": [],
    },
    "rubber_base": {
        "labor_type": "Project Wall Base",
        "base": ["4 inch cove base"],
        "size_tiers": [],
        "addons": [],
    },
    "rubber_tile": {
        "labor_type": "Project Resilient Tile",
        "base": ["rubber", "tile"],
        "size_tiers": [],
        "addons": [],
    },
    "sound_mat": {
        "labor_type": "Project Resilient Rolled Sheet",
        "base": ["install sound mat"],
        "size_tiers": [],
        "size_field": "installed_qty",
        "addons": [],
    },
    "rubber_sheet": {
        "labor_type": "Project Resilient Rolled Sheet",
        "base": ["commercial sheet vinyl", "over 1/4"],
        "size_tiers": [
            (1000, "more than 1000"),
            (0, "1000sy or less"),
        ],
        "size_field": "installed_qty_sy",
        "addons": ["weld commercial vinyl"],
    },
    "vct": {
        "labor_type": "Project Resilient Tile",
        "base": ["vct", "glue down"],
        "size_tiers": [
            (1000, "more than 1000"),
            (0, "1000sf or less"),
        ],
        "size_field": "installed_qty",
        "addons": [],
    },
    "wood": {
        "labor_type": "Project Hardwood",
        "base": ["prefinished hardwood", "glue down"],
        "size_tiers": [],
        "addons": [],
    },
    "tread_riser": {
        "labor_type": "Project Resilient Tile",
        "base": ["stair", "tread and riser"],
        "size_tiers": [],
        "addons": [],
    },
    "waterproofing": {
        "labor_type": "Project Tile Add Ons",
        "base": ["waterproofing roll on"],
        "size_tiers": [],
        "addons": [],
    },
    "transitions": {
        "labor_type": "Project Tile Add Ons",
        "base": ["schluter schiene"],
        "size_tiers": [],
        "addons": [],
    },
}


def load_labor_catalog(file_path: str) -> list[dict]:
    """
    Parse the Labor Catalog Excel file and persist to database.
    Sheet1, columns A-F: Labor Type, Description, Cost, Retail Display, Unit, GPM Markup
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # Sheet1

    catalog = []
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        labor_type = str(row[0]).strip() if row[0] else ""
        description = str(row[1]).strip() if row[1] else ""
        cost = _safe_float(row[2])
        retail_display = str(row[3]).strip() if row[3] else ""
        unit = str(row[4]).strip() if row[4] else ""
        gpm_markup = _safe_float(row[5])

        if not labor_type:
            continue

        catalog.append({
            "labor_type": labor_type,
            "description": description,
            "cost": cost,
            "retail_display": retail_display,
            "unit": unit,
            "gpm_markup": gpm_markup,
        })

    wb.close()
    save_labor_catalog_entries(catalog)
    return catalog


def load_labor_catalog_from_pdf(file_path: str, api_key: str = None, model: str = "gpt-5-mini") -> list[dict]:
    """
    Parse a Labor Catalog PDF using pdfplumber + OpenAI and persist to database.
    """
    # Extract text from PDF
    text_parts = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    text = "\n".join(text_parts)

    if not text.strip():
        raise ValueError("Could not extract text from labor catalog PDF")

    # Call AI to parse the labor entries
    content = chat_complete(
        system=LABOR_CATALOG_PROMPT,
        user=text,
        api_key=api_key,
        model=model,
        json_mode=True,
    )
    parsed = json.loads(content)
    entries = parsed.get("entries", [])

    catalog = []
    for entry in entries:
        labor_type = str(entry.get("labor_type", "")).strip()
        if not labor_type:
            continue
        catalog.append({
            "labor_type": labor_type,
            "description": str(entry.get("description", "")).strip(),
            "cost": _safe_float(entry.get("cost")),
            "retail_display": str(entry.get("retail_display", "")).strip(),
            "unit": str(entry.get("unit", "")).strip(),
            "gpm_markup": _safe_float(entry.get("gpm_markup")),
        })

    save_labor_catalog_entries(catalog)
    return catalog


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


_TILE_TYPES = {"floor_tile", "wall_tile", "backsplash", "tub_shower_surround"}

_TILE_DIM_RE = re.compile(r'(\d+(?:\.\d+)?)\s*["\u201d]?\s*x\s*(\d+(?:\.\d+)?)\s*["\u201d]?', re.IGNORECASE)


def _parse_tile_dims(description: str):
    """Parse tile dimensions from a material description like '12" x 24"'.
    Returns (min_edge, max_edge) or None."""
    m = _TILE_DIM_RE.search(description)
    if not m:
        return None
    a, b = float(m.group(1)), float(m.group(2))
    return (min(a, b), max(a, b))


def _tile_dim_tier(w: float, h: float) -> str:
    """Map tile dimensions to the catalog dimension keyword for filtering.
    Order matters: most specific first so 48x48 doesn't get caught by "greater than 36in".
    """
    max_edge = max(w, h)
    min_edge = min(w, h)
    if min_edge >= 48 and max_edge >= 48:
        return "48x48"
    if max_edge >= 48 and min_edge >= 24:
        return "24x48"
    if max_edge > 36:
        return "greater than 36in"
    elif max_edge > 24:
        return "greater than 24in"
    elif min_edge >= 24 and max_edge >= 24:
        return "24x24"
    elif max_edge <= 13:
        return "0-13x0-13"
    else:
        return "12x24"


_THICKNESS_MM_RE = re.compile(r'(\d+(?:\.\d+)?)\s*mm\b', re.IGNORECASE)


def _parse_mat_thickness_mm(description: str):
    """Parse mm thickness from material description, e.g. '8.2mm Thick', '5mm Acoustical'.
    Returns a float or None if no mm value found."""
    m = _THICKNESS_MM_RE.search(description or "")
    if not m:
        return None
    try:
        return float(m.group(1))
    except (TypeError, ValueError):
        return None


def _sound_mat_tier_kw(thickness_mm: float) -> str:
    """Map sound mat thickness to the catalog keyword used to pick the right labor row."""
    if thickness_mm < 3:
        return "less than 3mm"
    if thickness_mm <= 6:
        return "4mm to 6mm"
    return "more than 7mm"


def _find_labor_entries(
    material_type: str,
    installed_qty: float,
    catalog: list[dict] = None,
    rule_override: dict = None,
    material: dict = None,
) -> list[dict]:
    """Find all matching labor catalog entries for a material type.

    Returns a list of catalog entries: the base rate entry + any X ADD entries.
    Uses LABOR_RULES for precise matching with size tier selection.
    rule_override: optional dict to override specific rule fields (e.g. base keywords for mosaic).
    material: optional full material dict for tile dimension matching.
    """
    rule = LABOR_RULES.get(material_type)
    if not rule:
        return []
    if rule_override:
        rule = {**rule, **rule_override}

    if catalog is None:
        catalog = get_labor_catalog_entries()

    # Filter catalog to matching labor_type
    type_prefix = rule["labor_type"].lower()
    type_entries = [e for e in catalog if e["labor_type"].lower().startswith(type_prefix)]

    results = []

    # Find base entry: ALL substrings in "base" must match the description
    base_keywords = [kw.lower() for kw in rule["base"]]
    candidates = []
    for entry in type_entries:
        desc = entry["description"].lower()
        if all(kw in desc for kw in base_keywords):
            candidates.append(entry)

    # For tile types, filter candidates by tile dimensions from material description
    if material_type in _TILE_TYPES and material and candidates:
        dims = _parse_tile_dims(material.get("description", ""))
        if dims:
            tier_kw = _tile_dim_tier(dims[0], dims[1]).lower()
            dim_filtered = [c for c in candidates if tier_kw in c["description"].lower()]
            if dim_filtered:
                candidates = dim_filtered

    # For sound mat, pick the catalog row that matches the material's mm thickness.
    # Special case: descriptions containing "Premium" (with no explicit mm) bump to
    # the 4-6mm tier — Josh bills Premium at $0.75 even when the product is the same
    # as Standard (see Sun Valley Block 2: both are Pliteq RST05, but Standard bills
    # at $0.50 and Premium at $0.75).
    if material_type == "sound_mat" and material and candidates:
        desc_lower = (material.get("description") or "").lower()
        thickness = _parse_mat_thickness_mm(desc_lower)
        if thickness is not None:
            tier_kw = _sound_mat_tier_kw(thickness)
        elif "premium" in desc_lower:
            tier_kw = "4mm to 6mm"
        else:
            # No thickness in description and not Premium — default to thinnest tier
            # so we don't accidentally pick an expensive rate by alphabetical luck.
            tier_kw = "less than 3mm"
        thickness_filtered = [c for c in candidates if tier_kw in c["description"].lower()]
        if thickness_filtered:
            candidates = thickness_filtered

    # For rubber_sheet materials, route to "Rolled Rubber" rows when the product
    # isn't explicitly commercial sheet vinyl. Many rubber products (Ecofit, Mondo,
    # Sportec, etc.) don't contain the word "rubber" in the description, so we
    # exclude on "vinyl" rather than include on "rubber".
    if material_type == "rubber_sheet" and material and candidates:
        desc_lower = (material.get("description") or "").lower()
        if "vinyl" not in desc_lower:
            rolled = [e for e in type_entries if "rolled rubber" in e["description"].lower()]
            if rolled:
                thickness = _parse_mat_thickness_mm(desc_lower)
                if thickness is not None:
                    tier_kw = "3mm or under" if thickness <= 3 else "over 3mm"
                    filtered = [c for c in rolled if tier_kw in c["description"].lower()]
                    if filtered:
                        rolled = filtered
                candidates = rolled

    # Apply size tier selection if defined
    size_tiers = rule.get("size_tiers", [])
    if size_tiers and candidates:
        # Convert qty to SY if the rule's thresholds are in SY
        size_qty = installed_qty
        if rule.get("size_field") == "installed_qty_sy":
            size_qty = installed_qty / 9

        best = None
        for threshold, tier_kw in size_tiers:
            if size_qty > threshold:
                # Find candidate matching this tier keyword
                for c in candidates:
                    if tier_kw.lower() in c["description"].lower():
                        best = c
                        break
                if best:
                    break
        if best:
            results.append(best)
        elif candidates:
            results.append(candidates[0])
    elif candidates:
        results.append(candidates[0])

    # Find addon entries (X ADD lines)
    for addon_kw in rule.get("addons", []):
        addon_kw_lower = addon_kw.lower()
        for entry in type_entries:
            if addon_kw_lower in entry["description"].lower():
                results.append(entry)
                break

    return results


def _get_labor_qty_rule(material_type: str) -> str:
    """Get the labor quantity rule for a material type.
    Matches longest key first so 'corridor_broadloom' beats 'broadloom'.
    """
    for key in sorted(LABOR_QTY_RULES.keys(), key=len, reverse=True):
        if key == "default":
            continue
        if key in material_type:
            return LABOR_QTY_RULES[key]
    return LABOR_QTY_RULES["default"]


def calculate_labor(
    material_type: str,
    installed_qty: float,
    waste_pct: Optional[float] = None,
    measure_qty: Optional[float] = None,
    material: dict = None,
) -> list[dict]:
    """
    Calculate labor for a single material line item.
    Returns a list of labor line items (base + add-ons), or empty list.
    material: optional full material dict for mosaic detection.
    """
    # Build rule override for mosaic tiles
    rule_override = None
    if material and material.get("is_mosaic"):
        if material.get("is_penny_hex"):
            rule_override = {"base": ["penny round or hex mosaic"]}
        else:
            rule_override = {"base": ["mosaic sheet backed porcelain"]}

    entries = _find_labor_entries(material_type, installed_qty, rule_override=rule_override, material=material)
    if not entries:
        return []

    if waste_pct is None:
        waste_pct = WASTE_FACTORS.get(material_type, 0)

    rule = _get_labor_qty_rule(material_type)
    if rule == "with_waste":
        labor_qty = installed_qty * (1 + waste_pct)
    elif rule == "no_waste":
        labor_qty = installed_qty
    elif rule == "from_measure" and measure_qty is not None:
        labor_qty = measure_qty
    else:
        labor_qty = installed_qty * (1 + waste_pct)

    results = []
    for entry in entries:
        rate = entry["cost"]
        extended_cost = labor_qty * rate
        results.append({
            "labor_description": f"{entry['labor_type']} - {entry['description']}",
            "qty": round(labor_qty, 2),
            "unit": entry["unit"],
            "rate": rate,
            "extended_cost": round(extended_cost, 2),
        })
    return results


def calculate_labor_for_materials(materials: list[dict]) -> list[dict]:
    """
    Calculate labor for a list of material line items.
    Loads labor catalog from DB once for efficiency.
    Each material can produce multiple labor line items (base + add-ons).
    """
    catalog = get_labor_catalog_entries()
    results = []
    for mat in materials:
        material_type = mat.get("material_type", "")
        installed_qty = _safe_float(mat.get("installed_qty", 0))
        waste_pct = mat.get("waste_pct")
        material_id = mat.get("id") or mat.get("item_code")

        # Skip labor for non-Schluter transitions (pin metal, etc.)
        # Schluter products: named "Schluter", Jolly, Exposed Edge Trim, metal trim to tile
        if material_type == "transitions":
            desc = mat.get("description", "").lower()
            is_schluter = ("schluter" in desc or "jolly" in desc
                           or "exposed edge trim" in desc
                           or "metal trim" in desc)
            if not is_schluter:
                continue

        # Build rule override for mosaic tiles
        rule_override = None
        if mat.get("is_mosaic"):
            if mat.get("is_penny_hex"):
                rule_override = {"base": ["penny round or hex mosaic"]}
            else:
                rule_override = {"base": ["mosaic sheet backed porcelain"]}

        entries = _find_labor_entries(material_type, installed_qty, catalog, rule_override=rule_override, material=mat)
        if not entries:
            continue

        if waste_pct is None:
            waste_pct = WASTE_FACTORS.get(material_type, 0)

        rule = _get_labor_qty_rule(material_type)
        if rule == "with_waste":
            labor_qty = installed_qty * (1 + waste_pct)
        elif rule == "no_waste":
            labor_qty = installed_qty
        else:
            labor_qty = installed_qty * (1 + waste_pct)

        for entry in entries:
            rate = entry["cost"]
            entry_desc = entry.get("description", "").lower()
            entry_unit = (entry.get("unit") or "").upper()

            # Weld addon: use weld_rod_lf from material instead of installed qty
            if "weld" in entry_desc and entry_desc.startswith("x add"):
                weld_lf = _safe_float(mat.get("weld_rod_lf", 0))
                qty = weld_lf if weld_lf > 0 else labor_qty
            # Convert units if needed (catalog SY vs material SF)
            elif entry_unit == "SY" and material_type not in ("unit_carpet_no_pattern", "unit_carpet_pattern", "corridor_broadloom", "cpt_tile"):
                # Material is in SF, convert qty to SY
                qty = labor_qty / 9
            elif entry_unit in ("SF", "") and material_type in ("unit_carpet_no_pattern", "unit_carpet_pattern", "corridor_broadloom"):
                # Material is in SY, convert qty to SF
                qty = labor_qty * 9
            else:
                qty = labor_qty

            extended_cost = qty * rate

            results.append({
                "labor_description": f"{entry['labor_type']} - {entry['description']}",
                "qty": round(qty, 2),
                "unit": entry_unit or entry.get("unit", ""),
                "rate": rate,
                "extended_cost": round(extended_cost, 2),
                "material_id": material_id,
            })

    return results


def get_labor_catalog() -> list[dict]:
    """Return the labor catalog from the database."""
    return get_labor_catalog_entries()
